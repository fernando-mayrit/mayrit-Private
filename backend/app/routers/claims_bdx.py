"""
Claims BDX (bordereau de siniestros de un binder) — modelo replicado de Alea.

A diferencia de Risk/Premium, el Claims BDX es ACUMULATIVO: cada mes se presenta el estado
actual de TODOS los siniestros del binder y se conserva lo presentado. El "To pay this month"
se deriva = pagado acumulado actual − lo presentado el último periodo anterior (si no hay
presentación previa → 0, todo a "Previously Paid"). Presentar un mes lo BLOQUEA (BdxBloqueo
tipo='claims'); bloquear impide presentar. Export = Claims Bordereau de Lloyd's (32 columnas),
con las celdas cambiadas respecto a la última presentación en AZUL.
"""
from __future__ import annotations

import calendar
import datetime as dt
import io
import json
import re
from decimal import Decimal

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..db import get_db
from ..models.maestras import BdxBloqueo, Binder, ClaimsPresentacion, Productor, Siniestro

router = APIRouter(tags=["Claims BDX"])

# 32 columnas EXACTAS del Claims Bordereau de Lloyd's (orden del template).
HEADERS = [
    "UCR", "Coverholder Name", "Unique Market Reference (UMR)",
    "Binding authority or coverholder appointment agreement inception date",
    "Binding authority or coverholder appointment agreement expiry date",
    "Reporting Period (End Date)", "Lloyd's Risk Code", "Original Currency",
    "Certificate Reference", "Claim Reference / Number", "Insured Full Name or Company Name",
    "Insured Country", "Risk Inception Date", "Risk Expiry Date", "Location of loss Country",
    "Loss Description", "Date Claim First Advised/Date Claim Made", "Claim Status",
    "Refer to Underwriters", "Denial (Y/N)", "Claimant Name", "Amount Claimed",
    "Paid this month - Indemnity", "Paid this month - Fees",
    "Previously Paid - Indemnity", "Previously Paid - Fees",
    "Reserve - Indemnity", "Reserve - Fees",
    "Total Incurred - Indemnity", "Total Incurred - Fees",
    "Date Claim Opened", "Date Closed",
]
H_FECHA = {
    "Binding authority or coverholder appointment agreement inception date",
    "Binding authority or coverholder appointment agreement expiry date",
    "Reporting Period (End Date)", "Risk Inception Date", "Risk Expiry Date",
    "Date Claim First Advised/Date Claim Made", "Date Claim Opened", "Date Closed",
}
H_NUM = {
    "Amount Claimed", "Paid this month - Indemnity", "Paid this month - Fees",
    "Previously Paid - Indemnity", "Previously Paid - Fees", "Reserve - Indemnity",
    "Reserve - Fees", "Total Incurred - Indemnity", "Total Incurred - Fees",
}


def _f(x) -> float:
    return float(x) if x is not None else 0.0


def _fin_mes(periodo: str) -> dt.date:
    y, m = (int(x) for x in periodo.split("-"))
    return dt.date(y, m, calendar.monthrange(y, m)[1])


def _ord(periodo: str) -> int:
    y, m = (int(x) for x in periodo.split("-"))
    return y * 100 + m


def _yn(v) -> str:
    return "Yes" if str(v or "").strip().lower() in ("1", "yes", "y", "sí", "si", "true") else "No"


def _binder_o_404(binder_id: int, db: Session) -> Binder:
    b = db.get(Binder, binder_id)
    if b is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    return b


def _bloqueado(db: Session, binder_id: int, periodo: str) -> bool:
    return db.scalar(
        select(BdxBloqueo).where(BdxBloqueo.binder_id == binder_id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == periodo)
    ) is not None


def _construir(db: Session, b: Binder, periodo: str) -> tuple[list[dict], list[dict]]:
    """Filas (32 col) + meta (por siniestro) del Claims BDX del binder para ese periodo."""
    siniestros = db.scalars(
        select(Siniestro).where(Siniestro.binder_id == b.id).order_by(Siniestro.certificate, Siniestro.reference, Siniestro.id)
    ).all()
    coverholder = db.get(Productor, b.productor_id).nombre if b.productor_id else None
    fin = _fin_mes(periodo)
    po = _ord(periodo)

    # Última presentación ANTERIOR por siniestro → base de "Previously Paid".
    prev_rows = db.scalars(
        select(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo_ord < po)
    ).all()
    prev: dict[int, tuple[float, float]] = {}
    for p in sorted(prev_rows, key=lambda x: x.periodo_ord):
        prev[p.siniestro_id] = (_f(p.paid_indemnity_acum), _f(p.paid_fees_acum))  # se queda con la más reciente

    filas, meta = [], []
    for s in siniestros:
        paid_i, paid_f = _f(s.paid_indemnity), _f(s.paid_fees)
        res_i, res_f = _f(s.reserves_indemnity), _f(s.reserves_fees)
        prev_i, prev_f = prev.get(s.id, (paid_i, paid_f))  # sin previa: to_pay=0, todo previamente pagado
        fila = {
            "UCR": s.ucr,
            "Coverholder Name": coverholder,
            "Unique Market Reference (UMR)": b.umr,
            "Binding authority or coverholder appointment agreement inception date": b.fecha_efecto,
            "Binding authority or coverholder appointment agreement expiry date": b.fecha_vencimiento,
            "Reporting Period (End Date)": fin,
            "Lloyd's Risk Code": s.risk_code,
            "Original Currency": s.currency or "EUR",
            "Certificate Reference": s.certificate,
            "Claim Reference / Number": s.reference,
            "Insured Full Name or Company Name": s.insured,
            "Insured Country": "Spain",
            "Risk Inception Date": s.risk_inception,
            "Risk Expiry Date": s.risk_expiry,
            "Location of loss Country": "Spain",
            "Loss Description": s.description,
            "Date Claim First Advised/Date Claim Made": s.claim_first_advised,
            "Claim Status": s.status,
            "Refer to Underwriters": _yn(s.refer),
            "Denial (Y/N)": _yn(s.denial),
            "Claimant Name": s.claimant,
            "Amount Claimed": _f(s.amount_claimed),
            "Paid this month - Indemnity": paid_i - prev_i,
            "Paid this month - Fees": paid_f - prev_f,
            "Previously Paid - Indemnity": prev_i,
            "Previously Paid - Fees": prev_f,
            "Reserve - Indemnity": res_i,
            "Reserve - Fees": res_f,
            # Total Incurred = LO QUE SE VE EN PANTALLA: pagado + reservas (igual que la app/front). El
            # campo `total_indemnity`/`total_fees` de la BD (import de SharePoint) NO se usa: puede estar
            # desfasado y no se muestra en ningún sitio de la app.
            "Total Incurred - Indemnity": paid_i + res_i,
            "Total Incurred - Fees": paid_f + res_f,
            "Date Claim Opened": s.date_opened,
            "Date Closed": s.date_closed,
        }
        filas.append({h: fila.get(h) for h in HEADERS})
        meta.append({
            "siniestro_id": s.id, "paid_indemnity_acum": paid_i, "paid_fees_acum": paid_f,
            "to_pay_indemnity": paid_i - prev_i, "to_pay_fees": paid_f - prev_f,
            "reserves_indemnity": res_i, "reserves_fees": res_f, "status": s.status,
        })
    return filas, meta


def _norm(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float, Decimal)):
        return f"{float(v):.2f}"
    return str(v)


def _diff(filas: list[dict], base: dict[str, dict]) -> list[set]:
    """Conjunto de columnas cambiadas por fila respecto a la baseline (por clave UCR/Certificate+Ref)."""
    out = []
    for f in filas:
        clave = (f.get("Certificate Reference") or "", f.get("Claim Reference / Number") or "")
        prev = base.get(clave)
        if prev is None:
            out.append({h for h in HEADERS if _norm(f.get(h))})  # siniestro nuevo: marca lo no vacío
            continue
        out.append({h for h in HEADERS if h != "Reporting Period (End Date)" and _norm(f.get(h)) != _norm(prev.get(h))})
    return out


def _baseline(db: Session, binder_id: int, antes_de_ord: int | None = None) -> dict[str, dict]:
    """Filas (fila_json) de la última presentación del binder (la más reciente, o la más reciente
    ESTRICTAMENTE anterior a `antes_de_ord`), indexadas por (certificate, reference)."""
    stmt = select(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == binder_id)
    if antes_de_ord is not None:
        stmt = stmt.where(ClaimsPresentacion.periodo_ord < antes_de_ord)
    rows = db.scalars(stmt).all()
    if not rows:
        return {}
    ult = max(r.periodo_ord for r in rows)
    base: dict[str, dict] = {}
    for r in rows:
        if r.periodo_ord != ult or not r.fila_json:
            continue
        fila = json.loads(r.fila_json)
        base[(fila.get("Certificate Reference") or "", fila.get("Claim Reference / Number") or "")] = fila
    return base


# ── Excel ──
HEAD_FONT = Font(name="Calibri", size=9, bold=True)
HEAD_FILL = PatternFill("solid", fgColor="D9D9D9")
BODY_FONT = Font(name="Calibri", size=9)
AZUL = Font(name="Calibri", size=9, bold=True, color="0070C0")


def _excel(filas: list[dict], cambios: list[set]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claims BDX"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 70
    for i, f in enumerate(filas):
        ws.append([f.get(h) for h in HEADERS])
        fila_xl = ws[i + 2]
        camb = cambios[i] if i < len(cambios) else set()
        for j, h in enumerate(HEADERS):
            c = fila_xl[j]
            c.font = AZUL if h in camb else BODY_FONT
            if h in H_FECHA:
                c.number_format = "dd/mm/yyyy"
            elif h in H_NUM:
                c.number_format = "#,##0.00"
    for j, h in enumerate(HEADERS, start=1):
        ancho = max([len(h)] + [len(_norm(f.get(h))) for f in filas]) if filas else len(h)
        ws.column_dimensions[get_column_letter(j)].width = min(max(ancho + 1, 10), 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(ws.max_row, 1)}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────── Endpoints ───────────────────────────────
@router.get("/binders/{binder_id}/claims-bdx")
def vista(binder_id: int, periodo: str | None = None, db: Session = Depends(get_db)):
    """Vista en vivo del Claims BDX. `meses` = solo periodos YA presentados (los que existen);
    para presentar uno nuevo se elige el mes libremente en el frontend. Por defecto abre en el
    último presentado (o el mes actual si aún no hay ninguno)."""
    b = _binder_o_404(binder_id, db)
    presentadas = sorted(
        {p.periodo for p in db.scalars(select(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id)).all()},
        reverse=True,
    )
    periodo = periodo or (presentadas[0] if presentadas else dt.date.today().strftime("%Y-%m"))
    filas, _ = _construir(db, b, periodo)
    # Meses candidatos a PRESENTAR: del efecto del binder a hoy, menos los ya presentados.
    pend, ini, hoy, ya = [], (b.fecha_efecto or dt.date.today()), dt.date.today(), set(presentadas)
    y, m = ini.year, ini.month
    while (y, m) <= (hoy.year, hoy.month):
        mm = f"{y:04d}-{m:02d}"
        if mm not in ya:
            pend.append(mm)
        m += 1
        if m > 12:
            y, m = y + 1, 1
    return {
        "periodo": periodo,
        "meses": presentadas,
        "meses_pendientes": pend,
        "presentado": periodo in presentadas,
        "bloqueado": _bloqueado(db, b.id, periodo),
        "headers": HEADERS,
        "filas": filas,
    }


@router.get("/binders/{binder_id}/claims-bdx/periodos")
def periodos_presentados(binder_id: int, db: Session = Depends(get_db)):
    rows = db.scalars(
        select(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == binder_id)
    ).all()
    agg: dict[str, dict] = {}
    for r in rows:
        g = agg.setdefault(r.periodo, {"periodo": r.periodo, "n": 0, "fecha": r.fecha_presentacion})
        # Solo cuentan los siniestros reales: una fila sin siniestro es el placeholder de un mes
        # presentado en blanco (NIL report), que debe figurar con n = 0.
        if r.siniestro_id is not None:
            g["n"] += 1
        if r.fecha_presentacion and (not g["fecha"] or r.fecha_presentacion > g["fecha"]):
            g["fecha"] = r.fecha_presentacion
    return sorted(agg.values(), key=lambda x: x["periodo"], reverse=True)


class PresentarPayload(BaseModel):
    periodo: str
    usuario: str | None = None


@router.post("/binders/{binder_id}/claims-bdx/presentar")
def presentar(binder_id: int, payload: PresentarPayload, db: Session = Depends(get_db)):
    """Congela el snapshot del periodo y lo BLOQUEA. Reemplaza si ya existía."""
    b = _binder_o_404(binder_id, db)
    if (b.estado or "") == "Cerrado":
        raise HTTPException(status_code=409, detail="El binder está «Cerrado»: no se pueden cargar más claims.")
    periodo = payload.periodo
    if _bloqueado(db, b.id, periodo):
        raise HTTPException(status_code=409, detail=f"El Claims BDX de {periodo} está bloqueado. Desbloquéalo para volver a presentarlo.")
    filas, meta = _construir(db, b, periodo)
    db.execute(delete(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == periodo))
    hoy = dt.date.today()
    po = _ord(periodo)
    for f, m in zip(filas, meta):
        db.add(ClaimsPresentacion(
            binder_id=b.id, periodo=periodo, periodo_ord=po, siniestro_id=m["siniestro_id"],
            paid_indemnity_acum=m["paid_indemnity_acum"], paid_fees_acum=m["paid_fees_acum"],
            to_pay_indemnity=m["to_pay_indemnity"], to_pay_fees=m["to_pay_fees"],
            reserves_indemnity=m["reserves_indemnity"], reserves_fees=m["reserves_fees"],
            status=m["status"], fila_json=json.dumps(f, ensure_ascii=False, default=str),
            fecha_presentacion=hoy, usuario=payload.usuario,
        ))
    # Presentar = bloquear ese mes (BdxBloqueo claims).
    if not _bloqueado(db, b.id, periodo):
        db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=periodo))
    db.commit()
    return {"periodo": periodo, "presentados": len(filas)}


@router.get("/binders/{binder_id}/claims-bdx/excel")
def excel(binder_id: int, periodo: str, modo: str = "vivo", db: Session = Depends(get_db)):
    """Descarga el Claims BDX (32 col) del periodo. modo=vivo (estado actual) o presentado (snapshot)."""
    b = _binder_o_404(binder_id, db)
    if modo == "presentado":
        rows = db.scalars(
            select(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == periodo).order_by(ClaimsPresentacion.id)
        ).all()
        filas = [json.loads(r.fila_json) for r in rows if r.fila_json]
        # Diff respecto a la presentación ANTERIOR a este periodo.
        cambios = _diff(filas, _baseline(db, b.id, _ord(periodo)))
    else:
        filas, _ = _construir(db, b, periodo)
        cambios = _diff(filas, _baseline(db, b.id, _ord(periodo)))
    contenido = _excel(filas, cambios)
    nombre = f"Claims BDX {b.umr or binder_id} {periodo}.xlsx"
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ─────────────── Comparar un Claims BDX subido con los siniestros de la app ───────────────
BLUE_FILL = PatternFill("solid", fgColor="9BC2E6")   # relleno azul para las celdas que difieren
EXTRA_COL = "Comparación"


def _hk(v) -> str:
    """Normaliza un texto de cabecera (minúsculas, sin signos) para emparejar columnas de forma
    tolerante (paréntesis, barras, etc.)."""
    s = "" if v is None else str(v).strip().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _hv(v) -> str:
    """Normaliza un VALOR de texto para comparar (trim + colapsa espacios + minúsculas)."""
    if v is None:
        return ""
    return " ".join(str(v).strip().lower().split())


def _to_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def _to_num(v) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float, Decimal)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return 0.0


def _igual(h: str, a, b) -> bool:
    if h in H_FECHA:
        return _to_date(a) == _to_date(b)
    if h in H_NUM:
        return round(_to_num(a), 2) == round(_to_num(b), 2)
    return _hv(a) == _hv(b)


def _clave_fila(d: dict) -> tuple:
    return (_hv(d.get("Certificate Reference")), _hv(d.get("Claim Reference / Number")))


def _leer_bdx_subido(contenido: bytes) -> list[dict]:
    """Lee un Claims BDX (plantilla Lloyd's) subido y devuelve sus filas mapeadas a HEADERS.
    Detecta la fila de cabecera y empareja columnas de forma tolerante a signos."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    norm_h = {_hk(h): h for h in HEADERS}
    hdr_idx, colmap = None, {}
    for i, row in enumerate(filas[:25]):
        cells = [_hk(c) for c in row]
        if sum(1 for c in cells if c in norm_h) >= 5:
            hdr_idx = i
            for j, c in enumerate(cells):
                if c in norm_h and norm_h[c] not in colmap:
                    colmap[norm_h[c]] = j
            break
    if hdr_idx is None:
        raise HTTPException(status_code=400,
                            detail="No se reconocen las columnas del Claims BDX en el fichero (¿es la plantilla de Lloyd's?).")
    out = []
    for row in filas[hdr_idx + 1:]:
        if row is None or all(c in (None, "") for c in row):
            continue
        d = {h: (row[colmap[h]] if (h in colmap and colmap[h] < len(row)) else None) for h in HEADERS}
        if not (d.get("Certificate Reference") or d.get("Claim Reference / Number") or d.get("UCR")):
            continue
        out.append(d)
    return out


def _periodo_de_filas(file_rows: list[dict]) -> str:
    fechas = [f for f in (_to_date(d.get("Reporting Period (End Date)")) for d in file_rows) if f]
    f = max(fechas) if fechas else dt.date.today()
    return f"{f.year:04d}-{f.month:02d}"


def _excel_comparacion(app_filas: list[dict], file_by_key: dict[tuple, dict]) -> bytes:
    cols = [EXTRA_COL] + HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparación Claims"
    ws.append(cols)
    for c in ws[1]:
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 70

    def _formatos(fila_xl):
        for j, h in enumerate(HEADERS):
            c = fila_xl[j + 1]
            c.font = BODY_FONT
            if h in H_FECHA:
                c.number_format = "dd/mm/yyyy"
            elif h in H_NUM:
                c.number_format = "#,##0.00"

    def _todas_azul(fila_xl, d):
        for j, h in enumerate(HEADERS):
            if d.get(h) not in (None, ""):
                fila_xl[j + 1].fill = BLUE_FILL

    # Se MANTIENE la información del BDX SUBIDO: en las filas que casan se escriben los valores del FICHERO,
    # sombreando en azul lo que difiere de la app (con comentario del valor de la app). Se itera sobre las
    # filas de la app (no se colapsan por clave) para traer siempre bien su dato en el comentario.
    usados, r = set(), 2
    for fila in app_filas:
        k = _clave_fila(fila)
        fr = file_by_key.get(k)
        if fr is not None:
            usados.add(k)
            difs = [h for h in HEADERS if not _igual(h, fr.get(h), fila.get(h))]
            estado = "Difiere" if difs else "Coincide"
            ws.append([estado] + [fr.get(h) for h in HEADERS])   # valores del FICHERO subido
            fila_xl = ws[r]
            fila_xl[0].font = BODY_FONT
            _formatos(fila_xl)
            for h in difs:
                c = fila_xl[HEADERS.index(h) + 1]
                c.fill = BLUE_FILL
                av = fila.get(h)
                c.comment = Comment(f"En la app: {av if av not in (None, '') else '(vacío)'}", "Mayrit")
        else:
            # Solo en la app (no venía en el fichero): valores de la app, todas las celdas en azul.
            ws.append(["Solo en la app"] + [fila.get(h) for h in HEADERS])
            fila_xl = ws[r]
            fila_xl[0].font = BODY_FONT
            _formatos(fila_xl)
            _todas_azul(fila_xl, fila)
        r += 1

    # Siniestros NUEVOS (en el fichero subido, no en la app): sus valores, TODAS las celdas en azul.
    for k, fr in file_by_key.items():
        if k in usados:
            continue
        ws.append(["Solo en el BDX subido"] + [fr.get(h) for h in HEADERS])
        fila_xl = ws[r]
        fila_xl[0].font = BODY_FONT
        _formatos(fila_xl)
        _todas_azul(fila_xl, fr)
        r += 1

    ws.column_dimensions["A"].width = 20
    _todas = list(file_by_key.values()) + app_filas
    for j, h in enumerate(HEADERS, start=2):
        ancho = max([len(h)] + [len(_norm(f.get(h))) for f in _todas]) if _todas else len(h)
        ws.column_dimensions[get_column_letter(j)].width = min(max(ancho + 1, 10), 45)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(ws.max_row, 1)}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/binders/{binder_id}/claims-bdx/comparar")
def comparar(binder_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Compara un Claims BDX (Excel) subido con los siniestros de la app y devuelve un Excel que MANTIENE
    los valores del fichero SUBIDO, con las celdas que difieren de la app resaltadas en azul (y un
    comentario con el valor que hay en la app)."""
    b = _binder_o_404(binder_id, db)
    contenido = file.file.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="El fichero está vacío.")
    file_rows = _leer_bdx_subido(contenido)
    if not file_rows:
        raise HTTPException(status_code=400, detail="El fichero no contiene filas de siniestros reconocibles.")
    periodo = _periodo_de_filas(file_rows)
    app_filas, _ = _construir(db, b, periodo)
    file_by_key: dict[tuple, dict] = {}
    for d in file_rows:
        file_by_key.setdefault(_clave_fila(d), d)
    contenido_xlsx = _excel_comparacion(app_filas, file_by_key)
    nombre = f"Comparacion Claims {b.umr or binder_id}.xlsx"
    return Response(
        content=contenido_xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ─────────────── SUBIR un Claims BDX: aplicar SOLO las celdas AZULES a los siniestros ───────────────
# El fichero es la plantilla de Lloyd's donde las celdas cambiadas van con FONDO AZUL (las que pinta el
# «Comparar», o las que el gestor marque a mano). «Subir» copia SOLO esas celdas a su campo del siniestro
# (empareja por Claim Reference + Certificate). Una fila que no casa con ningún siniestro = NUEVO (se crea).
# Las reservas vienen YA NETAS de los pagos (fichero bien hecho): se copian tal cual, NO se recalcula el
# Total Incurred (= pagado + reservas, se deduce solo) → no se duplica nada.

# Cabecera del BDX → campo DIRECTO del Siniestro (los importes de pago se tratan aparte).
_H2F = {
    "UCR": "ucr", "Lloyd's Risk Code": "risk_code", "Original Currency": "currency",
    "Certificate Reference": "certificate", "Claim Reference / Number": "reference",
    "Insured Full Name or Company Name": "insured", "Risk Inception Date": "risk_inception",
    "Risk Expiry Date": "risk_expiry", "Loss Description": "description",
    "Date Claim First Advised/Date Claim Made": "claim_first_advised", "Claim Status": "status",
    "Refer to Underwriters": "refer", "Denial (Y/N)": "denial", "Claimant Name": "claimant",
    "Amount Claimed": "amount_claimed", "Reserve - Indemnity": "reserves_indemnity",
    "Reserve - Fees": "reserves_fees", "Date Claim Opened": "date_opened", "Date Closed": "date_closed",
}
_F_FECHA = {"risk_inception", "risk_expiry", "claim_first_advised", "date_opened", "date_closed"}
_F_NUM = {"amount_claimed", "reserves_indemnity", "reserves_fees", "paid_indemnity", "paid_fees"}
_F_YN = {"refer", "denial"}


def _es_azul(cell) -> bool:
    """¿La celda tiene fondo AZUL? Acepta el 9BC2E6 del «Comparar» y cualquier azul parecido (azul
    claramente dominante, ni gris ni blanco). Solo rellenos sólidos con color RGB explícito."""
    f = getattr(cell, "fill", None)
    if not f or f.patternType != "solid":
        return False
    rgb = getattr(getattr(f, "fgColor", None), "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return False
    try:
        r, g, b = int(rgb[-6:-4], 16), int(rgb[-4:-2], 16), int(rgb[-2:], 16)
    except ValueError:
        return False
    return b >= 120 and b > r + 20 and b > g + 10 and not (r > 205 and g > 205 and b > 205)


def _leer_bdx_con_azules(contenido: bytes) -> list[dict]:
    """Lee el Claims BDX subido devolviendo, por fila, {cabecera: (valor, es_azul)}. Necesita los
    ESTILOS (para el azul), así que NO usa read_only. Detecta la cabecera igual que `_leer_bdx_subido`."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")
    ws = wb.active
    filas = list(ws.iter_rows())
    norm_h = {_hk(h): h for h in HEADERS}
    hdr_idx, colmap = None, {}
    for i, row in enumerate(filas[:25]):
        cells = [_hk(c.value) for c in row]
        if sum(1 for c in cells if c in norm_h) >= 5:
            hdr_idx = i
            for j, c in enumerate(cells):
                if c in norm_h and norm_h[c] not in colmap:
                    colmap[norm_h[c]] = j
            break
    if hdr_idx is None:
        raise HTTPException(status_code=400,
                            detail="No se reconocen las columnas del Claims BDX en el fichero (¿es la plantilla de Lloyd's?).")
    out = []
    for row in filas[hdr_idx + 1:]:
        if all(c.value in (None, "") for c in row):
            continue
        d = {h: ((row[j].value, _es_azul(row[j])) if j < len(row) else (None, False)) for h, j in colmap.items()}
        if not any((d.get(h) or (None,))[0] for h in ("Certificate Reference", "Claim Reference / Number", "UCR")):
            continue
        out.append(d)
    return out


def _coerce_campo(campo: str, val):
    if campo in _F_FECHA:
        return _to_date(val)
    if campo in _F_NUM:
        return Decimal(f"{_to_num(val):.2f}")
    if campo in _F_YN:
        return _yn(val)
    return str(val).strip() if val not in (None, "") else None


def _difiere(campo: str, viejo, nuevo) -> bool:
    if campo in _F_NUM:
        return round(_to_num(viejo), 2) != round(_to_num(nuevo), 2)
    if campo in _F_FECHA:
        return _to_date(viejo) != _to_date(nuevo)
    if campo in _F_YN:
        return _yn(viejo) != _yn(nuevo)
    return _hv(viejo) != _hv(nuevo)


def _val(par):
    return (par or (None, False))[0]


def _es_az(d: dict, h: str) -> bool:
    return (d.get(h) or (None, False))[1]


def _paid_nuevo(d: dict, suf: str) -> Decimal:
    """Pagado acumulado = «Previously Paid» + «Paid this month» del fichero (el Total Incurred NO se toca)."""
    return Decimal(f"{_to_num(_val(d.get(f'Previously Paid - {suf}'))) + _to_num(_val(d.get(f'Paid this month - {suf}'))):.2f}")


def _aplicar_claims(db: Session, b: Binder, file_rows: list[dict], dry_run: bool) -> dict:
    siniestros = db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all()
    por_key: dict[tuple, Siniestro] = {(_hv(s.certificate), _hv(s.reference)): s for s in siniestros}
    por_ref: dict[str, list] = {}
    for s in siniestros:
        por_ref.setdefault(_hv(s.reference), []).append(s)

    nuevos, actualizados, ambiguos = [], [], []
    sin_cambios = 0
    for d in file_rows:
        ref = _hv(_val(d.get("Claim Reference / Number")))
        cert = _hv(_val(d.get("Certificate Reference")))
        ident = {"reference": _val(d.get("Claim Reference / Number")), "certificate": _val(d.get("Certificate Reference")),
                 "insured": _val(d.get("Insured Full Name or Company Name"))}
        s = por_key.get((cert, ref))
        if s is None:
            cands = por_ref.get(ref, [])
            if len(cands) == 1:
                s = cands[0]
            elif len(cands) > 1:
                ambiguos.append(ident)          # misma referencia en varios siniestros y sin casar por certificado
                continue

        if s is None:
            # ── NUEVO siniestro: se crea con TODOS los campos del fichero (es nuevo, no importa el azul) ──
            campos = {campo: _coerce_campo(campo, _val(d.get(h))) for h, campo in _H2F.items() if _val(d.get(h)) not in (None, "")}
            campos["paid_indemnity"] = _paid_nuevo(d, "Indemnity")
            campos["paid_fees"] = _paid_nuevo(d, "Fees")
            fin = _to_date(_val(d.get("Reporting Period (End Date)")))
            if fin:
                campos.setdefault("reporting_period", fin.replace(day=1))
            nuevos.append({**ident, "campos": {k: str(v) for k, v in campos.items()}})
            if not dry_run:
                db.add(Siniestro(binder_id=b.id, **campos))
            continue

        # ── EXISTENTE: solo se copian las celdas AZULES ──
        cambios = []
        for h, campo in _H2F.items():
            if not _es_az(d, h):
                continue
            nuevo = _coerce_campo(campo, _val(d.get(h)))
            if _difiere(campo, getattr(s, campo), nuevo):
                cambios.append({"campo": campo, "de": _fmt_val(getattr(s, campo)), "a": _fmt_val(nuevo)})
                if not dry_run:
                    setattr(s, campo, nuevo)
        # Pagado: si el flujo de pago viene en azul, recalcula el acumulado (Previously + This month).
        for suf, campo in (("Indemnity", "paid_indemnity"), ("Fees", "paid_fees")):
            if _es_az(d, f"Paid this month - {suf}") or _es_az(d, f"Previously Paid - {suf}"):
                nuevo = _paid_nuevo(d, suf)
                if round(_to_num(getattr(s, campo)), 2) != round(_to_num(nuevo), 2):
                    cambios.append({"campo": campo, "de": _fmt_val(getattr(s, campo)), "a": _fmt_val(nuevo)})
                    if not dry_run:
                        setattr(s, campo, nuevo)
        if cambios:
            actualizados.append({**ident, "cambios": cambios})
        else:
            sin_cambios += 1

    if not dry_run:
        db.commit()
    return {
        "dry_run": dry_run,
        "n_filas": len(file_rows),
        "nuevos": nuevos, "actualizados": actualizados, "ambiguos": ambiguos,
        "n_nuevos": len(nuevos), "n_actualizados": len(actualizados),
        "n_campos": sum(len(a["cambios"]) for a in actualizados), "sin_cambios": sin_cambios,
    }


def _fmt_val(v) -> str:
    if v in (None, ""):
        return "(vacío)"
    if isinstance(v, dt.date):
        return v.strftime("%d/%m/%Y")
    return str(v)


@router.post("/binders/{binder_id}/claims-bdx/aplicar")
def aplicar(binder_id: int, file: UploadFile = File(...), dry_run: bool = True, db: Session = Depends(get_db)):
    """SUBE un Claims BDX y aplica SOLO las celdas azules a los siniestros del binder (crea los nuevos).
    Con `dry_run=true` (por defecto) NO escribe: devuelve el resumen para confirmar. Con `dry_run=false`
    aplica y guarda. El binder «Cerrado» no admite escrituras."""
    b = _binder_o_404(binder_id, db)
    if not dry_run and (b.estado or "") == "Cerrado":
        raise HTTPException(status_code=409, detail="El binder está «Cerrado»: no se pueden cargar más claims.")
    contenido = file.file.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="El fichero está vacío.")
    file_rows = _leer_bdx_con_azules(contenido)
    if not file_rows:
        raise HTTPException(status_code=400, detail="El fichero no contiene filas de siniestros reconocibles.")
    return _aplicar_claims(db, b, file_rows, dry_run)
