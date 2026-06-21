"""
Cierre contable mensual.

- Cada mes se envían los recibos a contabilidad: se CIERRA ese (año, mes) y los recibos
  cuya FechaContable cae en ese mes pasan a estado 'Contabilizado' (no editables).
- El Excel que se envía a contabilidad es ACUMULADO del año hasta el mes elegido,
  con 2 hojas (TRecibosEUR / TRecibosUSD), con el formato de la casa (cabecera gris,
  Calibri 11 negrita; cuerpo Calibri 9) y las columnas calculadas finales.
"""
from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..db import get_db
from ..models.maestras import CierreContable, Recibo
from ..schemas import maestras as sch

router = APIRouter(tags=["Cierre contable"])

CONTABILIZADO = "Contabilizado"
MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _f(x):
    return float(x) if x is not None else None


def _pct(x):
    return float(x) / 100 if x is not None else None  # 8.15 -> 0.0815 (formato 0.00%)


def _sub(a, b):
    return float((a or Decimal(0)) - (b or Decimal(0)))


def _es(r: Recibo, *tipos: str) -> bool:
    return (r.tipo_poliza or "") in tipos


def _rango_anio(anio: int) -> tuple[dt.date, dt.date]:
    """[1-ene-anio, 1-ene-(anio+1)). Para filtrar por año usando el índice de fecha_contable
    (un rango medio-abierto, en vez de extract(year), que impide usar el índice)."""
    return dt.date(anio, 1, 1), dt.date(anio + 1, 1, 1)


def _rango_mes(anio: int, mes: int) -> tuple[dt.date, dt.date]:
    """[1-mes, 1-(mes+1)). El fin es exclusivo; diciembre pasa al 1-ene del año siguiente."""
    ini = dt.date(anio, mes, 1)
    fin = dt.date(anio + 1, 1, 1) if mes == 12 else dt.date(anio, mes + 1, 1)
    return ini, fin


# ── Formatos numéricos del template ──
MONEY = "#,##0.00"
PCT = "0.00%"
DATE = "mm-dd-yy"
ACC = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'

# Columnas del Excel (exactamente como el template): (cabecera, getter, formato).
COLS: list[tuple[str, object, str]] = [
    ("NumeroRecibo", lambda r: r.numero, "General"),
    ("Mercado", lambda r: r.mercado, "General"),
    ("NumeroPoliza", lambda r: r.numero_poliza, "General"),
    ("Asegurado", lambda r: r.asegurado, "General"),
    ("Corredor", lambda r: r.corredor, "General"),
    ("TipoPoliza", lambda r: r.tipo_poliza, "General"),
    ("FechaEfecto", lambda r: r.fecha_efecto, DATE),
    ("FechaVencimiento", lambda r: r.fecha_vencimiento, DATE),
    ("YOA", lambda r: r.yoa, "General"),
    ("Pago", lambda r: r.pago, "General"),
    ("Moneda", lambda r: r.moneda, "General"),
    ("Recibo", lambda r: r.recibo_num, "General"),
    ("FechaEfectoRecibo", lambda r: r.fecha_efecto_recibo, DATE),
    ("FechaVctoRecibo", lambda r: r.fecha_vcto_recibo, DATE),
    ("PrimaNetaRecibo", lambda r: _f(r.prima_neta_recibo), MONEY),
    ("ImpuestosPorc", lambda r: _pct(r.impuestos_porc), PCT),
    ("ImpuestossobreTotalPorc", lambda r: _pct(r.impuestos_sobre_total_porc), PCT),
    ("ImpuestossobreReciboPorc", lambda r: _pct(r.impuestos_sobre_recibo_porc), PCT),
    ("OtrosImpuestos", lambda r: _f(r.otros_impuestos), ACC),
    ("ImpuestosRecibo", lambda r: _f(r.impuestos_recibo), MONEY),
    ("PrimaBrutaRecibo", lambda r: _f(r.prima_bruta_recibo), MONEY),
    ("DeduccionTotalPorc", lambda r: _pct(r.deduccion_total_porc), PCT),
    ("DeduccionTotal", lambda r: _f(r.deduccion_total), MONEY),
    ("Honorarios", lambda r: _f(r.honorarios), MONEY),
    ("ComisionCedidaPorc", lambda r: _pct(r.comision_cedida_porc), PCT),
    ("ComisionCedida", lambda r: _f(r.comision_cedida), MONEY),
    ("ComisionRetenidaPorc", lambda r: _pct(r.comision_retenida_porc), PCT),
    ("ComisionRetenida", lambda r: _f(r.comision_retenida), MONEY),
    ("Pagador", lambda r: r.pagador, "General"),
    ("PrimaAdeudada", lambda r: _f(r.prima_adeudada), MONEY),
    ("PrimaCobrada", lambda r: _f(r.prima_cobrada), MONEY),
    ("PrimaFechaCobro", lambda r: r.prima_fecha_cobro, DATE),
    ("PrimaPendientecobro", lambda r: _sub(r.prima_adeudada, r.prima_cobrada), MONEY),
    ("ComisionRetenidaCobrada", lambda r: _f(r.comision_retenida_cobrada), MONEY),
    ("ComisionRetenidaTraspasada", lambda r: _f(r.comision_retenida_traspasada), MONEY),
    ("ComisionFechaTraspaso", lambda r: r.comision_fecha_traspaso, DATE),
    ("ComisionPendienteCobro", lambda r: _sub(r.comision_retenida, r.comision_retenida_cobrada), MONEY),
    ("ComisionPendienteTraspaso", lambda r: _sub(r.comision_retenida_cobrada, r.comision_retenida_traspasada), MONEY),
    ("Liquidar", lambda r: _f(r.liquidar), MONEY),
    ("LiquidarCobrado", lambda r: _f(r.liquidar_cobrado), MONEY),
    ("LiquidarPendienteCobro", lambda r: _sub(r.liquidar, r.liquidar_cobrado), MONEY),
    ("LiquidarLiquidado", lambda r: _f(r.liquidar_liquidado), MONEY),
    ("LiquidarFechaLiquidacion", lambda r: r.liquidar_fecha_liquidacion, DATE),
    ("LiquidarPendienteLiquidacion", lambda r: _sub(r.liquidar_cobrado, r.liquidar_liquidado), MONEY),
    ("ComisionCedidaaPagar", lambda r: _f(r.comision_cedida_a_pagar), MONEY),
    ("ComisionCedidaPagada", lambda r: _f(r.comision_cedida_pagada), MONEY),
    ("ComisionCedidaFechaPago", lambda r: r.comision_cedida_fecha_pago, DATE),
    ("ComisionCedidaPendientePago", lambda r: _sub(r.comision_cedida_a_pagar, r.comision_cedida_pagada), MONEY),
    ("Notas", lambda r: r.notas, "General"),
    ("Cuenta", lambda r: r.cuenta, "General"),
    ("FechaContable", lambda r: r.fecha_contable, DATE),
    # Columnas contables calculadas (según TipoPoliza), como en el template:
    ("Facturación Bruta", lambda r: _f(r.deduccion_total) if _es(r, "Póliza", "Comisiones") else _f(r.comision_retenida), ACC),
    ("Facturación Neta", lambda r: _f(r.comision_retenida), ACC),
    ("Comisión Cedida", lambda r: (_f(r.comision_cedida) if _es(r, "Póliza", "Comisiones") else 0.0), ACC),
    ("Base Imponible", lambda r: (_f(r.comision_retenida) if _es(r, "Consultoría") else 0.0), ACC),
    ("IVA", lambda r: (_f(r.impuestos_recibo) if _es(r, "Consultoría") else 0.0), ACC),
]

HEAD_FONT = Font(name="Calibri", size=11, bold=True)
HEAD_FILL = PatternFill("solid", fgColor="C0C0C0")
HEAD_ALIGN = Alignment(horizontal="center")
BODY_FONT = Font(name="Calibri", size=9)


def _volcar_hoja(ws, recibos: list[Recibo]) -> None:
    ws.append([h for h, _, _ in COLS])
    for c in ws[1]:
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = HEAD_ALIGN
    for r in recibos:
        ws.append([getter(r) for _, getter, _ in COLS])
    # Formatos + fuente de cuerpo
    for fila in ws.iter_rows(min_row=2):
        for j, c in enumerate(fila):
            c.font = BODY_FONT
            c.number_format = COLS[j][2]
    # Anchos aproximados
    for j, (h, _, _) in enumerate(COLS, start=1):
        ancho = len(str(h))
        for r in recibos:
            v = COLS[j - 1][1](r)
            if v is not None:
                ancho = max(ancho, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(ancho + 1, 9), 40)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(ws.max_row, 1)}"


def _excel_acumulado(db: Session, anio: int, mes: int) -> bytes:
    """Acumulado del año hasta el mes (inclusive), por FechaContable, en 2 hojas EUR/USD."""
    ini, _ = _rango_anio(anio)
    _, fin = _rango_mes(anio, mes)  # acumulado del año hasta el final del mes elegido
    recibos = db.scalars(
        select(Recibo)
        .where(Recibo.fecha_contable >= ini, Recibo.fecha_contable < fin)
        .order_by(Recibo.numero)
    ).all()
    eur = [r for r in recibos if (r.moneda or "EUR") != "USD"]
    usd = [r for r in recibos if (r.moneda or "") == "USD"]
    wb = openpyxl.Workbook()
    ws_eur = wb.active
    ws_eur.title = "TRecibosEUR"
    _volcar_hoja(ws_eur, eur)
    _volcar_hoja(wb.create_sheet("TRecibosUSD"), usd)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────── Endpoints ───────────────────────────────
@router.get("/cierres", response_model=list[sch.CierreRead])
def listar(anio: int | None = None, db: Session = Depends(get_db)):
    stmt = select(CierreContable)
    if anio is not None:
        stmt = stmt.where(CierreContable.anio == anio)
    return db.scalars(stmt.order_by(CierreContable.anio.desc(), CierreContable.mes.desc())).all()


def _anio_cerrado(db: Session, anio: int) -> bool:
    """Año cerrado = existe un cierre con mes=0 (marca de cierre anual; bloquea reabrir meses)."""
    return db.scalar(select(CierreContable).where(CierreContable.anio == anio, CierreContable.mes == 0)) is not None


@router.get("/cierres/resumen")
def resumen(anio: int, db: Session = Depends(get_db)):
    """Por cada mes del año: nº de recibos (por FechaContable), acumulado y si está cerrado.
    Incluye `anio_cerrado` y `puede_cerrar_anio` (todos los meses con recibos están cerrados)."""
    ini, fin = _rango_anio(anio)
    filas = db.execute(
        select(func.extract("month", Recibo.fecha_contable), func.count())
        .where(Recibo.fecha_contable >= ini, Recibo.fecha_contable < fin)
        .group_by(func.extract("month", Recibo.fecha_contable))
    ).all()
    por_mes = {int(m): c for m, c in filas}
    cierres = {c.mes: c for c in db.scalars(select(CierreContable).where(CierreContable.anio == anio)).all()}
    out, acum = [], 0
    meses_con_recibos = [m for m in range(1, 13) if por_mes.get(m, 0) > 0]
    for m in range(1, 13):
        n = por_mes.get(m, 0)
        acum += n
        c = cierres.get(m)
        out.append({
            "mes": m, "nombre": MESES[m], "recibos": n, "acumulado": acum,
            "cerrado": c is not None,
            "fecha": c.fecha.isoformat() if c else None,  # fecha de envío a contabilidad
        })
    anio_cerrado = _anio_cerrado(db, anio)
    # Se puede cerrar el año si hay recibos y todos los meses con recibos están cerrados.
    puede_cerrar_anio = (
        not anio_cerrado
        and len(meses_con_recibos) > 0
        and all(m in cierres for m in meses_con_recibos)
    )
    cierre_anio = cierres.get(0)
    return {
        "anio": anio,
        "meses": out,
        "anio_cerrado": anio_cerrado,
        "puede_cerrar_anio": puede_cerrar_anio,
        "anio_fecha": cierre_anio.fecha.isoformat() if cierre_anio else None,
    }


class CierrePayload(BaseModel):
    anio: int
    mes: int
    fecha: dt.date | None = None     # fecha de envío físico a contabilidad
    usuario: str | None = None


class CierreAnioPayload(BaseModel):
    fecha: dt.date | None = None
    usuario: str | None = None


@router.post("/cierres", response_model=sch.CierreRead, status_code=201)
def cerrar(payload: CierrePayload, db: Session = Depends(get_db)):
    if not (1 <= payload.mes <= 12):
        raise HTTPException(status_code=422, detail="Mes inválido.")
    ya = db.scalar(select(CierreContable).where(CierreContable.anio == payload.anio, CierreContable.mes == payload.mes))
    if ya is not None:
        raise HTTPException(status_code=409, detail=f"{MESES[payload.mes]} {payload.anio} ya está cerrado.")
    # Marca como Contabilizado los recibos de ese mes (por FechaContable).
    ini, fin = _rango_mes(payload.anio, payload.mes)
    recibos = db.scalars(
        select(Recibo).where(Recibo.fecha_contable >= ini, Recibo.fecha_contable < fin)
    ).all()
    for r in recibos:
        r.estado = CONTABILIZADO
    cierre = CierreContable(
        anio=payload.anio, mes=payload.mes,
        fecha=payload.fecha or dt.date.today(), usuario=payload.usuario,
    )
    db.add(cierre)
    db.commit()
    db.refresh(cierre)
    return cierre


@router.post("/cierres/{anio}/cerrar-anio", response_model=sch.CierreRead, status_code=201)
def cerrar_anio(anio: int, payload: CierreAnioPayload, db: Session = Depends(get_db)):
    """Cierre ANUAL: solo si todos los meses con recibos del año están cerrados. Bloquea la
    reapertura de los meses. Se materializa como un cierre con mes=0."""
    if _anio_cerrado(db, anio):
        raise HTTPException(status_code=409, detail=f"El año {anio} ya está cerrado.")
    ini, fin = _rango_anio(anio)
    filas = db.execute(
        select(func.extract("month", Recibo.fecha_contable), func.count())
        .where(Recibo.fecha_contable >= ini, Recibo.fecha_contable < fin)
        .group_by(func.extract("month", Recibo.fecha_contable))
    ).all()
    meses_con_recibos = {int(m) for m, c in filas if c > 0}
    if not meses_con_recibos:
        raise HTTPException(status_code=409, detail=f"No hay recibos contables en {anio}.")
    cerrados = {c.mes for c in db.scalars(select(CierreContable).where(CierreContable.anio == anio)).all()}
    faltan = sorted(meses_con_recibos - cerrados)
    if faltan:
        nombres = ", ".join(MESES[m] for m in faltan)
        raise HTTPException(status_code=409, detail=f"No se puede cerrar el año: faltan meses por cerrar ({nombres}).")
    cierre = CierreContable(anio=anio, mes=0, fecha=payload.fecha or dt.date.today(), usuario=payload.usuario)
    db.add(cierre)
    db.commit()
    db.refresh(cierre)
    return cierre


@router.delete("/cierres/{anio}/anio", status_code=204)
def reabrir_anio(anio: int, db: Session = Depends(get_db)):
    """Reabre el año (quita la marca mes=0): vuelve a permitir reabrir los meses."""
    cierre = db.scalar(select(CierreContable).where(CierreContable.anio == anio, CierreContable.mes == 0))
    if cierre is None:
        raise HTTPException(status_code=404, detail=f"El año {anio} no está cerrado.")
    db.delete(cierre)
    db.commit()


@router.delete("/cierres/{anio}/{mes}", status_code=204)
def reabrir(anio: int, mes: int, db: Session = Depends(get_db)):
    if _anio_cerrado(db, anio):
        raise HTTPException(
            status_code=409,
            detail=f"El año {anio} está cerrado: reabre primero el año para poder reabrir sus meses.",
        )
    cierre = db.scalar(select(CierreContable).where(CierreContable.anio == anio, CierreContable.mes == mes))
    if cierre is None:
        raise HTTPException(status_code=404, detail=f"{MESES[mes]} {anio} no está cerrado.")
    # Reabre: los recibos contabilizados de ese mes vuelven a 'Emitido'.
    ini, fin = _rango_mes(anio, mes)
    recibos = db.scalars(
        select(Recibo).where(
            Recibo.fecha_contable >= ini, Recibo.fecha_contable < fin,
            Recibo.estado == CONTABILIZADO,
        )
    ).all()
    for r in recibos:
        r.estado = "Emitido"
    db.delete(cierre)
    db.commit()


@router.get("/cierres/{anio}/{mes}/excel")
def excel(anio: int, mes: int, db: Session = Depends(get_db)):
    """Excel acumulado del año hasta ese mes (el que se envía a contabilidad)."""
    if not (1 <= mes <= 12):
        raise HTTPException(status_code=422, detail="Mes inválido.")
    contenido = _excel_acumulado(db, anio, mes)
    nombre = f"TRecibos Total {MESES[mes]} {anio}.xlsx"
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
