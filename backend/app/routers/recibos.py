"""
Recibos: núcleo de facturación/contabilidad. Modelo basado en SharePoint 'Mayrit - TRecibos'.

En la app se **emite 1 recibo por Risk BDX** (binder + periodo 'YYYY-MM'); la comisión de Mayrit
es `comision_retenida` = Σ `brokerage_amount` de las líneas Risk de ese periodo. El cobro llega
con los Premium BDX (rara vez coinciden con el Risk BDX) → puede ser parcial. Numeración por año
natural 'AÑO-NNNN'. Los "pendientes" (cobro/liquidación) los recalcula el backend.
"""
import calendar
import datetime as dt
import io
import re
import time
import uuid
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import delete, func, select, update
from sqlalchemy.orm import Session, load_only

from ..db import get_db
from ..models.maestras import (
    Bdx,
    BdxBloqueo,
    BdxLinea,
    Binder,
    BinderSeccion,
    CierreContable,
    CuentaBancaria,
    Lpan,
    Mercado,
    Poliza,
    PremiumNota,
    Programa,
    Recibo,
    SeccionMercado,
)
from ..schemas import maestras as sch
from .. import transferencias_auto

router = APIRouter(tags=["Recibos"])

D0 = Decimal(0)
# Nº de Risk BDX al año según el intervalo del binder (para "Recibo Nº X de N").
INTERVALO_N = {"Mensual": 12, "Trimestral": 4, "Semestral": 2, "Anual": 1}


def _q2(x) -> Decimal:
    return Decimal(x).quantize(Decimal("0.01"), ROUND_HALF_UP)


def _q4(x):
    return None if x is None else Decimal(x).quantize(Decimal("0.0001"), ROUND_HALF_UP)


def _exigir_mes_abierto(db: Session, fecha: dt.date | None) -> None:
    """Bloquea crear recibos cuya FechaContable caiga en un mes contable ya cerrado."""
    if not fecha:
        return
    cerrado = db.scalar(
        select(CierreContable).where(CierreContable.anio == fecha.year, CierreContable.mes == fecha.month)
    )
    if cerrado is not None:
        raise HTTPException(
            status_code=409,
            detail=f"El mes contable {fecha.month:02d}/{fecha.year} está cerrado: no se pueden generar recibos en ese periodo.",
        )


def _max_numero(db: Session, anio: int) -> int:
    """Mayor correlativo NNNN usado en 'AÑO-NNNN' para ese año (0 si no hay)."""
    numeros = db.scalars(select(Recibo.numero).where(Recibo.anio == anio)).all()
    maximo = 0
    for n in numeros:
        try:
            maximo = max(maximo, int(str(n).split("-")[-1]))
        except (ValueError, IndexError):
            pass
    return maximo


def _siguiente_numero(db: Session, anio: int) -> str:
    """'AÑO-NNNN' correlativo por año natural (último + 1)."""
    return f"{anio}-{_max_numero(db, anio) + 1:04d}"


# ── Numeración automática de pólizas OM: B1634OM + AA (año efecto) + NNNNNNNN (correlativo global) ──
POLIZA_PREFIX = "B1634OM"
POLIZA_SEQ_INICIO = 43  # primera póliza emitida con numeración automática


def _siguiente_numero_poliza(db: Session, anio: int) -> str:
    """B1634OM + AA + correlativo de 8 dígitos. El correlativo es GLOBAL (no reinicia por año)
    y arranca en 43; AA son los 2 últimos dígitos del año de la fecha de efecto."""
    yy = f"{anio % 100:02d}"
    rx = re.compile(rf"^{POLIZA_PREFIX}\d{{2}}(\d{{8}})$")
    maxseq = POLIZA_SEQ_INICIO - 1
    for num in db.scalars(select(Poliza.numero_poliza).where(Poliza.numero_poliza.isnot(None))):
        m = rx.match(num or "")
        if m:
            maxseq = max(maxseq, int(m.group(1)))
    return f"{POLIZA_PREFIX}{yy}{maxseq + 1:08d}"


@router.get("/polizas/siguiente-numero")
def siguiente_numero_poliza(anio: int, db: Session = Depends(get_db)):
    """Devuelve el próximo nº de póliza automático para un año (de la fecha de efecto)."""
    return {"numero_poliza": _siguiente_numero_poliza(db, anio)}


def _rango_mes(periodo: str) -> tuple[dt.date, dt.date]:
    """'YYYY-MM' → (primer día del mes, primer día del mes siguiente)."""
    try:
        y, m = (int(x) for x in periodo.split("-"))
        ini = dt.date(y, m, 1)
    except (ValueError, TypeError):
        raise HTTPException(status_code=422, detail=f"Periodo inválido: {periodo!r} (use 'YYYY-MM').")
    fin = dt.date(y + 1, 1, 1) if m == 12 else dt.date(y, m + 1, 1)
    return ini, fin


def _lineas_risk_periodo(db: Session, binder_id: int, periodo: str):
    """Líneas del BDX (Risk) del binder cuyo reporting_period_start cae en el mes `periodo`."""
    ini, fin = _rango_mes(periodo)
    return db.scalars(
        select(BdxLinea)
        .join(Bdx, BdxLinea.bdx_id == Bdx.id)
        .where(
            Bdx.binder_id == binder_id,
            BdxLinea.reporting_period_start >= ini,
            BdxLinea.reporting_period_start < fin,
        )
    ).all()


# A los sindicatos de Lloyd's se les liquida a través de esta entidad única (Lloyd's Bruselas).
LLOYDS_COMPANY = "Lloyds Insurance Company"


def _mercado_recibo_binder(db: Session, binder_id: int) -> tuple[str | None, str | None]:
    """(mercado, nombre_mercado) del recibo de un binder:
      - `nombre_mercado` = nombres de los mercados del binder (sindicatos/compañías), sin repetir.
      - `mercado` = la entidad por la que se AGRUPA/liquida: para sindicatos de Lloyd's
        (tipo_mercado='Lloyds') es SIEMPRE 'Lloyds Insurance Company' (el settlement de Lloyd's);
        para compañías es el propio nombre. Así los recibos de Lloyd's agrupan por Lloyd's."""
    rows = db.execute(
        select(Mercado.nombre, Mercado.tipo_mercado)
        .join(SeccionMercado, SeccionMercado.mercado_id == Mercado.id)
        .join(BinderSeccion, BinderSeccion.id == SeccionMercado.seccion_id)
        .where(BinderSeccion.binder_id == binder_id)
        .distinct()
    ).all()
    nombres = sorted({n for n, _ in rows if n})
    if not nombres:
        return None, None
    nombre_mercado = ", ".join(nombres)
    tipos = {t for _, t in rows}
    mercado = LLOYDS_COMPANY if tipos and tipos <= {"Lloyds"} else nombre_mercado
    return mercado, nombre_mercado


def _mercado_pago(db: Session, nombre: str | None) -> str | None:
    """Entidad por la que se agrupa/liquida un mercado (para el campo `mercado` del recibo):
    'Lloyds Insurance Company' si es sindicato de Lloyd's (tipo_mercado='Lloyds'); si no, el nombre."""
    if not nombre:
        return nombre
    tipo = db.scalar(select(Mercado.tipo_mercado).where(Mercado.nombre == nombre))
    return LLOYDS_COMPANY if tipo == "Lloyds" else nombre


def _yoa_int(binder: Binder) -> int | None:
    return int(binder.yoa) if binder.yoa and str(binder.yoa).isdigit() else None


def _cuenta_binder(db: Session, binder: Binder) -> str | None:
    if not binder.cuenta_bancaria_id:
        return None
    c = db.get(CuentaBancaria, binder.cuenta_bancaria_id)
    return c.nombre if c else None


def _ramos_binder(db: Session, binder_id: int) -> str | None:
    ramos = db.scalars(
        select(BinderSeccion.ramo).where(BinderSeccion.binder_id == binder_id).distinct()
    ).all()
    ramos = [r for r in ramos if r]
    return ", ".join(sorted(set(ramos))) if ramos else None


def _pos_bdx_anual(binder: Binder, periodo: str) -> tuple[int | None, str | None]:
    """'Recibo Nº X de N' = posición de este Risk BDX en el año según el intervalo del binder."""
    mes = int(periodo.split("-")[1])
    n = INTERVALO_N.get(binder.risk_bdx_intervalo or "")
    if n == 12:
        x = mes
    elif n == 4:
        x = (mes - 1) // 3 + 1
    elif n == 2:
        x = (mes - 1) // 6 + 1
    elif n == 1:
        x = 1
    else:
        x, n = mes, None  # intervalo desconocido: usa el mes, total sin fijar
    return x, (str(n) if n else None)


def _recompute(r: Recibo) -> None:
    """Recalcula los 'pendientes' a partir de los importes base."""
    r.comision_pendiente_cobro = (r.comision_retenida or D0) - (r.comision_retenida_cobrada or D0)
    r.liquidar_pendiente_cobro = (r.liquidar or D0) - (r.liquidar_cobrado or D0)


def _read(db: Session, r: Recibo) -> sch.ReciboRead:
    """ReciboRead enriquecido con UMR del binder (o nº de póliza en OM) y nº de líneas enlazadas."""
    binder = db.get(Binder, r.binder_id) if r.binder_id else None
    poliza = db.get(Poliza, r.poliza_id) if r.poliza_id else None
    num_lineas = db.scalar(select(func.count(BdxLinea.id)).where(BdxLinea.recibo_id == r.id)) or 0
    data = sch.ReciboRead.model_validate(r)
    data.binder_umr = (binder.umr or binder.agreement_number) if binder else None
    data.poliza_numero = poliza.numero_poliza if poliza else None
    data.num_lineas = num_lineas
    return data


def _read_lote(db: Session, recibos: list[Recibo]) -> list[sch.ReciboRead]:
    """Como _read pero para una lista, cargando binders, pólizas y conteos EN LOTE (evita N+1)."""
    if not recibos:
        return []
    bids = {r.binder_id for r in recibos if r.binder_id}
    pids = {r.poliza_id for r in recibos if r.poliza_id}
    binders = {b.id: b for b in db.scalars(select(Binder).where(Binder.id.in_(bids))).all()} if bids else {}
    polizas = {p.id: p for p in db.scalars(select(Poliza).where(Poliza.id.in_(pids))).all()} if pids else {}
    rids = [r.id for r in recibos]
    counts = {
        rid: c
        for rid, c in db.execute(
            select(BdxLinea.recibo_id, func.count(BdxLinea.id))
            .where(BdxLinea.recibo_id.in_(rids))
            .group_by(BdxLinea.recibo_id)
        ).all()
    }
    out: list[sch.ReciboRead] = []
    for r in recibos:
        data = sch.ReciboRead.model_validate(r)
        b = binders.get(r.binder_id) if r.binder_id else None
        p = polizas.get(r.poliza_id) if r.poliza_id else None
        data.binder_umr = (b.umr or b.agreement_number) if b else None
        data.poliza_numero = p.numero_poliza if p else None
        data.num_lineas = counts.get(r.id, 0)
        out.append(data)
    return out


# ──────────────────────────────── Listados ──────────────────────────────────
@router.get("/recibos", response_model=list[sch.ReciboRead])
def listar(anio: int | None = None, binder_id: int | None = None, poliza_id: int | None = None, q: str | None = None, db: Session = Depends(get_db)):
    stmt = select(Recibo)
    if anio is not None:
        stmt = stmt.where(Recibo.anio == anio)
    if binder_id is not None:
        stmt = stmt.where(Recibo.binder_id == binder_id)
    if poliza_id is not None:
        stmt = stmt.where(Recibo.poliza_id == poliza_id)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            Recibo.numero.ilike(like) | Recibo.nombre_mercado.ilike(like) | Recibo.asegurado.ilike(like)
        )
    stmt = stmt.order_by(Recibo.anio.desc(), Recibo.numero.desc())
    return _read_lote(db, list(db.scalars(stmt).all()))


@router.get("/binders/{binder_id}/recibos", response_model=list[sch.ReciboRead])
def listar_de_binder(binder_id: int, db: Session = Depends(get_db)):
    filas = db.scalars(
        select(Recibo).where(Recibo.binder_id == binder_id).order_by(Recibo.periodo.desc())
    ).all()
    return _read_lote(db, list(filas))


@router.get("/recibos/{recibo_id}", response_model=sch.ReciboRead)
def obtener(recibo_id: int, db: Session = Depends(get_db)):
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    return _read(db, r)


# ───────────────────────── Generar desde un Risk BDX ─────────────────────────
def _validar(db: Session, binder: Binder, periodo: str):
    """Valida y devuelve las líneas del Risk BDX del periodo. Aborta 409/400 si procede."""
    _rango_mes(periodo)  # valida el formato
    existe = db.scalar(
        select(Recibo).where(Recibo.binder_id == binder.id, Recibo.periodo == periodo)
    )
    if existe is not None:
        raise HTTPException(
            status_code=409,
            detail=f"Ya existe el recibo {existe.numero} para este Risk BDX ({periodo}).",
        )
    lineas = _lineas_risk_periodo(db, binder.id, periodo)
    if not lineas:
        raise HTTPException(status_code=400, detail=f"No hay líneas Risk en el periodo {periodo}.")
    return lineas


def _campos_emision(db: Session, binder: Binder, periodo: str, lineas, fecha: dt.date) -> dict:
    """Auto-relleno COMPLETO del recibo a partir de las líneas del Risk BDX (our line).
    Pagador = Agencia de Suscripción → Prima Adeudada = Prima Total − Comisión Cedida."""
    def S(attr) -> Decimal:
        return sum((getattr(l, attr) or D0) for l in lineas)

    prima_neta = _q2(S("total_gwp_our_line"))       # Prima Neta Bordereau (our line, sin impuestos)
    impuestos = _q2(S("total_taxes_levies"))
    prima_bruta = _q2(prima_neta + impuestos)        # Prima Total Bordereau
    cedida = _q2(S("commission_coverholder_amount")) # comisión a la agencia (coverholder)
    retenida = _q2(S("brokerage_amount"))            # comisión de Mayrit
    honorarios = _q2(S("fees"))
    deduccion = _q2(cedida + retenida + honorarios)
    gwp100 = _q2(S("gross_written_premium"))         # GWP al 100% (para la participación)
    # Impuestos liquidados localmente (p. ej. agencias italianas): se excluyen del cobro y de
    # 'A Liquidar' (no se liquidan a través nuestro); prima_bruta/impuestos sí se reflejan.
    excl_imp = _impuestos_locales(db, binder.id)
    if binder.programa and binder.programa.reaseguro:
        # Reaseguro (p. ej. caución Iberian/Hamilton): hay una capa extra (comisión del reasegurado),
        # así que el Cobro y la liquidación salen directos de las columnas del bordereau:
        #   Cobro      = Net Premium to pay to Reinsurance Broker by Reinsured (net_premium_to_broker)
        #   A Liquidar = Final Net Premium to UW/Hamilton (final_net_premium_uw)
        # La comisión retenida de Mayrit (brokerage) = Cobro − Liquidar.
        adeudada = _q2(S("net_premium_to_broker"))
        liquidar = _q2(S("final_net_premium_uw"))
    else:
        adeudada = _q2(prima_bruta - cedida - (impuestos if excl_imp else D0))   # pagador = Agencia
        liquidar = _q2(adeudada - retenida)

    def pct(x):
        return _q4(x / prima_neta * 100) if prima_neta else None

    ini, fin = _rango_mes(periodo)
    mercado_pago, nombre_mercado = _mercado_recibo_binder(db, binder.id)
    pos, total = _pos_bdx_anual(binder, periodo)

    return dict(
        binder_id=binder.id,
        periodo=periodo,
        anio=fecha.year,
        estado="Emitido",
        # Contexto. En binder el nº de póliza es el UMR (como el histórico); referencia no se usa.
        numero_poliza=binder.umr or binder.agreement_number,
        tipo_poliza="Binder",
        referencia=None,
        nombre_mercado=nombre_mercado,
        mercado=mercado_pago,
        # Corredor = alias de la agencia (coverholder); si no tiene alias, el nombre.
        corredor=((binder.productor.alias or binder.productor.nombre) if binder.productor else None),
        ramo=_ramos_binder(db, binder.id),
        produccion=None,
        fecha_efecto=binder.fecha_efecto,
        fecha_vencimiento=binder.fecha_vencimiento,
        yoa=_yoa_int(binder),
        # El recibo de binder es de un solo pago por periodo (no fraccionado).
        pago="Único",
        moneda=binder.moneda or "EUR",
        prima_neta_poliza=prima_neta,
        participacion=(_q4(prima_neta / gwp100 * 100) if gwp100 else None),
        recibo_num=pos,
        recibos_totales=total,
        # Importe del recibo + impuestos
        fecha_efecto_recibo=ini,
        fecha_vcto_recibo=fin - dt.timedelta(days=1),
        prima_neta_recibo=prima_neta,
        impuestos_porc=pct(impuestos),
        impuestos_recibo=impuestos,
        prima_bruta_recibo=prima_bruta,
        deduccion_total_porc=pct(deduccion),
        deduccion_total=deduccion,
        honorarios=honorarios,
        # Comisiones
        comision_cedida_porc=pct(cedida),
        comision_cedida=cedida,
        comision_retenida_porc=pct(retenida),
        comision_retenida=retenida,
        pagador="Agencia de Suscripción",
        # Cobro (a 0 al emitir; llega con los Premium BDX)
        prima_adeudada=adeudada,
        # Liquidación a la Cía
        liquidar=liquidar,
        # Contable
        cuenta=_cuenta_binder(db, binder),
        fecha_contable=fecha,
    )


@router.get("/binders/{binder_id}/recibos/preview", response_model=sch.ReciboPreview)
def preview(binder_id: int, periodo: str, db: Session = Depends(get_db)):
    """Calcula el recibo SIN guardarlo, para precumplimentar el formulario de emisión."""
    binder = db.get(Binder, binder_id)
    if binder is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    lineas = _validar(db, binder, periodo)
    fecha = _rango_mes(periodo)[0]   # fecha contable por defecto = primer día del mes del periodo
    campos = _campos_emision(db, binder, periodo, lineas, fecha)
    return sch.ReciboPreview(
        numero=_siguiente_numero(db, fecha.year),
        binder_umr=binder.umr or binder.agreement_number,
        num_lineas=len(lineas),
        **campos,
    )


@router.post("/binders/{binder_id}/recibos/generar", response_model=sch.ReciboRead, status_code=201)
def generar(binder_id: int, payload: sch.ReciboGenerar, db: Session = Depends(get_db)):
    binder = db.get(Binder, binder_id)
    if binder is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")

    periodo = payload.periodo
    lineas = _validar(db, binder, periodo)
    overrides = payload.model_dump(exclude_unset=True, exclude={"periodo"})
    # Fecha contable por defecto = primer día del mes del periodo (se contabiliza en SU mes, no en el
    # día en que se genera). Editable en el formulario (override); si ese mes ya estuviera cerrado,
    # hay que fijar a mano una fecha de un mes abierto.
    fecha = overrides.get("fecha_contable") or _rango_mes(periodo)[0]
    _exigir_mes_abierto(db, fecha)  # no contabilizar en un mes cerrado

    campos = _campos_emision(db, binder, periodo, lineas, fecha)
    campos.update(overrides)  # lo editado en el formulario prevalece
    recibo = Recibo(numero=_siguiente_numero(db, campos["anio"]), **campos)
    _recompute(recibo)
    db.add(recibo)
    db.flush()                  # asigna recibo.id

    # Enlaza las líneas del periodo con el recibo (y guarda el nº en texto).
    for l in lineas:
        l.recibo_id = recibo.id
        l.recibo = recibo.numero

    db.commit()
    db.refresh(recibo)
    return _read(db, recibo)


# ═══════════════════ Emisión de Póliza OM (póliza + sus recibos) ═══════════════════
# Reglas (acordadas con negocio):
#  - Pago Único/Dos/Tres/Cuatro Pagos → 1..4 plazos; la prima de participación se reparte
#    a partes iguales (el último plazo absorbe el redondeo).
#  - Comisión por recibo = cedida% (al corredor) + retenida% (Mayrit), independientes.
#  - Fechas de plazos: 1º en fecha de efecto y los siguientes cada 12/N meses (editables).
#  - Importes por recibo: prima_bruta = prima + impuestos + recargos;
#    adeudada = prima_bruta − cedida; liquidar = adeudada − retenida.
PAGO_LABEL = {1: "Único", 2: "Semestral", 3: "Cuatrimestral", 4: "Trimestral"}


def _sumar_meses(d: dt.date, meses: int) -> dt.date:
    m = d.month - 1 + meses
    y = d.year + m // 12
    m = m % 12 + 1
    return dt.date(y, m, min(d.day, calendar.monthrange(y, m)[1]))


class _EmisionParams(BaseModel):
    n_plazos: int = 1                                   # 1..4
    comision_cedida_porc: Decimal | None = None         # al corredor
    comision_retenida_porc: Decimal | None = None       # de Mayrit
    plazos_fechas: list[dt.date] | None = None          # override opcional de fechas


class PolizaEmitir(sch.PolizaCreate, _EmisionParams):
    pass


class EmisionLinea(BaseModel):
    recibo_num: int
    recibos_totales: int
    fecha_efecto_recibo: dt.date | None = None
    fecha_vcto_recibo: dt.date | None = None
    prima_neta_recibo: Decimal
    impuestos_porc: Decimal | None = None
    impuestos_recibo: Decimal
    recargos: Decimal
    prima_bruta_recibo: Decimal
    comision_cedida_porc: Decimal | None = None
    comision_cedida: Decimal
    comision_retenida_porc: Decimal | None = None
    comision_retenida: Decimal
    prima_adeudada: Decimal
    liquidar: Decimal


class EmisionPreview(BaseModel):
    pago: str
    prima_participacion: Decimal
    impuestos: Decimal
    prima_total: Decimal
    comision_total: Decimal
    lineas: list[EmisionLinea]


def _emision_lineas(p: PolizaEmitir) -> EmisionPreview:
    n = max(1, min(4, p.n_plazos or 1))
    cap = p.capacidad if p.capacidad is not None else Decimal(1)
    base_total = _q2((p.prima_neta or D0) * cap)        # prima de participación (our line)
    recargos_total = _q2(p.recargos or D0)
    imp_pct, ced_pct, ret_pct = p.impuestos_porc, p.comision_cedida_porc, p.comision_retenida_porc

    # Quién paga: "Tomador" → adeudada = bruta (100%); "Corredor" → adeudada = bruta − cedida.
    paga_tomador = (p.pagador or "Corredor") == "Tomador"
    cuota = _q2(base_total / n)
    rec_cuota = _q2(recargos_total / n)
    step = 12 // n
    lineas: list[EmisionLinea] = []
    for i in range(n):
        prima_i = cuota if i < n - 1 else _q2(base_total - cuota * (n - 1))
        rec_i = rec_cuota if i < n - 1 else _q2(recargos_total - rec_cuota * (n - 1))
        imp_i = _q2(prima_i * imp_pct / 100) if imp_pct else D0
        bruta_i = _q2(prima_i + imp_i + rec_i)
        ced_i = _q2(prima_i * ced_pct / 100) if ced_pct else D0
        ret_i = _q2(prima_i * ret_pct / 100) if ret_pct else D0
        adeudada = bruta_i if paga_tomador else _q2(bruta_i - ced_i)
        # fechas
        def fecha_plazo(idx):
            if p.plazos_fechas and idx < len(p.plazos_fechas):
                return p.plazos_fechas[idx]
            return _sumar_meses(p.fecha_efecto, idx * step) if p.fecha_efecto else None
        fe = fecha_plazo(i)
        if i < n - 1:
            sig = fecha_plazo(i + 1)
            fv = (sig - dt.timedelta(days=1)) if sig else None
        else:
            fv = p.fecha_vencimiento
        lineas.append(EmisionLinea(
            recibo_num=i + 1, recibos_totales=n,
            fecha_efecto_recibo=fe, fecha_vcto_recibo=fv,
            prima_neta_recibo=prima_i, impuestos_porc=imp_pct, impuestos_recibo=imp_i,
            recargos=rec_i, prima_bruta_recibo=bruta_i,
            comision_cedida_porc=ced_pct, comision_cedida=ced_i,
            comision_retenida_porc=ret_pct, comision_retenida=ret_i,
            prima_adeudada=adeudada, liquidar=_q2(bruta_i - ced_i - ret_i),
        ))
    imp_total = _q2(sum((l.impuestos_recibo for l in lineas), D0))
    com_total = _q2(sum((l.comision_cedida + l.comision_retenida for l in lineas), D0))
    return EmisionPreview(
        pago=PAGO_LABEL[n], prima_participacion=base_total, impuestos=imp_total,
        prima_total=_q2(base_total + imp_total + recargos_total), comision_total=com_total, lineas=lineas,
    )


@router.post("/polizas/emitir/preview", response_model=EmisionPreview)
def emitir_preview(payload: PolizaEmitir):
    """Calcula la póliza y sus recibos SIN guardar, para precumplimentar el formulario."""
    return _emision_lineas(payload)


def _plazos_de_pago(pago: str | None) -> int:
    """Pago → nº de plazos al año (Único=1, Semestral=2, Trimestral=4)."""
    return {"Único": 1, "Semestral": 2, "Trimestral": 4}.get((pago or "").strip(), 1)


def _mercado_nombre(db: Session, valor: str | None) -> str | None:
    """Resuelve el mercado a su NOMBRE canónico: si `valor` coincide con un alias (p. ej. 'LSM'),
    devuelve el nombre completo ('Liberty Specialty Markets'); si ya es un nombre —o no se reconoce—
    lo deja igual. Así el recibo guarda el nombre del mercado aunque la póliza tenga el alias."""
    v = (valor or "").strip()
    if not v:
        return valor
    # Prioridad al match por nombre; si no, por alias.
    nombre = db.scalar(select(Mercado.nombre).where(Mercado.nombre == v))
    if nombre:
        return nombre
    return db.scalar(select(Mercado.nombre).where(Mercado.alias == v)) or valor


def _generar_recibos(db: Session, poliza: Poliza, n: int) -> None:
    """Genera los recibos de una póliza YA guardada: n plazos × compañías (coaseguro).
    Prima de cada compañía = prima_neta × su % sobre el total; repartida entre los plazos.
    Comisión por recibo repartida en cedida (corredor) y retenida (Mayrit)."""
    n = max(1, min(4, n or 1))
    step = 12 // n
    prima_neta = poliza.prima_neta or D0
    cap = poliza.capacidad if poliza.capacidad is not None else Decimal(1)   # fracción (0..1)
    imp_pct = poliza.impuestos_porc or D0
    com_pct = poliza.comision_porc or D0
    ced_pct = poliza.comision_cedida_porc or D0
    ret_pct = com_pct - ced_pct
    pagador = poliza.pagador or "Corredor"
    paga_tomador = pagador == "Tomador"

    if poliza.coaseguro and poliza.coaseguro_lineas:
        companias = [
            (str(l.get("mercado") or poliza.mercado or ""), Decimal(str(l.get("participacion") or 0)) / 100)
            for l in poliza.coaseguro_lineas
        ]
    else:
        companias = [(poliza.mercado or "", cap)]

    contadores: dict[int, int] = {}

    def numero(anio: int) -> str:
        if anio not in contadores:
            contadores[anio] = _max_numero(db, anio)
        contadores[anio] += 1
        return f"{anio}-{contadores[anio]:04d}"

    def fecha_plazo(idx: int) -> dt.date:
        return _sumar_meses(poliza.fecha_efecto, idx * step)

    # Ningún plazo puede caer en un mes contable cerrado.
    for i in range(n):
        _exigir_mes_abierto(db, fecha_plazo(i))

    for mercado_nom, share in companias:
        mercado_nom = _mercado_nombre(db, mercado_nom)  # alias ('LSM') → nombre ('Liberty Specialty Markets')
        prima_comp = _q2(prima_neta * share)            # prima de esta compañía (su parte del total)
        cuota = _q2(prima_comp / n)
        for i in range(n):
            prima_i = cuota if i < n - 1 else _q2(prima_comp - cuota * (n - 1))
            imp_i = _q2(prima_i * imp_pct / 100) if imp_pct else D0
            bruta_i = _q2(prima_i + imp_i)
            ced_i = _q2(prima_i * ced_pct / 100) if ced_pct else D0   # comisión corredor
            ret_i = _q2(prima_i * ret_pct / 100) if ret_pct else D0   # comisión Mayrit
            # Tomador → cobramos el 100% (bruta); Corredor → cobramos neto (bruta − cedida).
            adeudada = bruta_i if paga_tomador else _q2(bruta_i - ced_i)
            fe = fecha_plazo(i)
            fv = (fecha_plazo(i + 1) - dt.timedelta(days=1)) if i < n - 1 else poliza.fecha_vencimiento
            anio = fe.year
            recibo = Recibo(
                numero=numero(anio), poliza_id=poliza.id, binder_id=None,
                periodo=fe.strftime("%Y-%m"), anio=anio, yoa=anio, estado="Emitido",
                numero_poliza=poliza.numero_poliza,
                asegurado=poliza.asegurado, corredor=poliza.corredor, ramo=poliza.ramo,
                mercado=_mercado_pago(db, mercado_nom), nombre_mercado=mercado_nom, produccion=poliza.produccion,
                tipo_poliza="Póliza", fecha_efecto=poliza.fecha_efecto, fecha_vencimiento=poliza.fecha_vencimiento,
                pago=PAGO_LABEL.get(n, ""), moneda=poliza.moneda or "EUR",
                prima_neta_poliza=prima_comp, participacion=_q4(share * 100),
                recibo_num=i + 1, recibos_totales=str(n),
                fecha_efecto_recibo=fe, fecha_vcto_recibo=fv,
                prima_neta_recibo=prima_i, impuestos_porc=(imp_pct or None),
                impuestos_recibo=imp_i, prima_bruta_recibo=bruta_i,
                comision_cedida_porc=(ced_pct or None), comision_cedida=ced_i,
                comision_retenida_porc=(ret_pct or None), comision_retenida=ret_i,
                prima_adeudada=adeudada, liquidar=_q2(bruta_i - ced_i - ret_i),
                comision_cedida_a_pagar=ced_i,
                pagador=pagador, fecha_contable=fe,
            )
            _recompute(recibo)
            db.add(recibo)


@router.post("/polizas/emitir", response_model=sch.PolizaRead, status_code=201)
def emitir(payload: PolizaEmitir, db: Session = Depends(get_db)):
    """Crea la póliza y genera sus recibos (plazos × compañías). Ej.: trimestral + 2 compañías = 8."""
    if not payload.fecha_efecto:
        raise HTTPException(status_code=422, detail="La fecha de efecto es obligatoria para emitir.")
    if not payload.prima_neta:
        raise HTTPException(status_code=422, detail="La prima neta es obligatoria para emitir.")

    n = max(1, min(4, payload.n_plazos or 1))
    prima_neta = payload.prima_neta or D0
    cap = payload.capacidad if payload.capacidad is not None else Decimal(1)
    imp_pct = payload.impuestos_porc or D0
    com_pct = payload.comision_porc or D0
    prima_part = _q2(prima_neta * cap)
    imp_part = _q2(prima_part * imp_pct / 100) if imp_pct else D0
    base = payload.model_dump(exclude={"n_plazos", "comision_cedida_porc", "comision_retenida_porc", "plazos_fechas"})
    base.update(
        pago=PAGO_LABEL.get(n, payload.pago),
        prima_participacion=prima_part,
        impuestos=imp_part,
        prima_total=_q2(prima_part + imp_part + _q2(payload.recargos or D0)),
        comision_total=_q2(prima_part * com_pct / 100) if com_pct else D0,
    )
    poliza = Poliza(**base)
    db.add(poliza)
    db.flush()  # asigna poliza.id
    _generar_recibos(db, poliza, n)
    db.commit()
    db.refresh(poliza)
    return sch.PolizaRead.model_validate(poliza)


@router.post("/polizas/{poliza_id}/emitir-recibos", response_model=sch.PolizaRead)
def emitir_recibos_existente(poliza_id: int, db: Session = Depends(get_db)):
    """Genera los recibos de una póliza YA existente que aún no los tiene (mismo criterio:
    plazos según Pago × compañías de coaseguro)."""
    poliza = db.get(Poliza, poliza_id)
    if poliza is None:
        raise HTTPException(status_code=404, detail=f"Póliza {poliza_id} no encontrada")
    ya = db.scalar(select(func.count()).select_from(Recibo).where(Recibo.poliza_id == poliza_id)) or 0
    if ya:
        raise HTTPException(status_code=409, detail="La póliza ya tiene recibos.")
    if not poliza.fecha_efecto:
        raise HTTPException(status_code=422, detail="La fecha de efecto es obligatoria para emitir.")
    if not poliza.prima_neta:
        raise HTTPException(status_code=422, detail="La prima neta es obligatoria para emitir.")
    _generar_recibos(db, poliza, _plazos_de_pago(poliza.pago))
    db.commit()
    db.refresh(poliza)
    return sch.PolizaRead.model_validate(poliza)


# ──────────────────────────── Editar / borrar ───────────────────────────────
# Un recibo "Contabilizado" (enviado al cierre contable mensual) queda BLOQUEADO:
# no se puede editar ni borrar. Para corregir un error hay que reabrirlo primero.
CONTABILIZADO = "Contabilizado"


def _exigir_no_contabilizado(r: Recibo, accion: str = "modificar") -> None:
    if (r.estado or "") == CONTABILIZADO:
        raise HTTPException(
            status_code=409,
            detail=f"El recibo {r.numero} está contabilizado: no se puede {accion}. Reábrelo primero.",
        )


@router.put("/recibos/{recibo_id}", response_model=sch.ReciboRead)
def editar(recibo_id: int, payload: sch.ReciboUpdate, db: Session = Depends(get_db)):
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    _exigir_no_contabilizado(r)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(r, k, v)
    _recompute(r)
    # Una edición directa puede cambiar el cobro/traspaso/liquidación/pago: re-sincroniza sus
    # transferencias (solo recibos no-binder; los de binder van por el Premium).
    transferencias_auto.sync_recibo_todas(db, r)
    db.commit()
    db.refresh(r)
    return _read(db, r)


# ─────────────────── Documento Word del recibo (una plantilla por tipo) ───────────────────
@router.get("/recibos/{recibo_id}/word")
def recibo_word(recibo_id: int, db: Session = Depends(get_db)):
    """Genera el Word del recibo según su tipo (una plantilla por tipo) y lo devuelve para descargar.
    De momento solo Consultoría (su factura de honorarios, plantilla 'Plantilla Factura.dotx')."""
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    tipo = r.tipo_poliza or ""
    if tipo == "Consultoría":
        from .consultoria import factura_docx_para_recibo   # lazy: evita import circular
        data, nombre = factura_docx_para_recibo(db, r)
    elif tipo == "Comisiones":
        from .comisiones import factura_comisiones_docx_para_recibo   # lazy: evita import circular
        data, nombre = factura_comisiones_docx_para_recibo(db, r)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Aún no hay plantilla de Word para recibos de tipo «{tipo or '—'}». De momento solo Consultoría y Comisiones.",
        )
    from urllib.parse import quote
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(nombre)}"},
    )


@router.post("/recibos/{recibo_id}/contabilizar", response_model=sch.ReciboRead)
def contabilizar(recibo_id: int, db: Session = Depends(get_db)):
    """Envía el recibo a contabilidad (cierre mensual): pasa a estado 'Contabilizado' y queda bloqueado."""
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    r.estado = CONTABILIZADO
    db.commit()
    db.refresh(r)
    return _read(db, r)


@router.post("/recibos/{recibo_id}/descontabilizar", response_model=sch.ReciboRead)
def descontabilizar(recibo_id: int, db: Session = Depends(get_db)):
    """Reabre un recibo contabilizado (vuelve a 'Emitido') para poder corregir errores."""
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    r.estado = "Emitido"
    db.commit()
    db.refresh(r)
    return _read(db, r)


# ─────────── Gestión íntegra de un recibo (OM / Fees / Comisiones, uno a uno) ───────────
# Los recibos de BINDER se cobran/traspasan/liquidan parcialmente desde el Premium BDX.
# Los demás (póliza OM, y futuros Fees/Comisiones) se gestionan aquí, ÍNTEGRAMENTE:
#   · Cobrar   → cobra toda la prima (y con ella la comisión retenida y lo 'a liquidar').
#   · Liquidar → paga a la compañía lo cobrado pendiente de liquidar.
#   · Traspasar→ pasa NUESTRA comisión (retenida cobrada) a la cuenta de gastos.
#   · Pagar    → paga la comisión cedida al tercero (corredor).
class GestionRecibo(BaseModel):
    accion: str                      # cobrar | liquidar | traspasar | pagar
    fecha: dt.date | None = None     # por defecto, hoy
    deshacer: bool = False           # revertir la acción
    cuenta_id: int | None = None     # cuenta del movimiento (origen en traspaso)
    cuenta_destino_id: int | None = None  # solo traspaso: cuenta destino (ambas de Mayrit)


@router.post("/recibos/{recibo_id}/gestion", response_model=sch.ReciboRead)
def gestion(recibo_id: int, payload: GestionRecibo, db: Session = Depends(get_db)):
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    if r.binder_id is not None:
        raise HTTPException(status_code=409, detail="Los recibos de binder se gestionan desde el Premium BDX.")
    f = payload.fecha or dt.date.today()
    off = payload.deshacer
    a = payload.accion
    cta = payload.cuenta_id
    cta_dest = payload.cuenta_destino_id
    # Si paga el Corredor, su comisión cedida se salda automáticamente al cobrar (la descuenta él).
    paga_corredor = (r.pagador or "") == "Corredor" and r.poliza_id is not None

    if a == "cobrar":
        # Consultoría y Comisiones se cobran DIRECTO en la cuenta de gastos (no hay custodia de primas):
        # el cobro ES el traspaso → se marca traspasado en el acto, sin movimiento aparte, y nada queda
        # pendiente de traspasar.
        cobro_es_traspaso = r.tipo_poliza in ("Consultoría", "Comisiones")
        if off:
            # El pago de comisión solo bloquea si es manual (Tomador); con Corredor se revierte aquí.
            pago_manual = (r.comision_cedida_pagada or D0) > 0 and not paga_corredor
            # El traspaso automático (consultoría/comisiones) NO bloquea el descobro: se revierte aquí.
            traspaso_bloquea = (r.comision_retenida_traspasada or D0) > 0 and not cobro_es_traspaso
            if (r.liquidar_liquidado or D0) > 0 or traspaso_bloquea or pago_manual:
                raise HTTPException(status_code=409, detail="Deshaz antes la liquidación, el traspaso y el pago de comisión.")
            r.prima_cobrada = r.comision_retenida_cobrada = r.liquidar_cobrado = D0
            r.prima_fecha_cobro = None
            r.cuenta_cobro_id = None
            if cobro_es_traspaso:
                r.comision_retenida_traspasada = D0
                r.comision_fecha_traspaso = None
                r.cuenta_traspaso_origen_id = None
                r.cuenta_traspaso_destino_id = None
            if paga_corredor:
                r.comision_cedida_pagada = D0
                r.comision_cedida_fecha_pago = None
                r.cuenta_pago_id = None
        else:
            # En recibos de Comisiones no hay prima: lo que nos pagan (y se cobra) es la comisión,
            # que vive en deduccion_total. En el resto, la prima adeudada.
            r.prima_cobrada = r.deduccion_total if r.tipo_poliza == "Comisiones" else r.prima_adeudada
            r.comision_retenida_cobrada = r.comision_retenida
            r.liquidar_cobrado = r.liquidar
            r.prima_fecha_cobro = f
            r.cuenta_cobro_id = cta
            if cobro_es_traspaso:   # cobro = traspaso (mismo día y cuenta de gastos), sin transferencia aparte
                r.comision_retenida_traspasada = r.comision_retenida_cobrada
                r.comision_fecha_traspaso = f
                r.cuenta_traspaso_origen_id = cta
                r.cuenta_traspaso_destino_id = cta
            if paga_corredor:
                base = r.comision_cedida_a_pagar or r.comision_cedida or D0
                r.comision_cedida_a_pagar = base
                r.comision_cedida_pagada = base   # el corredor la retuvo → saldada
                r.comision_cedida_fecha_pago = f
                r.cuenta_pago_id = cta            # misma cuenta del cobro (la retuvo el corredor)
    elif a == "traspasar":
        if off:
            r.comision_retenida_traspasada = D0
            r.comision_fecha_traspaso = None
            r.cuenta_traspaso_origen_id = None
            r.cuenta_traspaso_destino_id = None
        else:
            if (r.comision_retenida_cobrada or D0) <= 0:
                raise HTTPException(status_code=409, detail="Primero hay que cobrar la comisión antes de traspasarla.")
            r.comision_retenida_traspasada = r.comision_retenida_cobrada
            r.comision_fecha_traspaso = f
            r.cuenta_traspaso_origen_id = cta
            r.cuenta_traspaso_destino_id = cta_dest
    elif a == "liquidar":
        if off:
            r.liquidar_liquidado = D0
            r.liquidar_fecha_liquidacion = None
            r.cuenta_liquidacion_id = None
        else:
            if (r.liquidar_cobrado or D0) <= 0:
                raise HTTPException(status_code=409, detail="Primero hay que cobrar la prima antes de liquidar a la compañía.")
            r.liquidar_liquidado = r.liquidar_cobrado
            r.liquidar_fecha_liquidacion = f
            r.cuenta_liquidacion_id = cta
    elif a == "pagar":
        if paga_corredor:
            raise HTTPException(status_code=409, detail="Paga el corredor: la comisión cedida se salda automáticamente al cobrar.")
        if off:
            r.comision_cedida_pagada = D0
            r.comision_cedida_fecha_pago = None
            r.cuenta_pago_id = None
        else:
            base = r.comision_cedida_a_pagar or r.comision_cedida or D0
            if base <= 0:
                raise HTTPException(status_code=409, detail="Este recibo no tiene comisión cedida que pagar.")
            r.comision_cedida_a_pagar = base
            r.comision_cedida_pagada = base
            r.comision_cedida_fecha_pago = f
            r.cuenta_pago_id = cta
    else:
        raise HTTPException(status_code=422, detail=f"Acción desconocida: {a!r}")

    _recompute(r)
    # Cierra el ciclo: genera/borra la transferencia (movimiento de dinero) de esta acción.
    transferencias_auto.sync_recibo_accion(db, r, a)
    db.commit()
    db.refresh(r)
    return _read(db, r)


@router.delete("/recibos/{recibo_id}", status_code=204)
def borrar(recibo_id: int, db: Session = Depends(get_db)):
    r = db.get(Recibo, recibo_id)
    if r is None:
        raise HTTPException(status_code=404, detail=f"Recibo {recibo_id} no encontrado")
    _exigir_no_contabilizado(r, "borrar")
    # Borra los movimientos automáticos de este recibo (si no, quedarían huérfanos por el SET NULL).
    transferencias_auto.borrar_recibo(db, recibo_id)
    # Desenlaza las líneas antes de borrar (el FK es SET NULL, pero limpiamos también el texto).
    db.execute(
        update(BdxLinea).where(BdxLinea.recibo_id == recibo_id).values(recibo_id=None, recibo=None)
    )
    db.execute(delete(Recibo).where(Recibo.id == recibo_id))
    db.commit()


# ─────────────────── Cobro vía Premium BDX (deriva el cobro del recibo) ───────────────────
def _comp_linea(l: BdxLinea, excluir_impuestos: bool = False, reaseguro: bool = False):
    """(adeudada, retenida, a_liquidar) de una línea, sobre our line (igual que en la emisión).

    En REASEGURO (caución) la economía es distinta (hay la capa del reasegurado): el cobro es el
    Net Premium to pay to Reinsurance Broker (net_premium_to_broker) y lo 'A Liquidar' es el Final
    Net Premium to UW (final_net_premium_uw); la comisión Mayrit (retenida) es la diferencia. NO
    aplica la fórmula GWP − comisión cedida.

    Si `excluir_impuestos` (binders con impuestos liquidados localmente por la agencia, p. ej.
    agencias italianas), los impuestos NO se cobran ni se liquidan a través de Mayrit: se excluyen
    TANTO del cobro (adeudada) como de 'A Liquidar'. El traspaso (retenida) no cambia."""
    if reaseguro:
        adeudada = l.net_premium_to_broker or D0
        return adeudada, (l.brokerage_amount or D0), (l.final_net_premium_uw or D0)
    neta = l.total_gwp_our_line or D0
    imp = l.total_taxes_levies or D0
    cedida = l.commission_coverholder_amount or D0
    retenida = l.brokerage_amount or D0
    adeudada = (neta - cedida) if excluir_impuestos else (neta + imp - cedida)
    return adeudada, retenida, adeudada - retenida


def _es_reaseguro(db: Session, binder_id: int | None) -> bool:
    """True si el binder cuelga de un programa de reaseguro (economía de recibo distinta)."""
    if not binder_id:
        return False
    return bool(db.execute(
        select(Programa.reaseguro).join(Binder, Binder.programa_id == Programa.id).where(Binder.id == binder_id)
    ).scalar())


def _impuestos_locales(db: Session, binder_id: int | None) -> bool:
    """True si el binder cuelga de un programa con impuestos liquidados localmente (p. ej.
    agencias italianas): sus impuestos NO se liquidan a través nuestro → se excluyen de 'A Liquidar'."""
    if not binder_id:
        return False
    return bool(db.execute(
        select(Programa.impuestos_locales)
        .join(Binder, Binder.programa_id == Programa.id)
        .where(Binder.id == binder_id)
    ).scalar())


def _es_lloyds(db: Session, binder_id: int | None) -> bool:
    """True si alguna sección del binder tiene un mercado de tipo 'Lloyds'. Solo los Lloyd's pasan
    por Xchanging (paso 'Liberado' del LPAN); los de Compañía no lo requieren para liquidar."""
    if not binder_id:
        return False
    tipos = db.execute(
        select(Mercado.tipo_mercado)
        .join(SeccionMercado, SeccionMercado.mercado_id == Mercado.id)
        .join(BinderSeccion, BinderSeccion.id == SeccionMercado.seccion_id)
        .where(BinderSeccion.binder_id == binder_id)
    ).scalars().all()
    return any((t or "").strip().lower() == "lloyds" for t in tipos)


def _recalcular_cobro_recibo(db: Session, recibo: Recibo) -> None:
    """El cobro/traspaso/liquidación del recibo se DERIVAN de sus líneas (vía Premium)."""
    lineas = db.scalars(select(BdxLinea).where(BdxLinea.recibo_id == recibo.id)).all()
    excl = _impuestos_locales(db, recibo.binder_id)
    rea = _es_reaseguro(db, recibo.binder_id)
    adeu = ret = liq = ret_tras = liq_liq = D0
    f_cobro, f_tras, f_liq = [], [], []
    for l in lineas:
        a, r, q = _comp_linea(l, excl, rea)
        if l.prima_cobrada:
            adeu += a
            ret += r
            liq += q
            if l.premium_payment_date:
                f_cobro.append(l.premium_payment_date)
        if l.traspaso:
            ret_tras += r
            if l.fecha_traspaso:
                f_tras.append(l.fecha_traspaso)
        if l.liquidado:
            liq_liq += q
            if l.fecha_liquidacion:
                f_liq.append(l.fecha_liquidacion)
    recibo.prima_cobrada = _q2(adeu)
    recibo.comision_retenida_cobrada = _q2(ret)
    recibo.liquidar_cobrado = _q2(liq)
    recibo.comision_retenida_traspasada = _q2(ret_tras)
    recibo.liquidar_liquidado = _q2(liq_liq)
    recibo.prima_fecha_cobro = max(f_cobro) if f_cobro else None
    recibo.comision_fecha_traspaso = max(f_tras) if f_tras else None
    recibo.liquidar_fecha_liquidacion = max(f_liq) if f_liq else None
    _recompute(recibo)


def _premium_bloqueado(db: Session, binder_id: int, periodo: str) -> bool:
    return db.scalar(
        select(BdxBloqueo).where(
            BdxBloqueo.binder_id == binder_id, BdxBloqueo.tipo == "premium", BdxBloqueo.periodo == periodo
        )
    ) is not None


def _exigir_premium_no_bloqueado(db: Session, binder_id: int, periodo: str):
    if _premium_bloqueado(db, binder_id, periodo):
        raise HTTPException(
            status_code=409,
            detail=f"El Premium {periodo} está bloqueado: no se puede modificar ni cambiar su cobro.",
        )


def _lineas_premium(db: Session, binder_id: int, periodo: str):
    ini, fin = _rango_mes(periodo)
    return db.scalars(
        select(BdxLinea)
        .join(Bdx, BdxLinea.bdx_id == Bdx.id)
        .where(
            Bdx.binder_id == binder_id,
            BdxLinea.incluido_en_premium.is_(True),
            BdxLinea.premium_bdx >= ini,
            BdxLinea.premium_bdx < fin,
        )
    ).all()


class PremiumGrupo(BaseModel):
    periodo: str            # 'YYYY-MM' del Premium
    num_lineas: int
    prima: Decimal          # Σ adeudada de sus líneas
    comision: Decimal       # Σ comisión retenida (brokerage)
    a_liquidar: Decimal     # Σ (adeudada − retenida)
    prima_lloyds: Decimal   # Σ Net Premium to Lloyd's Broker (net_premium_to_broker)
    cobrado: bool           # todas sus líneas cobradas
    traspasado: bool        # todas sus líneas traspasadas
    liquidado: bool         # todas sus líneas liquidadas
    tiene_recibo: bool = True  # todas sus líneas tienen recibo generado (si no, no se puede cobrar/etc.)
    fecha_pago: dt.date | None = None
    fecha_traspaso: dt.date | None = None
    fecha_liquidacion: dt.date | None = None
    nota: str | None = None    # nota libre del mes (PremiumNota)


class AccionPremium(BaseModel):
    periodo: str
    fecha: dt.date


@router.get("/binders/{binder_id}/premium", response_model=list[PremiumGrupo])
def listar_premium(binder_id: int, db: Session = Depends(get_db)):
    """Grupos de Premium del binder (líneas incluidas en premium, agrupadas por mes)."""
    # load_only: solo las 11 columnas que usa el bucle (+ _comp_linea), en vez de las ~90 de la
    # entidad. Misma lógica y mismo resultado; solo reduce transferencia/hidratación.
    lineas = db.scalars(
        select(BdxLinea)
        .join(Bdx, BdxLinea.bdx_id == Bdx.id)
        .where(Bdx.binder_id == binder_id, BdxLinea.incluido_en_premium.is_(True), BdxLinea.premium_bdx.is_not(None))
        .options(load_only(
            BdxLinea.premium_bdx,
            BdxLinea.total_gwp_our_line, BdxLinea.total_taxes_levies,
            BdxLinea.commission_coverholder_amount, BdxLinea.brokerage_amount,
            BdxLinea.net_premium_to_broker, BdxLinea.final_net_premium_uw, BdxLinea.recibo_id,
            BdxLinea.prima_cobrada, BdxLinea.premium_payment_date,
            BdxLinea.traspaso, BdxLinea.fecha_traspaso,
            BdxLinea.liquidado, BdxLinea.fecha_liquidacion,
        ))
    ).all()
    excl = _impuestos_locales(db, binder_id)
    rea = _es_reaseguro(db, binder_id)
    grupos: dict[str, dict] = {}
    for l in lineas:
        per = l.premium_bdx.strftime("%Y-%m")
        g = grupos.setdefault(per, {"num": 0, "conrec": 0, "prima": D0, "com": D0, "liq": D0, "npb": D0, "cob": 0, "tra": 0, "liqd": 0, "fc": [], "ft": [], "fl": []})
        a, r, q = _comp_linea(l, excl, rea)
        g["num"] += 1
        if l.recibo_id:
            g["conrec"] += 1
        g["prima"] += a
        g["com"] += r
        g["liq"] += q
        g["npb"] += (l.net_premium_to_broker or D0)
        if l.prima_cobrada:
            g["cob"] += 1
            if l.premium_payment_date:
                g["fc"].append(l.premium_payment_date)
        if l.traspaso:
            g["tra"] += 1
            if l.fecha_traspaso:
                g["ft"].append(l.fecha_traspaso)
        if l.liquidado:
            g["liqd"] += 1
            if l.fecha_liquidacion:
                g["fl"].append(l.fecha_liquidacion)
    notas = {n.periodo: n.nota for n in db.scalars(
        select(PremiumNota).where(PremiumNota.binder_id == binder_id)).all()}
    return [
        PremiumGrupo(
            periodo=per,
            num_lineas=g["num"],
            prima=_q2(g["prima"]),
            comision=_q2(g["com"]),
            a_liquidar=_q2(g["liq"]),
            prima_lloyds=_q2(g["npb"]),
            cobrado=g["num"] > 0 and g["cob"] == g["num"],
            traspasado=g["num"] > 0 and g["tra"] == g["num"],
            liquidado=g["num"] > 0 and g["liqd"] == g["num"],
            # "Tiene recibo" = todas sus líneas están enlazadas a un recibo (el de su mes de RIESGO).
            # No depende de que exista recibo del mes del Premium (los binders en run-off siguen
            # recibiendo Premium de riesgos pasados, sin Risk nuevo ese mes).
            tiene_recibo=g["num"] > 0 and g["conrec"] == g["num"],
            fecha_pago=max(g["fc"]) if g["fc"] else None,
            fecha_traspaso=max(g["ft"]) if g["ft"] else None,
            fecha_liquidacion=max(g["fl"]) if g["fl"] else None,
            nota=notas.get(per),
        )
        for per, g in sorted(grupos.items())
    ]


class PremiumNotaIn(BaseModel):
    periodo: str            # 'YYYY-MM'
    nota: str | None = None


@router.put("/binders/{binder_id}/premium/nota")
def guardar_nota_premium(binder_id: int, payload: PremiumNotaIn, db: Session = Depends(get_db)):
    """Crea/actualiza/borra la nota libre de un mes de Premium del binder. nota vacía = se borra."""
    if db.get(Binder, binder_id) is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    texto = (payload.nota or "").strip() or None
    fila = db.scalar(select(PremiumNota).where(
        PremiumNota.binder_id == binder_id, PremiumNota.periodo == payload.periodo))
    if texto is None:
        if fila is not None:
            db.delete(fila)
            db.commit()
        return {"periodo": payload.periodo, "nota": None}
    if fila is None:
        fila = PremiumNota(binder_id=binder_id, periodo=payload.periodo, nota=texto)
        db.add(fila)
    else:
        fila.nota = texto
    db.commit()
    return {"periodo": payload.periodo, "nota": texto}


def _accion_premium(db: Session, binder_id: int, periodo: str, setter, exigir_recibo: bool = False, mov: dict | None = None) -> dict:
    """Aplica una acción (setter) a todas las líneas del Premium y recalcula los recibos.
    Si `exigir_recibo`, exige que el periodo tenga su Recibo generado (líneas con recibo_id):
    no se puede cobrar/liquidar/traspasar una prima sin recibo emitido.
    Si `mov` (dict con tipo/subtipo/fecha/importe_de), genera la transferencia del binder con la
    suma de las líneas (o la borra al deshacer, cuando la suma queda a 0)."""
    _exigir_premium_no_bloqueado(db, binder_id, periodo)
    lineas = _lineas_premium(db, binder_id, periodo)
    if not lineas:
        raise HTTPException(status_code=400, detail=f"No hay líneas en el Premium {periodo}.")
    # Cada línea se cobra/liquida contra el recibo de su mes de RIESGO (recibo_id). No se exige
    # un recibo del mes del Premium: un binder en run-off recibe Premium de riesgos pasados.
    if exigir_recibo and any(l.recibo_id is None for l in lineas):
        raise HTTPException(
            status_code=409,
            detail="Hay líneas de este Premium sin recibo (su mes de riesgo no tiene recibo generado). Genera primero el recibo del mes de riesgo correspondiente.",
        )
    for l in lineas:
        setter(l)
    db.flush()
    rids = {l.recibo_id for l in lineas if l.recibo_id}
    recibos = db.scalars(select(Recibo).where(Recibo.id.in_(rids))).all() if rids else []
    for r in recibos:
        _recalcular_cobro_recibo(db, r)
    if mov is not None:
        binder = db.get(Binder, binder_id)
        importe = sum((Decimal(mov["importe_de"](l) or 0) for l in lineas), Decimal(0))
        transferencias_auto.sync_binder(
            db, binder, periodo=periodo, tipo=mov["tipo"], subtipo=mov["subtipo"],
            importe=importe, fecha=mov["fecha"],
        )
    db.commit()
    return {"lineas": len(lineas), "recibos_actualizados": len(recibos)}


@router.post("/binders/{binder_id}/premium/cobrar")
def cobrar_premium(binder_id: int, payload: AccionPremium, db: Session = Depends(get_db)):
    """💰 Cobrar: marca las líneas como cobradas (fecha real) → Cantidad Cobrada y Pdte. Cobro en los recibos."""
    def setter(l):
        l.prima_cobrada = True
        l.premium_payment_date = payload.fecha
        l.ingresado = l.net_premium_to_broker   # importe cobrado de la línea (para la columna 'Cobrado' del BDX)
    return _accion_premium(db, binder_id, payload.periodo, setter, exigir_recibo=True,
                           mov={"tipo": "Primas", "subtipo": "Cobro", "fecha": payload.fecha, "importe_de": lambda l: l.ingresado})


@router.post("/binders/{binder_id}/premium/descobrar")
def descobrar_premium(binder_id: int, payload: AccionPremium, db: Session = Depends(get_db)):
    """Deshace el cobro de un Premium (vuelve a pendiente)."""
    def setter(l):
        l.prima_cobrada = False
        l.premium_payment_date = None
        l.ingresado = None                       # revierte el importe cobrado de la línea
    # importe queda a 0 → la transferencia de Cobro se borra sola.
    return _accion_premium(db, binder_id, payload.periodo, setter,
                           mov={"tipo": "Primas", "subtipo": "Cobro", "fecha": None, "importe_de": lambda l: l.ingresado})


@router.post("/binders/{binder_id}/premium/traspasar")
def traspasar_premium(binder_id: int, payload: AccionPremium, db: Session = Depends(get_db)):
    """🔁 Traspasar: lleva NUESTRA comisión de la cuenta de primas a la de gastos."""
    def setter(l):
        l.traspaso = True
        l.fecha_traspaso = payload.fecha
        l.traspasado = l.brokerage_amount
    return _accion_premium(db, binder_id, payload.periodo, setter, exigir_recibo=True,
                           mov={"tipo": "Comisiones", "subtipo": "Traspaso", "fecha": payload.fecha, "importe_de": lambda l: l.traspasado})


@router.post("/binders/{binder_id}/premium/liquidar")
def liquidar_premium(binder_id: int, payload: AccionPremium, db: Session = Depends(get_db)):
    """🏦 Liquidar: paga a la compañía/Lloyd's la parte a liquidar (adeudada − comisión retenida).
    EXIGE que existan LPAN que cubran el neto a pagar al mercado (los LPAN controlan la liquidación,
    en Lloyd's Y en Compañía; la única diferencia es que los Lloyd's requieren FDO previo al LPAN, lo
    cual se valida al generar el LPAN), que su neto CUADRE con el del Premium, y que estén **Liberados**.
    Al liquidar, sella su fecha de pago (=liquidación) con la fecha de la liquidación."""
    lpans = db.scalars(select(Lpan).where(Lpan.binder_id == binder_id, Lpan.periodo == payload.periodo)).all()

    # 1) Los LPAN son OBLIGATORIOS para liquidar (Lloyd's y Compañía): controlan la liquidación al
    # mercado. Su neto (Σ net_premium) debe cuadrar con el neto a pagar del Premium (Σ Final Net Premium
    # to UW de sus líneas), con el que se construye el propio LPAN. Si no hay LPAN, o su neto no cuadra
    # (falta/sobra un LPAN, o cambiaron las líneas), no se liquida.
    neto_premium = _q2(sum((l.final_net_premium_uw or D0) for l in _lineas_premium(db, binder_id, payload.periodo)))
    neto_lpan = _q2(sum((lp.net_premium or D0) for lp in lpans))
    tol = Decimal("0.01") * max(len(lpans), 1)   # solo absorbe céntimos de redondeo
    if abs(neto_premium - neto_lpan) > tol:
        b = db.get(Binder, binder_id)
        moneda = (b.moneda if b else None) or "EUR"
        if not lpans:
            detalle = (f"No se puede liquidar: este Premium no tiene LPAN generados (neto a pagar al "
                       f"mercado {neto_premium:,.2f} {moneda}). Genera primero el/los LPAN de este periodo.")
        else:
            detalle = (f"No se puede liquidar: las cantidades no coinciden. Neto del Premium "
                       f"{neto_premium:,.2f} {moneda} vs suma de los {len(lpans)} LPAN {neto_lpan:,.2f} "
                       f"{moneda} (diferencia {abs(neto_premium - neto_lpan):,.2f} {moneda}). Revisa o "
                       f"regenera los LPAN de este periodo antes de liquidar.")
        raise HTTPException(status_code=409, detail=detalle)

    # 2) Solo en binders LLOYD'S: todos los LPAN deben estar Liberados (por Xchanging) antes de pagar
    # al mercado. Los de Compañía no pasan por Xchanging, así que NO se exige el paso 'Liberado'.
    if _es_lloyds(db, binder_id):
        sin_liberar = [lp for lp in lpans if lp.liberado is None]
        if sin_liberar:
            refs = ", ".join((lp.broker_ref2 or f"LPAN {lp.id}") for lp in sin_liberar[:6])
            mas = f" y {len(sin_liberar) - 6} más" if len(sin_liberar) > 6 else ""
            raise HTTPException(
                status_code=409,
                detail=f"No se puede liquidar: {len(sin_liberar)} LPAN de este Premium sin fecha de Liberado "
                       f"({refs}{mas}). Cumpliméntala primero.",
            )

    # 3) Al liquidar el Premium, sella la fecha de liquidación (pagado) en los LPAN que aún no la tengan.
    for lp in lpans:
        if lp.pagado is None:
            lp.pagado = payload.fecha

    excl = _impuestos_locales(db, binder_id)
    rea = _es_reaseguro(db, binder_id)
    def setter(l):
        a, r, q = _comp_linea(l, excl, rea)
        l.liquidado = True
        l.fecha_liquidacion = payload.fecha
        l.liquidado_uw = _q2(q)
    return _accion_premium(db, binder_id, payload.periodo, setter, exigir_recibo=True,
                           mov={"tipo": "Primas", "subtipo": "Liquidación", "fecha": payload.fecha, "importe_de": lambda l: l.liquidado_uw})


# ─────────────── Macheo automático del Premium desde un Excel subido ───────────────
def _cabecera(ws, max_scan: int = 12):
    """Detecta la fila de cabecera (la que más celdas de texto tiene) y devuelve (idx, columnas)."""
    filas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        filas.append(row)
        if i >= max_scan:
            break
    best_i, best_n = 0, -1
    for i, row in enumerate(filas):
        n = sum(1 for v in row if isinstance(v, str) and v.strip())
        if n > best_n:
            best_i, best_n = i, n
    cols = [(str(v).strip() if v is not None else "") for v in filas[best_i]]
    return best_i, cols


def _sugerir(cols: list[str], guardado: str | None, claves: list[str]) -> str | None:
    """Sugiere una columna: primero la guardada (si existe), si no por palabras clave EN ORDEN
    de prioridad (la 1ª clave manda; así 'our line' gana a un genérico 'premium')."""
    if guardado and guardado in cols:
        return guardado
    for k in claves:
        for c in cols:
            if k in c.lower():
                return c
    return None


# Caché en memoria del Excel subido, para NO re-subir/re-parsear el mismo fichero en cada paso del
# macheo (preview inicial, cambio de hoja y machear). El front sube el fichero UNA vez, recibe un
# `token` y lo reutiliza; si el token caduca (10 min) o falla, reintenta subiendo el fichero.
# (El backend corre con 1 worker gunicorn → la caché en proceso es suficiente.)
_XLSX_TTL = 600.0
_XLSX_CACHE: dict[str, tuple[float, bytes]] = {}


def _xlsx_cache_put(content: bytes) -> str:
    ahora = time.time()
    for k in [k for k, (exp, _) in _XLSX_CACHE.items() if exp < ahora]:
        _XLSX_CACHE.pop(k, None)
    tok = uuid.uuid4().hex
    _XLSX_CACHE[tok] = (ahora + _XLSX_TTL, content)
    return tok


def _xlsx_cache_get(token: str | None) -> bytes | None:
    if not token:
        return None
    v = _XLSX_CACHE.get(token)
    if not v:
        return None
    exp, content = v
    if exp < time.time():
        _XLSX_CACHE.pop(token, None)
        return None
    return content


async def _xlsx_contenido(file: UploadFile | None, token: str | None) -> tuple[bytes, str]:
    """Devuelve (bytes, token) del Excel: reutiliza el de la caché si el token es válido; si no, lee
    el fichero subido y lo cachea. Si no hay ni token válido ni fichero, pide re-subir (409)."""
    cached = _xlsx_cache_get(token)
    if cached is not None:
        return cached, token   # reutiliza (no se re-sube ni re-parsea la subida)
    if file is None:
        raise HTTPException(status_code=409, detail="token_caducado")   # el front reintenta con el fichero
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se admite .xlsx (convierte los .xls antes de subir).")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="El fichero está vacío.")
    return content, _xlsx_cache_put(content)


def _wb_de_bytes(content: bytes):
    """Carga un workbook openpyxl (solo lectura) desde bytes."""
    try:
        return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")


@router.post("/binders/{binder_id}/premium/excel-preview")
async def excel_preview(binder_id: int, file: UploadFile | None = File(None), hoja: str | None = Form(None),
                        token: str | None = Form(None), db: Session = Depends(get_db)):
    """Lee hojas/cabeceras del Excel subido y sugiere el mapeo (recordado de la agencia o por palabras clave)."""
    binder = db.get(Binder, binder_id)
    if binder is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    content, token = await _xlsx_contenido(file, token)
    wb = _wb_de_bytes(content)
    hoja = hoja if (hoja and hoja in wb.sheetnames) else wb.sheetnames[0]
    ws = wb[hoja]
    hdr_i, cols = _cabecera(ws)
    columnas = [c for c in cols if c]
    # Cuenta TODAS las filas de datos (no vacías) tras la cabecera; muestra las 3 primeras.
    n_filas = 0
    muestra = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hdr_i:
            continue
        fila = {cols[j]: ("" if v is None else str(v)) for j, v in enumerate(row) if j < len(cols) and cols[j]}
        if not any(fila.values()):
            continue
        n_filas += 1
        if len(muestra) < 3:
            muestra.append(fila)
    prod = binder.productor
    return {
        "hojas": wb.sheetnames,
        "hoja": hoja,
        "token": token,
        "columnas": columnas,
        "n_filas": n_filas,
        "muestra": muestra,
        "mapeo": {
            "certificado": _sugerir(columnas, prod.premium_col_certificado if prod else None, ["certificate", "certificado", "cert ref", "policy", "poliza"]),
            # Se compara contra el Net Premium to Lloyd's Broker del Risk → se sugiere SIEMPRE esa misma
            # columna del Excel (ignorando la recordada, que podía ser otra p. ej. "Gross ... Our Line").
            "importe": _sugerir(columnas, None, ["net premium to lloyd", "net premium to broker", "net premium to pay", "net to broker", "net premium"]),
        },
    }


class MatchRow(BaseModel):
    certificate_ref: str
    importe_excel: Decimal | None = None
    estado: str                 # 'match' | 'importe_distinto' | 'no_encontrada'
    linea_id: int | None = None
    importe_risk: Decimal | None = None
    risk_bdx: str | None = None  # periodo(s) 'YYYY-MM' del Risk de la(s) línea(s) macheada(s) ('a / b' si varias)
    risk_lineas: int = 1         # nº de líneas del Risk que representa esta fila (>1 si se machea por suma)


def _a_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "."))
    except (ValueError, TypeError, ArithmeticError):
        return None


@router.post("/binders/{binder_id}/premium/match-excel")
async def match_excel(binder_id: int, file: UploadFile | None = File(None), hoja: str = Form(...),
                      certificado: str = Form(...), importe: str | None = Form(None),
                      periodo: str = Form(...), token: str | None = Form(None), db: Session = Depends(get_db)):
    """Casa las filas del Excel subido con las líneas Risk del binder por Certificate Ref (importe como
    comprobación). Guarda el mapeo en la agencia. NO aplica: devuelve preview + ids macheados."""
    binder = db.get(Binder, binder_id)
    if binder is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    _rango_mes(periodo)
    content, _token = await _xlsx_contenido(file, token)
    wb = _wb_de_bytes(content)
    if hoja not in wb.sheetnames:
        raise HTTPException(status_code=404, detail=f"Hoja '{hoja}' no encontrada")
    ws = wb[hoja]
    hdr_i, cols = _cabecera(ws)
    try:
        cert_idx = cols.index(certificado)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Columna de certificado '{certificado}' no está en la hoja")
    imp_idx = cols.index(importe) if (importe and importe in cols) else None

    # Líneas Risk del binder indexadas por certificate_ref. load_only: solo las columnas que usan el
    # macheo (id, certificado, net) y el cálculo de totales (_comp_linea), en vez de las ~90 de la
    # entidad → mucho menos que traer y que hidratar (el binder puede tener miles de líneas).
    risk = db.scalars(
        select(BdxLinea).join(Bdx, BdxLinea.bdx_id == Bdx.id).where(Bdx.binder_id == binder_id)
        .options(load_only(
            BdxLinea.certificate_ref, BdxLinea.net_premium_to_broker, BdxLinea.reporting_period_start,
            BdxLinea.total_gwp_our_line, BdxLinea.total_taxes_levies,
            BdxLinea.commission_coverholder_amount, BdxLinea.brokerage_amount,
            BdxLinea.final_net_premium_uw,
        ))
    ).all()
    por_cert: dict[str, list[BdxLinea]] = {}
    for l in risk:
        if l.certificate_ref:
            por_cert.setdefault(l.certificate_ref.strip().lower(), []).append(l)

    filas: list[MatchRow] = []
    matched_ids: list[int] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hdr_i:
            continue
        cert_raw = row[cert_idx] if cert_idx < len(row) else None
        if cert_raw is None or str(cert_raw).strip() == "":
            continue
        cert = str(cert_raw).strip()
        imp = _a_decimal(row[imp_idx]) if (imp_idx is not None and imp_idx < len(row)) else None
        cands = por_cert.get(cert.lower(), [])
        if not cands:
            filas.append(MatchRow(certificate_ref=cert, importe_excel=imp, estado="no_encontrada"))
            continue
        # La comparación se hace contra el "Net Premium to Lloyd's Broker" del Risk
        # (net_premium_to_broker), que es el importe que DE VERDAD cuenta para conciliar el Premium.
        # (a) Línea individual más cercana al importe del Premium.
        best = cands[0]
        best_diff = None
        for l in cands:
            amt = l.net_premium_to_broker
            if amt is None:
                continue
            d = abs(Decimal(amt) - (imp if imp is not None else Decimal(amt)))
            if best_diff is None or d < best_diff:
                best, best_diff = l, d
        # (b) SUMA de TODAS las líneas del mismo certificado: un único apunte del Premium puede estar
        # liquidando VARIAS líneas del Risk (endosos/ajustes con el mismo Certificate). Si la suma
        # cuadra con el Premium, se machean TODAS esas líneas.
        suma = _q2(sum((l.net_premium_to_broker or D0) for l in cands))
        tol = max(Decimal("0.02"), abs(imp) * Decimal("0.01")) if imp is not None else D0
        diff_suma = abs(imp - suma) if imp is not None else None
        sum_ok = imp is not None and len(cands) > 1 and diff_suma is not None and diff_suma <= tol
        single_ok = imp is None or (best_diff is not None and best_diff <= tol)

        def periodos(ls):
            ps = sorted({l.reporting_period_start.strftime("%Y-%m") for l in ls if l.reporting_period_start})
            return " / ".join(ps) if ps else None
        def per_best():
            return best.reporting_period_start.strftime("%Y-%m") if best.reporting_period_start else None

        # Prioriza la SUMA cuando cuadra (es el caso de "un apunte liquida varias líneas"); si no,
        # la línea individual; si nada cuadra, importe distinto (mostrando lo más cercano).
        prefiere_suma = sum_ok and (not single_ok or (best_diff is not None and diff_suma <= best_diff))
        if prefiere_suma:
            filas.append(MatchRow(certificate_ref=cert, importe_excel=imp, estado="match", linea_id=best.id,
                                  importe_risk=suma, risk_bdx=periodos(cands), risk_lineas=len(cands)))
            matched_ids.extend(l.id for l in cands)
        elif single_ok:
            filas.append(MatchRow(certificate_ref=cert, importe_excel=imp, estado="match", linea_id=best.id,
                                  importe_risk=_q2(best.net_premium_to_broker or 0), risk_bdx=per_best(), risk_lineas=1))
            matched_ids.append(best.id)
        elif len(cands) > 1 and diff_suma is not None and best_diff is not None and diff_suma < best_diff:
            filas.append(MatchRow(certificate_ref=cert, importe_excel=imp, estado="importe_distinto", linea_id=best.id,
                                  importe_risk=suma, risk_bdx=periodos(cands), risk_lineas=len(cands)))
        else:
            filas.append(MatchRow(certificate_ref=cert, importe_excel=imp, estado="importe_distinto", linea_id=best.id,
                                  importe_risk=_q2(best.net_premium_to_broker or 0), risk_bdx=per_best(), risk_lineas=1))

    # Recordar el mapeo en la agencia
    if binder.productor:
        binder.productor.premium_col_certificado = certificado
        binder.productor.premium_col_importe = importe
        db.commit()

    resumen = {
        "total": len(filas),
        "match": sum(1 for f in filas if f.estado == "match"),
        "importe_distinto": sum(1 for f in filas if f.estado == "importe_distinto"),
        "no_encontrada": sum(1 for f in filas if f.estado == "no_encontrada"),
    }
    # Sumatorio del Premium (líneas macheadas): A Cobrar / A Traspasar / A Liquidar, con la economía
    # del binder (reaseguro incluido), igual que en la emisión.
    excl = _impuestos_locales(db, binder_id)
    rea = _es_reaseguro(db, binder_id)
    byid = {l.id: l for l in risk}
    t_cobrar = t_tras = t_liq = D0
    for lid in matched_ids:
        l = byid.get(lid)
        if l is None:
            continue
        a, r, q = _comp_linea(l, excl, rea)
        t_cobrar += a; t_tras += r; t_liq += q
    premium = {"cobrar": _q2(t_cobrar), "traspasar": _q2(t_tras), "liquidar": _q2(t_liq)}
    return {"periodo": periodo, "filas": filas, "matched_ids": matched_ids, "resumen": resumen, "premium": premium}


# ──────────────────────── Exportación genérica a Excel (.xlsx) ────────────────────────
# El frontend manda ya las columnas (cabeceras) y los valores formateados/numéricos que
# quiere exportar; aquí solo se genera el .xlsx con el estilo de la casa: Calibri 9 y
# cabecera en gris + negrita.
class ExportXlsx(BaseModel):
    nombre: str = "export"
    hoja: str = "Datos"
    headers: list[str]
    filas: list[list[str | float | int | None]]


@router.post("/export/xlsx")
def export_xlsx(payload: ExportXlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (payload.hoja or "Datos")[:31]

    head_font = Font(name="Calibri", size=9, bold=True)
    head_fill = PatternFill("solid", fgColor="D9D9D9")  # gris claro
    body_font = Font(name="Calibri", size=9)

    ws.append(payload.headers)
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill

    for fila in payload.filas:
        ws.append(fila)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = body_font

    # Anchos de columna aproximados al contenido (acotados).
    for i, h in enumerate(payload.headers, start=1):
        ancho = len(str(h))
        for fila in payload.filas:
            v = fila[i - 1] if i - 1 < len(fila) else None
            if v is not None:
                ancho = max(ancho, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max(ancho + 2, 8), 50)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = (payload.nombre or "export").replace('"', "")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}.xlsx"'},
    )
