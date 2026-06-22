"""
Triangulación de siniestralidad.

Estructura (la del usuario): eje de MESES de calendario, desde el inicio del binder hasta el
último mes con snapshot.
  - Filas    = mes de apertura del siniestro (date_opened; respaldo: claim_first_advised o su
               primer mes de aparición). Se muestran TODOS los meses del binder, aunque no haya
               siniestros (salen en 0). Cada fila lleva su GWP Our Line del mes.
  - Columnas = mes de valuación (calendario). La celda (mes_origen, mes_valuación) es la
               siniestralidad de los siniestros abiertos en `mes_origen` valuada a `mes_valuación`
               (último snapshot ≤ ese mes; arrastra el anterior). Vacío si valuación < origen.
  - Fila inferior = Total por columna (siniestralidad total a ese mes). La última columna = hoy.
  - 3 métricas conmutables: Incurrido (pagado+reservas), Pagado, Nº de siniestros.
  - IBNR sugerido por chain-ladder (sobre el desarrollo por antigüedad).
"""
from __future__ import annotations

import io

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..db import get_db
from ..models.maestras import Bdx, BdxLinea, Binder, ClaimsPresentacion, Programa, Siniestro

router = APIRouter(tags=["Triangulación"])

METRICAS = {"incurrido": "Incurrido", "pagado": "Pagado", "num": "Nº siniestros", "pct": "% Siniestralidad"}


def _scope_label(seccion: int | None, risk_code: str | None) -> str:
    if risk_code:
        return f"Código {risk_code}"
    if seccion is not None:
        return f"Sección {seccion}"
    return "Total"


def _mi(anio: int, mes: int) -> int:
    return anio * 12 + (mes - 1)


def _mi_de_ord(ord_: int) -> int:
    return _mi(ord_ // 100, ord_ % 100)


def _periodo_de_mi(mi: int) -> str:
    return f"{mi // 12:04d}-{mi % 12 + 1:02d}"


def _risk_scope(stmt, seccion: int | None, risk_code: str | None):
    if seccion is not None:
        stmt = stmt.where(BdxLinea.section_no == seccion)
    if risk_code:
        stmt = stmt.where(BdxLinea.risk_code == risk_code)
    return stmt


def _bases(db: Session, binder_id: int, seccion=None, risk_code=None) -> tuple[float, float]:
    """(GWP our line, Net to UWs) de las líneas Risk del ámbito. Net = GWP − com − brokerage."""
    q = (
        select(
            func.coalesce(func.sum(BdxLinea.total_gwp_our_line), 0),
            func.coalesce(func.sum(BdxLinea.commission_coverholder_amount), 0),
            func.coalesce(func.sum(BdxLinea.brokerage_amount), 0),
        )
        .select_from(BdxLinea).join(Bdx, Bdx.id == BdxLinea.bdx_id)
        .where(Bdx.binder_id == binder_id, Bdx.tipo == "Risk")
    )
    gwp, cc, brk = (float(x) for x in db.execute(_risk_scope(q, seccion, risk_code)).first())
    return round(gwp, 2), round(gwp - cc - brk, 2)


def _net_por_mes(db: Session, binder_id: int, seccion=None, risk_code=None) -> dict[int, float]:
    """Net to UWs por mes (reporting_period_start) de las líneas Risk del ámbito."""
    q = (
        select(
            BdxLinea.reporting_period_start,
            func.coalesce(func.sum(BdxLinea.total_gwp_our_line), 0)
            - func.coalesce(func.sum(BdxLinea.commission_coverholder_amount), 0)
            - func.coalesce(func.sum(BdxLinea.brokerage_amount), 0),
        )
        .join(Bdx, Bdx.id == BdxLinea.bdx_id)
        .where(Bdx.binder_id == binder_id, Bdx.tipo == "Risk", BdxLinea.reporting_period_start.is_not(None))
    )
    q = _risk_scope(q, seccion, risk_code).group_by(BdxLinea.reporting_period_start)
    out: dict[int, float] = {}
    for d, s in db.execute(q).all():
        out[_mi(d.year, d.month)] = out.get(_mi(d.year, d.month), 0.0) + float(s)
    return out


def _opciones(db: Session, binder_id: int) -> tuple[list[int], list[str]]:
    """Secciones y códigos de riesgo disponibles (de los siniestros del binder)."""
    secs, rcs = set(), set()
    for s in db.scalars(select(Siniestro).where(Siniestro.binder_id == binder_id)).all():
        if s.section is not None:
            secs.add(s.section)
        if s.risk_code:
            rcs.add(s.risk_code)
    return sorted(secs), sorted(rcs)


def _payload_binder(db: Session, b: Binder, seccion: int | None, risk_code: str | None) -> dict:
    """Triángulo del binder para un ámbito (Total / Sección / Código de riesgo)."""
    binder_id = b.id
    gwp_our_line, net_uw = _bases(db, binder_id, seccion, risk_code)
    net_mi = _net_por_mes(db, binder_id, seccion, risk_code)
    secciones, risk_codes = _opciones(db, binder_id)
    base = {
        "gwp_our_line": gwp_our_line, "net_uw": net_uw,
        "secciones": secciones, "risk_codes": risk_codes,
        "ambito": _scope_label(seccion, risk_code),
    }

    q = (
        select(
            ClaimsPresentacion.siniestro_id, ClaimsPresentacion.periodo_ord,
            ClaimsPresentacion.paid_indemnity_acum, ClaimsPresentacion.paid_fees_acum,
            ClaimsPresentacion.reserves_indemnity, ClaimsPresentacion.reserves_fees,
        )
        .join(Siniestro, Siniestro.id == ClaimsPresentacion.siniestro_id)
        .where(ClaimsPresentacion.binder_id == binder_id, ClaimsPresentacion.siniestro_id.is_not(None))
    )
    if seccion is not None:
        q = q.where(Siniestro.section == seccion)
    if risk_code:
        q = q.where(Siniestro.risk_code == risk_code)
    pres = db.execute(q.order_by(ClaimsPresentacion.siniestro_id, ClaimsPresentacion.periodo_ord)).all()

    snaps: dict[int, list[tuple[int, float, float]]] = {}
    for sid, ord_, pi, pf, ri, rf in pres:
        mi = _mi_de_ord(ord_)
        snaps.setdefault(sid, []).append(
            (mi, float(pi or 0) + float(pf or 0) + float(ri or 0) + float(rf or 0), float(pi or 0) + float(pf or 0))
        )

    if not snaps:
        return {**base, "meses": [], "net_premium_mes": [],
                "triangulos": {"incurrido": [], "pagado": [], "num": []},
                "incurrido_actual": 0.0, "ibnr_sugerido": 0.0, "ultimate_sugerido": 0.0}

    sins = {s.id: s for s in db.scalars(select(Siniestro).where(Siniestro.binder_id == binder_id)).all()}
    origen: dict[int, int] = {}
    for sid, lista in snaps.items():
        s = sins.get(sid)
        f = (s.date_opened if s else None) or (s.claim_first_advised if s else None)
        origen[sid] = _mi(f.year, f.month) if f else lista[0][0]

    latest = max(mi for lista in snaps.values() for (mi, _, _) in lista)
    inicio_cand = list(origen.values()) + list(net_mi) + [latest]
    if b.fecha_efecto:
        inicio_cand.append(_mi(b.fecha_efecto.year, b.fecha_efecto.month))
    start = min(inicio_cand)
    meses = list(range(start, latest + 1))

    def val_at(sid: int, C: int):
        v = None
        for mi, inc, paid in snaps[sid]:
            if mi <= C:
                v = (inc, paid)
            else:
                break
        return v

    cohortes: dict[int, list[int]] = {}
    for sid, o in origen.items():
        cohortes.setdefault(o, []).append(sid)

    n = len(meses)
    tri_inc = [[None] * n for _ in range(n)]
    tri_paid = [[None] * n for _ in range(n)]
    tri_num = [[None] * n for _ in range(n)]
    for i, o in enumerate(meses):
        claims_o = cohortes.get(o, [])
        for j in range(i, n):
            C = meses[j]
            sinc = spaid = 0.0
            cnt = 0
            for sid in claims_o:
                v = val_at(sid, C)
                if v is not None:
                    sinc += v[0]; spaid += v[1]; cnt += 1
            tri_inc[i][j] = round(sinc, 2)
            tri_paid[i][j] = round(spaid, 2)
            tri_num[i][j] = cnt

    incurrido_actual = round(sum((val_at(sid, latest) or (0.0, 0.0))[0] for sid in snaps), 2)
    # IBNR por Bornhuetter-Ferguson con el patrón de desarrollo y la ELR del PROGRAMA del binder
    # (los años maduros del programa estiman la cola del año joven). Sin programa → sin IBNR.
    edad = (latest - _mi(b.fecha_efecto.year, b.fecha_efecto.month)) if b.fecha_efecto else len(meses) - 1
    if b.programa_id is not None:
        factores_p, maxlen_p, elr = _programa_factores_elr(db, b.programa_id)
        ibnr, ultimate = _ibnr_bf(b.estado, net_uw, incurrido_actual, edad, factores_p, maxlen_p, elr)
    else:
        ibnr, ultimate = 0.0, incurrido_actual
    return {
        **base,
        "meses": [_periodo_de_mi(m) for m in meses],
        "net_premium_mes": [round(net_mi.get(m, 0.0), 2) for m in meses],
        "triangulos": {"incurrido": tri_inc, "pagado": tri_paid, "num": tri_num},
        "incurrido_actual": incurrido_actual,
        "ibnr_sugerido": ibnr,
        "ultimate_sugerido": ultimate,
    }


@router.get("/binders/{binder_id}/triangulacion")
def triangulacion_binder(
    binder_id: int,
    seccion: int | None = Query(None),
    risk_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    b = db.get(Binder, binder_id)
    if b is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    return _payload_binder(db, b, seccion, risk_code)


@router.get("/binders/{binder_id}/triangulacion/excel")
def triangulacion_binder_excel(
    binder_id: int,
    metrica: str = Query("incurrido"),
    seccion: int | None = Query(None),
    risk_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    b = db.get(Binder, binder_id)
    if b is None:
        raise HTTPException(status_code=404, detail=f"Binder {binder_id} no encontrado")
    if metrica not in METRICAS:
        raise HTTPException(status_code=422, detail=f"Métrica '{metrica}' no válida")
    d = _payload_binder(db, b, seccion, risk_code)
    meses = d["meses"]
    es_num = metrica == "num"
    es_pct = metrica == "pct"
    net_uw = d["net_uw"] or 0
    src = d["triangulos"]["incurrido" if es_pct else metrica]

    def cel(i, j):
        v = src[i][j] if (src and j < len(src[i])) else None
        if v is None:
            return None
        if es_pct:
            return round(v / net_uw * 100, 2) if net_uw else None
        return v

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Triangulación"[:31]
    head = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    # Columnas de valuación: del más reciente (izquierda) al más antiguo (derecha).
    cols_val = list(range(len(meses) - 1, -1, -1))
    cab = ["Mes", "Net to UWs"] + [meses[j] for j in cols_val]
    ws.append(cab)
    for c in ws[1]:
        c.font = head
        c.fill = fill
    for i, m in enumerate(meses):
        fila = [m, d["net_premium_mes"][i]] + [cel(i, j) for j in cols_val]
        ws.append(fila)
    total = ["Total", round(net_uw, 2)]
    for j in cols_val:
        s = sum((cel(i, j) or 0) for i in range(len(meses)))
        total.append(round(s, 2))
    ws.append(total)
    for c in ws[ws.max_row]:
        c.font = head
    fmt = "0.00%" if False else ("#,##0.00" if not es_num else "#,##0")
    for fila in ws.iter_rows(min_row=2, min_col=2):
        for c in fila:
            if isinstance(c.value, (int, float)):
                c.number_format = ("0.00\\%" if es_pct and c.column > 2 else fmt)
    for j in range(1, len(cab) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    scope = d["ambito"].replace(" ", "")
    nombre = f"Triangulacion {b.umr} {METRICAS[metrica]} {scope}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


def _factores_programa(curvas: list[list[float]]) -> tuple[list[float], int]:
    """Factores de desarrollo volumen-ponderados y longitud máxima, sobre el incurrido por antigüedad
    de TODAS las curvas (los años maduros guían a los jóvenes)."""
    tri = [c for c in curvas if c]
    maxlen = max((len(x) for x in tri), default=0)
    factores = []
    for d in range(max(maxlen - 1, 0)):
        num = sum(x[d + 1] for x in tri if len(x) > d + 1)
        den = sum(x[d] for x in tri if len(x) > d + 1)
        factores.append(num / den if den > 0 else 1.0)
    return factores, maxlen


def _pct_desarrollado(factores: list[float], maxlen: int, edad: int) -> float:
    """% del 'ultimate' ya incurrido a la edad `edad` = 1 / (producto de factores de `edad` al final).
    1.0 si ya se llegó al final del desarrollo."""
    if edad >= maxlen - 1:
        return 1.0
    cdf = 1.0
    for k in range(max(edad, 0), maxlen - 1):
        cdf *= factores[k] if factores[k] > 0 else 1.0
    return 1.0 / cdf if cdf > 0 else 1.0


def _programa_factores_elr(db: Session, programa_id: int) -> tuple[list[float], int, float]:
    """Patrón de desarrollo (factores, maxlen) y siniestralidad esperada a priori (ELR) del programa.
    ELR = siniestralidad observada de los años MADUROS (≥66% desarrollados, incurrido ≈ ultimate)."""
    pbs = db.scalars(select(Binder).where(Binder.programa_id == programa_id)).all()
    curvas, nets, incs = [], [], []
    for pb in pbs:
        c = _curva_binder(db, pb)
        curvas.append(c["inc"]); nets.append(c["net"]); incs.append(c["incurrido_actual"])
    factores, maxlen = _factores_programa(curvas)
    num = den = 0.0
    for inc_i, net_i, cur in zip(incs, nets, curvas):
        e = len(cur) - 1
        if e >= 0 and net_i > 0 and _pct_desarrollado(factores, maxlen, e) >= 0.66:
            num += inc_i; den += net_i
    if den > 0:
        elr = num / den
    else:  # ningún año maduro: cae a la LR media del programa
        tot_net = sum(nets)
        elr = (sum(incs) / tot_net) if tot_net > 0 else 0.0
    return factores, maxlen, elr


def _ibnr_bf(estado: str | None, net: float, incurrido_actual: float, edad: int,
             factores: list[float], maxlen: int, elr: float) -> tuple[float, float]:
    """Bornhuetter-Ferguson: IBNR = Net to UWs × ELR × (1 − %desarrollado). Devuelve (ibnr, ultimate).
    Binder en run-off ('Cerrado' exacto) → sin IBNR (ya está todo declarado)."""
    if (estado or "").strip() == "Cerrado":
        return 0.0, round(incurrido_actual, 2)
    pd = _pct_desarrollado(factores, maxlen, edad) if edad >= 0 else 0.0
    ibnr = round(net * elr * (1.0 - pd), 2)
    return ibnr, round(incurrido_actual + ibnr, 2)


def _curva_binder(db: Session, binder: Binder, risk_code: str | None = None) -> dict | None:
    """Desarrollo POR ANTIGÜEDAD de un binder: para cada edad d (meses desde el inicio del binder),
    el incurrido/pagado/nº de sus siniestros valuados a ese mes, y la PRIMA ACUMULADA (Net to UWs
    devengada hasta ese mes) para el ratio de siniestralidad. Filtrable por risk code.
    Incluye GWP our line, net to UWs total e incurrido actual. `start_mi` = mes de inicio."""
    gwp, net = _bases(db, binder.id, risk_code=risk_code)
    net_mes = _net_por_mes(db, binder.id, risk_code=risk_code)  # {mes_index: net del mes}
    q = (
        select(
            ClaimsPresentacion.siniestro_id, ClaimsPresentacion.periodo_ord,
            ClaimsPresentacion.paid_indemnity_acum, ClaimsPresentacion.paid_fees_acum,
            ClaimsPresentacion.reserves_indemnity, ClaimsPresentacion.reserves_fees,
        )
        .join(Siniestro, Siniestro.id == ClaimsPresentacion.siniestro_id)
        .where(ClaimsPresentacion.binder_id == binder.id, ClaimsPresentacion.siniestro_id.is_not(None))
    )
    if risk_code:
        q = q.where(Siniestro.risk_code == risk_code)
    pres = db.execute(q.order_by(ClaimsPresentacion.siniestro_id, ClaimsPresentacion.periodo_ord)).all()
    snaps: dict[int, list[tuple[int, float, float]]] = {}
    for sid, ord_, pi, pf, ri, rf in pres:
        mi = _mi_de_ord(ord_)
        snaps.setdefault(sid, []).append(
            (mi, float(pi or 0) + float(pf or 0) + float(ri or 0) + float(rf or 0), float(pi or 0) + float(pf or 0))
        )

    start = _mi(binder.fecha_efecto.year, binder.fecha_efecto.month) if binder.fecha_efecto else None
    if snaps:
        latest = max(mi for lista in snaps.values() for (mi, _, _) in lista)
        if start is None or start > latest:
            start = min(lista[0][0] for lista in snaps.values())
    elif net_mes:
        # Sin siniestros pero con prima: la curva va del inicio al último mes con prima.
        latest = max(net_mes)
        if start is None or start > latest:
            start = min(net_mes)
    else:
        return {"gwp": gwp, "net": net, "inc": [], "paid": [], "num": [], "prima_acum": [],
                "incurrido_actual": 0.0, "start_mi": start}

    def val_at(sid: int, C: int):
        v = None
        for mi, inc, paid in snaps[sid]:
            if mi <= C:
                v = (inc, paid)
            else:
                break
        return v

    inc, paid, num, prima_acum = [], [], [], []
    acum = 0.0
    for d in range(latest - start + 1):
        C = start + d
        acum += net_mes.get(C, 0.0)
        si = sp = 0.0
        cnt = 0
        for sid in snaps:
            v = val_at(sid, C)
            if v is not None:
                si += v[0]; sp += v[1]; cnt += 1
        inc.append(round(si, 2)); paid.append(round(sp, 2)); num.append(cnt)
        prima_acum.append(round(acum, 2))
    return {"gwp": gwp, "net": net, "inc": inc, "paid": paid, "num": num, "prima_acum": prima_acum,
            "incurrido_actual": inc[-1] if inc else 0.0, "start_mi": start}


@router.get("/programas/{programa_id}/triangulacion")
def triangulacion_programa(
    programa_id: int,
    risk_code: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Triángulo del PROGRAMA: una fila por binder/YOA, columnas = antigüedad (meses desde el
    inicio de cada binder). Los factores de desarrollo se calculan con TODO el programa (los años
    maduros guían la proyección de los jóvenes) y de ahí sale el IBNR de cada año y el total.
    Filtrable por `risk_code` (None = TOTAL)."""
    prog = db.get(Programa, programa_id)
    if prog is None:
        raise HTTPException(status_code=404, detail=f"Programa {programa_id} no encontrado")
    binders = db.scalars(
        select(Binder).where(Binder.programa_id == programa_id).order_by(Binder.yoa, Binder.id)
    ).all()

    # Risk codes disponibles en el programa (para el selector de categoría de la comparación).
    risk_codes: list[str] = []
    if binders:
        rcs = db.execute(
            select(BdxLinea.risk_code).distinct()
            .join(Bdx, Bdx.id == BdxLinea.bdx_id)
            .where(Bdx.binder_id.in_([b.id for b in binders]), Bdx.tipo == "Risk",
                   BdxLinea.risk_code.is_not(None))
        ).all()
        risk_codes = sorted({r[0] for r in rcs if r[0]})

    # Mes de inicio del programa (para etiquetar las filas como Año/Mes): el del binder más antiguo.
    efectos = [b.fecha_efecto for b in binders if b.fecha_efecto]
    mes_inicio = min(efectos).month if efectos else 1

    filas = []  # por binder: meta + curvas
    for b in binders:
        c = _curva_binder(db, b, risk_code=risk_code)
        filas.append({"binder": b, "curva": c})

    # Patrón de desarrollo (volumen-ponderado) con el incurrido de TODOS los binders del programa,
    # y siniestralidad esperada a priori (ELR) de los años maduros → IBNR por Bornhuetter-Ferguson.
    factores, maxlen = _factores_programa([f["curva"]["inc"] for f in filas])
    n_elr = d_elr = 0.0
    for f in filas:
        c = f["curva"]; e = len(c["inc"]) - 1
        if e >= 0 and c["net"] > 0 and _pct_desarrollado(factores, maxlen, e) >= 0.66:
            n_elr += c["incurrido_actual"]; d_elr += c["net"]
    if d_elr > 0:
        elr = n_elr / d_elr
    else:
        tot_net = sum(f["curva"]["net"] for f in filas)
        elr = (sum(f["curva"]["incurrido_actual"] for f in filas) / tot_net) if tot_net > 0 else 0.0

    edades = maxlen  # nº de columnas de antigüedad (0..maxlen-1)
    out_binders, m_inc, m_paid, m_num, m_prima = [], [], [], [], []
    premium_b, netuw_b, inc_act_b, ult_b, ibnr_b = [], [], [], [], []
    for f in filas:
        b, c = f["binder"], f["curva"]
        # IBNR por Bornhuetter-Ferguson (estable en años verdes). "Cerrado" exacto → sin IBNR.
        ibnr, ult = _ibnr_bf(b.estado, c["net"], c["incurrido_actual"], len(c["inc"]) - 1, factores, maxlen, elr)
        out_binders.append({"id": b.id, "umr": b.umr, "agreement": b.agreement_number, "yoa": b.yoa})
        # padding a la longitud común con None
        m_inc.append(c["inc"] + [None] * (edades - len(c["inc"])))
        m_paid.append(c["paid"] + [None] * (edades - len(c["paid"])))
        m_num.append(c["num"] + [None] * (edades - len(c["num"])))
        m_prima.append(c["prima_acum"] + [None] * (edades - len(c["prima_acum"])))
        premium_b.append(c["gwp"]); netuw_b.append(c["net"])
        inc_act_b.append(c["incurrido_actual"]); ult_b.append(ult); ibnr_b.append(ibnr)

    return {
        "programa": prog.nombre,
        "binders": out_binders,
        "max_edad": max(edades - 1, 0),
        "mes_inicio": mes_inicio,
        "risk_codes": risk_codes,
        "risk_code": risk_code,
        "elr_pct": round(elr * 100, 2),
        "triangulos": {"incurrido": m_inc, "pagado": m_paid, "num": m_num},
        "prima_acum_binder": m_prima,
        "premium_binder": premium_b,
        "net_uw_binder": netuw_b,
        "incurrido_binder": inc_act_b,
        "ultimate_binder": ult_b,
        "ibnr_binder": ibnr_b,
        "incurrido_total": round(sum(inc_act_b), 2),
        "ultimate_total": round(sum(ult_b), 2),
        "ibnr_total": round(sum(ibnr_b), 2),
        "premium_total": round(sum(premium_b), 2),
        "net_uw_total": round(sum(netuw_b), 2),
    }
