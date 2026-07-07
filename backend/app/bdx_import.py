"""
Importador de BDX desde SharePoint, controlado y binder a binder.

Lee la lista `Mayrit - <UMR>` del binder (vía app.sharepoint), coacciona los valores al tipo
de cada columna de `BdxLinea` y vuelca las líneas en el **BDX único** del binder (tipo Risk).
Idempotente por `sp_old_id` (el `_OldID` del origen): re-importar actualiza, no duplica.

Normalizaciones (decididas con datos reales):
  - Fechas SIN hora (ya las normaliza el lector a 'aaaa-mm-dd' → se parsean a date).
  - Importes con coma o punto decimal (y punto de miles europeo).
  - Porcentajes: en el origen vienen como fracción (0,8 = 80 %) → ×100.
"""
from __future__ import annotations

import datetime as dt
import io
import re
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation

import openpyxl
from sqlalchemy import Boolean, Date, Integer, Numeric, select
from sqlalchemy.orm import Session

from . import sharepoint
from .models.maestras import Bdx, Binder, BdxLinea

# Porcentajes que en el origen vienen como fracción (0,8) y guardamos como entero (80).
PCT_FIELDS = {
    "written_line_pct",
    "commission_coverholder_pct",
    "brokerage_pct",
    "pct_for_lloyds",
    "tax1_pct",
    "tax2_pct",
    "tax3_pct",
    "tax4_pct",
}


def _num(v) -> Decimal | None:
    """Importe → Decimal. Acepta int/float y texto con coma/punto (miles y decimal europeos)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip()
    if "." in s and "," in s:        # europeo: '.' miles, ',' decimal
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:                    # solo coma → decimal
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _fecha(v) -> dt.date | None:
    """Fecha desde: date/datetime, texto ISO (aaaa-mm-dd) o texto EUROPEO (dd/mm/aaaa o dd-mm-aaaa).
    OJO: algunas plantillas traen las fechas como TEXTO dd/mm/aaaa; si solo se aceptara ISO, el dato
    se perdería en silencio (bug real detectado con el Risk de junio del PI2725)."""
    if not v:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s[:10])   # ISO (con o sin hora)
    except ValueError:
        pass
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)   # dd/mm/aaaa europeo
    if m:
        d, mo, y = (int(x) for x in m.groups())
        if y < 100:
            y += 2000
        try:
            return dt.date(y, mo, d)
        except ValueError:
            return None
    return None


def _int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _coerce(field: str, value, coltype):
    """Coacciona el valor crudo de SharePoint al tipo de la columna del modelo."""
    if isinstance(coltype, Boolean):
        return bool(value)
    if value is None or value == "":
        return None
    if isinstance(coltype, Date):
        return _fecha(value)
    if isinstance(coltype, Integer):
        return _int(value)
    if isinstance(coltype, Numeric):
        n = _num(value)
        if n is None:
            return None
        if field in PCT_FIELDS:
            n = n * 100
        # Cuantizar a la escala de la columna (dinero=2, % =4) → quita el ruido de coma flotante.
        escala = coltype.scale if coltype.scale is not None else 2
        n = n.quantize(Decimal(1).scaleb(-escala), rounding=ROUND_HALF_UP)
        # Fuera de rango para la precisión de la columna → se anula (dato erróneo en origen, p. ej.
        # una fecha metida en una columna de %). Evita que un valor basura tumbe toda la importación.
        if coltype.precision is not None and abs(n) >= Decimal(10) ** (coltype.precision - escala):
            return None
        return n
    return str(value)  # String / Text


def importar(db: Session, binder: Binder) -> dict:
    """Importa (o re-importa) los BDX del binder desde su lista de SharePoint. Devuelve un
    resumen con la conciliación SharePoint ↔ Postgres."""
    list_title = f"Mayrit - {binder.umr}"
    filas = sharepoint.leer_lista_bdx(list_title)
    return importar_filas(db, binder, filas, origen=list_title)


def importar_filas(db: Session, binder: Binder, filas: list[dict], origen: str = "(excel)") -> dict:
    """Inserta/actualiza líneas Risk del binder desde `filas` (dicts con las claves del MAPEO).
    Origen-agnóstico: lo usan tanto la importación de SharePoint como la de Excel. Idempotente
    por `sp_old_id`. Devuelve la conciliación origen ↔ Postgres."""
    list_title = origen
    # BDX único por binder (tipo Risk): se reutiliza si ya existe.
    bdx = db.scalars(
        select(Bdx).where(Bdx.binder_id == binder.id, Bdx.tipo == "Risk")
    ).first()
    if bdx is None:
        bdx = Bdx(binder_id=binder.id, tipo="Risk", estado="Abierto", notas="Importado de SharePoint")
        db.add(bdx)
        db.flush()

    # Líneas existentes indexadas por sp_old_id (idempotencia).
    existentes = {
        l.sp_old_id: l
        for l in db.scalars(select(BdxLinea).where(BdxLinea.bdx_id == bdx.id)).all()
        if l.sp_old_id is not None
    }

    cols = {c.name: c.type for c in BdxLinea.__table__.columns}
    insertadas = actualizadas = sin_old_id = auto_seccion = 0

    # Fallback de sección: risk_code → nº de sección (1-based) SOLO cuando el código pertenece a una
    # única sección declarada del binder. Cubre líneas que llegan con "Section No" vacío en el origen.
    rc2sec: dict[str, int] = {}
    rc_ambiguos: set[str] = set()
    for i, s in enumerate(binder.secciones, start=1):
        for rc in s.risk_codes:
            code = (rc.codigo or "").strip()
            if not code:
                continue
            if code in rc2sec and rc2sec[code] != i:
                rc_ambiguos.add(code)
            rc2sec[code] = i

    for fila in filas:
        datos = {campo: _coerce(campo, fila.get(campo), cols[campo]) for campo in sharepoint.MAPEO}
        if datos.get("section_no") is None:
            code = (datos.get("risk_code") or "").strip()
            if code and code in rc2sec and code not in rc_ambiguos:
                datos["section_no"] = rc2sec[code]
                auto_seccion += 1
        oldid = datos.get("sp_old_id")
        if oldid is None:
            sin_old_id += 1
        linea = existentes.get(oldid) if oldid is not None else None
        if linea is None:
            linea = BdxLinea(bdx_id=bdx.id)
            db.add(linea)
            insertadas += 1
            if oldid is not None:
                existentes[oldid] = linea
        else:
            actualizadas += 1
        for k, v in datos.items():
            setattr(linea, k, v)

    db.flush()

    # Cabecera: rango global de periodos (informativo; el periodo real va por línea).
    starts = [l.reporting_period_start for l in bdx.lineas if l.reporting_period_start]
    ends = [l.reporting_period_end for l in bdx.lineas if l.reporting_period_end]
    bdx.reporting_period_start = min(starts) if starts else None
    bdx.reporting_period_end = max(ends) if ends else None

    db.commit()

    # ── Conciliación SharePoint ↔ Postgres ──
    # GWP de origen redondeado a céntimos por línea (igual que se guarda) para comparar de verdad.
    def _cent(v) -> Decimal:
        d = _num(v)
        return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if d is not None else Decimal(0)

    sp_total = len(filas)
    sp_gwp = sum(_cent(f.get("gross_written_premium")) for f in filas)
    db_lineas = db.scalars(select(BdxLinea).where(BdxLinea.bdx_id == bdx.id)).all()
    db_total = len(db_lineas)
    db_gwp = sum((l.gross_written_premium or Decimal(0)) for l in db_lineas)
    periodos = sorted({str(l.reporting_period_start) for l in db_lineas if l.reporting_period_start})

    return {
        "bdx_id": bdx.id,
        "list_title": list_title,
        "insertadas": insertadas,
        "actualizadas": actualizadas,
        "sin_old_id": sin_old_id,
        "auto_seccion": auto_seccion,
        "periodos": periodos,
        "conciliacion": {
            "lineas_sharepoint": sp_total,
            "lineas_postgres": db_total,
            "lineas_ok": sp_total == db_total,
            "gwp_sharepoint": float(round(sp_gwp, 2)),
            "gwp_postgres": float(round(db_gwp, 2)),
            "gwp_ok": abs(sp_gwp - db_gwp) < Decimal("0.01"),
        },
    }


# ── Importación de Risk desde un Excel subido (sin SharePoint) ───────────────────────────────────
# Reaprovecha el MAPEO (columna interna → nombre de columna del bordereau) y la coacción de tipos.
_CLAVES_FILA = ("certificate_ref", "total_gwp_our_line", "gross_written_premium", "section_no")


def _norm_col(h: str | None) -> str:
    """Normaliza un encabezado para comparar: ignora paréntesis y mayúsculas/espacios. Así
    'Sum insured (Our Line)' casa con el alias 'Sum insured Our Line' (un desajuste por paréntesis
    no debe tirar el dato en silencio)."""
    return sharepoint._norm((h or "").replace("(", " ").replace(")", " ")).lower()


def _resolver_columnas(headers: list[str]) -> dict[str, int]:
    """Para cada campo del MAPEO, el índice de la columna del Excel cuyo título casa (con alias)."""
    norm = {}
    for i, h in enumerate(headers):
        k = _norm_col(h)
        if k:
            norm.setdefault(k, i)
    out: dict[str, int] = {}
    for campo, alias in sharepoint.MAPEO.items():
        for nm in (alias if isinstance(alias, list) else [alias]):
            i = norm.get(_norm_col(nm))
            if i is not None:
                out[campo] = i
                break
    return out


def parse_risk_excel(content: bytes, hoja: str | None = None) -> tuple[list[dict], dict]:
    """Lee el Excel (la hoja indicada, o la primera), detecta la cabecera y devuelve (filas, meta). Cada
    fila es un dict {campo_interno: valor_crudo} con las columnas reconocidas del MAPEO."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = hoja if (hoja and hoja in wb.sheetnames) else wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # Cabecera = la fila con más celdas de texto entre las primeras 15.
    best_i, best_n = 0, -1
    for i in range(min(15, len(rows))):
        n = sum(1 for v in rows[i] if isinstance(v, str) and v.strip())
        if n > best_n:
            best_i, best_n = i, n
    headers = [str(v).strip() if v is not None else "" for v in (rows[best_i] if rows else [])]
    colmap = _resolver_columnas(headers)
    filas: list[dict] = []
    for r in rows[best_i + 1:]:
        if not any(v not in (None, "") for v in r):
            continue
        fila = {campo: (r[idx] if idx < len(r) else None) for campo, idx in colmap.items()}
        if not any(fila.get(k) not in (None, "") for k in _CLAVES_FILA):
            continue
        # Captura COMPLETA de la fila: TODAS las celdas con cabecera (no solo las reconocidas), para
        # no perder nada. Las no mapeadas se guardarán tal cual en `extra`.
        fila["_raw"] = {
            headers[i]: r[i]
            for i in range(len(headers))
            if headers[i] and i < len(r) and r[i] not in (None, "")
        }
        filas.append(fila)
    usados = set(colmap.values())
    meta = {
        "n_filas": len(filas),
        "cabecera_fila": best_i + 1,
        "mapeadas": {campo: headers[i] for campo, i in colmap.items()},
        "sin_mapear": [headers[i] for i in range(len(headers)) if headers[i] and i not in usados],
        "hojas": list(wb.sheetnames),
        "hoja": sheet,
    }
    return filas, meta


def _rc2sec(binder: Binder) -> tuple[dict[str, int], set[str]]:
    """Mapa risk code → nº de sección (1..n) del binder, y el conjunto de codes ambiguos (en >1 sección)."""
    rc2sec: dict[str, int] = {}
    rc_amb: set[str] = set()
    for i, s in enumerate(binder.secciones, start=1):
        for rc in s.risk_codes:
            code = (rc.codigo or "").strip()
            if not code:
                continue
            if code in rc2sec and rc2sec[code] != i:
                rc_amb.add(code)
            rc2sec[code] = i
    return rc2sec, rc_amb


def _seccion_de(datos: dict, rc2sec: dict[str, int], rc_amb: set[str]) -> int | None:
    """Sección de la línea: la declarada, o la deducida por su risk code (si es inequívoco)."""
    if datos.get("section_no") is not None:
        return datos["section_no"]
    code = (datos.get("risk_code") or "").strip()
    if code and code in rc2sec and code not in rc_amb:
        return rc2sec[code]
    return None


def _coerce_fila(cols: dict, fila: dict) -> dict:
    return {campo: _coerce(campo, fila.get(campo), cols[campo])
            for campo in sharepoint.MAPEO if campo in cols and campo != "sp_old_id"}


def _json_safe(v):
    """Convierte un valor de celda a algo serializable en JSONB (para `extra`)."""
    if isinstance(v, (dt.date, dt.datetime)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def _extra_no_mapeadas(fila: dict, meta_mapeadas: dict) -> dict | None:
    """De la captura completa de la fila (`_raw`), las columnas que NO han ido a un campo estructurado.
    Así toda la información del Excel queda guardada (las reconocidas en sus campos, el resto en `extra`)."""
    raw = fila.get("_raw") or {}
    cabeceras_mapeadas = set(meta_mapeadas.values())
    extra = {h: _json_safe(v) for h, v in raw.items() if h not in cabeceras_mapeadas}
    return extra or None


def preview_risk_excel(db: Session, binder: Binder, content: bytes, hoja: str | None = None) -> dict:
    """Coacciona las filas y devuelve un resumen (sin escribir): hojas, nº líneas, periodos, totales,
    reparto por sección (con asignación por risk code) y los meses que ya están cargados en el Risk."""
    filas, meta = parse_risk_excel(content, hoja)
    cols = {c.name: c.type for c in BdxLinea.__table__.columns}
    coerced = [_coerce_fila(cols, f) for f in filas]
    periodos = sorted({d["reporting_period_start"].strftime("%Y-%m") for d in coerced if d.get("reporting_period_start")})
    tot_our = sum((d.get("total_gwp_our_line") or Decimal(0)) for d in coerced)
    tot_100 = sum((d.get("gross_written_premium") or Decimal(0)) for d in coerced)
    tot_traspasar = sum((d.get("brokerage_amount") or Decimal(0)) for d in coerced)        # nuestra comisión
    tot_liquidar = sum((d.get("final_net_premium_uw") or Decimal(0)) for d in coerced)      # neto al UW

    # Reparto por sección: la declarada o la deducida por risk code; aviso de las que no casan.
    rc2sec, rc_amb = _rc2sec(binder)
    por_seccion: dict[str, int] = {}
    auto_seccion = sin_seccion = 0
    for d in coerced:
        sec = _seccion_de(d, rc2sec, rc_amb)
        if sec is None:
            sin_seccion += 1
            continue
        if d.get("section_no") is None:
            auto_seccion += 1
        por_seccion[str(sec)] = por_seccion.get(str(sec), 0) + 1

    # Meses (reporting) que ya están en el Risk del binder: al aplicar se omitirán para no recargarlos.
    bdx = db.scalars(select(Bdx).where(Bdx.binder_id == binder.id, Bdx.tipo == "Risk")).first()
    ya = set()
    if bdx is not None:
        ya = {l.reporting_period_start.strftime("%Y-%m") for l in bdx.lineas if l.reporting_period_start}
    periodos_ya_cargados = sorted(p for p in periodos if p in ya)

    muestra = [{
        "certificado": d.get("certificate_ref"), "asegurado": d.get("insured_name"),
        "section_no": _seccion_de(d, rc2sec, rc_amb), "risk_code": d.get("risk_code"),
        "reporting": d["reporting_period_start"].isoformat() if d.get("reporting_period_start") else None,
        "gwp_our_line": float(d["total_gwp_our_line"]) if d.get("total_gwp_our_line") is not None else None,
        "comision_pct": float((d.get("commission_coverholder_pct") or 0) + (d.get("brokerage_pct") or 0)),
        "prima_traspasar": float(d["brokerage_amount"]) if d.get("brokerage_amount") is not None else None,
        "liquidar": float(d["final_net_premium_uw"]) if d.get("final_net_premium_uw") is not None else None,
    } for d in coerced]   # todas las líneas (no solo una muestra), con su fila de totales en el modal
    return {
        "hojas": meta["hojas"], "hoja": meta["hoja"],
        "n_lineas": meta["n_filas"], "periodos": periodos,
        "total_gwp_our_line": float(round(tot_our, 2)), "total_gwp_100": float(round(tot_100, 2)),
        "total_prima_traspasar": float(round(tot_traspasar, 2)), "total_liquidar": float(round(tot_liquidar, 2)),
        "mapeadas": meta["mapeadas"], "sin_mapear": meta["sin_mapear"], "muestra": muestra,
        "por_seccion": por_seccion, "auto_seccion": auto_seccion, "sin_seccion": sin_seccion,
        "periodos_ya_cargados": periodos_ya_cargados,
        # Líneas sin periodo reconocible: si es > 0, el import ABORTA (guardarraíl). El front lo avisa.
        "sin_periodo": sum(1 for d in coerced if not d.get("reporting_period_start")),
    }


def importar_risk_excel(db: Session, binder: Binder, content: bytes, hoja: str | None = None) -> dict:
    """Añade TODAS las líneas del Excel al BDX Risk del binder (sin dedup por línea: se conservan los
    duplicados legítimos de pagos fraccionados). Asigna la sección por risk code cuando falta. La única
    protección es a nivel de mes: si un Reporting ya estaba cargado en el Risk, ese mes se omite entero
    para no recargarlo por error."""
    filas, meta = parse_risk_excel(content, hoja)
    mapeadas = meta.get("mapeadas", {})
    cols = {c.name: c.type for c in BdxLinea.__table__.columns}
    bdx = db.scalars(select(Bdx).where(Bdx.binder_id == binder.id, Bdx.tipo == "Risk")).first()
    if bdx is None:
        bdx = Bdx(binder_id=binder.id, tipo="Risk", estado="Abierto", notas="Importado de Excel")
        db.add(bdx)
        db.flush()

    rc2sec, rc_amb = _rc2sec(binder)

    # Meses (reporting) ya presentes en el Risk ANTES de esta subida → se omiten enteros.
    periodos_existentes = {
        l.reporting_period_start.strftime("%Y-%m")
        for l in db.scalars(select(BdxLinea).where(BdxLinea.bdx_id == bdx.id)).all()
        if l.reporting_period_start
    }

    # GUARDARRAÍL (crítico): NO importar en silencio líneas sin periodo. Un Risk BDX SIEMPRE tiene
    # 'Reporting Period Start Date'; si alguna línea queda sin periodo reconocible (columna con otro
    # encabezado, o fechas en un formato no soportado), se ABORTA la importación entera para que el
    # usuario lo revise. Sin periodo no se agrupa ni genera recibo, y encima se salta la protección
    # de 'mes ya cargado' → duplicados. (Bug detectado con el Risk de junio del PI2725.)
    coerced = [_coerce_fila(cols, f) for f in filas]
    sin_periodo = sum(1 for d in coerced if not d.get("reporting_period_start"))
    if sin_periodo:
        raise ValueError(
            f"{sin_periodo} de {len(filas)} líneas no tienen 'Reporting Period Start Date' reconocible. "
            f"Revisa que la hoja seleccionada tenga esa columna y que las fechas sean válidas. "
            f"No se ha importado NADA (para no dejar líneas sin periodo)."
        )

    insertadas = omitidas_periodo = auto_seccion = sin_seccion = 0
    periodos_omitidos: set[str] = set()
    for fila, datos in zip(filas, coerced):
        sec = _seccion_de(datos, rc2sec, rc_amb)
        if datos.get("section_no") is None and sec is not None:
            datos["section_no"] = sec
            auto_seccion += 1
        if sec is None:
            sin_seccion += 1
        rp = datos.get("reporting_period_start")
        pm = rp.strftime("%Y-%m") if rp else None
        if pm and pm in periodos_existentes:
            omitidas_periodo += 1
            periodos_omitidos.add(pm)
            continue
        linea = BdxLinea(bdx_id=bdx.id)
        for k, v in datos.items():
            setattr(linea, k, v)
        linea.extra = _extra_no_mapeadas(fila, mapeadas)  # todo lo no estructurado, sin perder nada
        db.add(linea)
        insertadas += 1

    db.flush()
    starts = [l.reporting_period_start for l in bdx.lineas if l.reporting_period_start]
    ends = [l.reporting_period_end for l in bdx.lineas if l.reporting_period_end]
    bdx.reporting_period_start = min(starts) if starts else None
    bdx.reporting_period_end = max(ends) if ends else None
    db.commit()
    periodos = sorted({str(p) for p in (l.reporting_period_start for l in bdx.lineas) if p})
    return {"bdx_id": bdx.id, "insertadas": insertadas, "omitidas_periodo": omitidas_periodo,
            "periodos_omitidos": sorted(periodos_omitidos), "auto_seccion": auto_seccion,
            "sin_seccion": sin_seccion, "total_lineas": len(bdx.lineas), "periodos": periodos}
