"""
Importa el HISTÓRICO de presentaciones del Claims BDX de un binder desde el Excel de bordereaux
(una pestaña por mes, formato Claims Bordereau de Lloyd's de 32 columnas).

- El periodo de cada pestaña se toma de la celda "Reporting Period (End Date)" (NO del nombre de
  la pestaña, que va desfasado).
- Crea una fila en `claims_presentaciones` por (binder, periodo, siniestro) con el snapshot
  congelado (fila_json) y los acumulados; y BLOQUEA cada mes (BdxBloqueo tipo='claims').
- Empareja el siniestro por Certificate (y, en su defecto, por Claim Reference).
- Idempotente: por cada periodo hace DELETE+INSERT. DRY-RUN por defecto (--apply para escribir).

Uso:  py -m tools.migrar_claims_presentaciones --excel "RUTA.xlsx" --binder-id 12 [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
from collections import Counter
from decimal import Decimal

import openpyxl
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.maestras import BdxBloqueo, Binder, ClaimsPresentacion, Siniestro
from app.routers.claims_bdx import HEADERS

H_FECHA = {
    "Binding authority or coverholder appointment agreement inception date",
    "Binding authority or coverholder appointment agreement expiry date",
    "Reporting Period (End Date)", "Risk Inception Date", "Risk Expiry Date",
    "Date Claim First Advised/Date Claim Made", "Date Claim Opened", "Date Closed",
}


_MESES = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july", "august",
     "september", "october", "november", "december"], start=1)}
_MESES.update({m: i for i, m in enumerate(
    ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
     "septiembre", "octubre", "noviembre", "diciembre"], start=1)})


def _periodo_de_hoja(nombre: str, anio_defecto: int | None = None):
    """'September 2020' -> (2020, 9). Entiende meses en inglés y español. Si la hoja es solo el
    mes (sin año, p. ej. 'March'), usa `anio_defecto`. Devuelve None si no reconoce el mes."""
    partes = str(nombre).strip().split()
    if not partes:
        return None
    mes = _MESES.get(partes[0].lower())
    if mes is None:
        return None
    if partes[-1].isdigit() and len(partes[-1]) == 4:
        return int(partes[-1]), mes
    if anio_defecto:
        return int(anio_defecto), mes
    return None


def _normh(s) -> str:
    """Normaliza una cabecera: colapsa saltos de línea/espacios múltiples a un solo espacio.
    Los Excel de bordereau traen las cabeceras con \\n dentro de la celda (p. ej.
    'Certificate\\nReference'); así casan con los nombres canónicos de HEADERS."""
    return " ".join(str(s).split()) if s is not None else ""


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _safe(h: str, v):
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    if h in H_FECHA:
        return str(v)[:10]
    return v


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--binder-id", type=int)
    ap.add_argument("--agreement")
    ap.add_argument("--periodo-desde", choices=["celda", "hoja"], default="celda",
                    help="De dónde sacar el periodo: 'celda' (Reporting Period End Date, por defecto) "
                         "o 'hoja' (nombre de la pestaña, p. ej. 'September 2020'). Usa 'hoja' cuando "
                         "la celda de periodo está sin mantener (repetida en muchas pestañas).")
    ap.add_argument("--periodo-override", default="",
                    help="Corrige periodos de hojas con el nombre mal escrito. Formato: "
                         "'Nombre Hoja=AAAA-MM,Otra Hoja=AAAA-MM'. Útil para typos (p. ej. 'November 20223').")
    ap.add_argument("--anio-defecto", type=int, default=None,
                    help="Año a usar para hojas cuyo nombre es solo el mes (p. ej. 'March' sin año).")
    ap.add_argument("--omitir-sin-casar", action="store_true",
                    help="Omite las filas de claim que no casan con ningún siniestro del binder "
                         "(en vez de volcarlas con siniestro_id nulo).")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    overrides: dict[str, str] = {}
    for par in args.periodo_override.split(","):
        if "=" in par:
            k, v = par.split("=", 1)
            overrides[k.strip()] = v.strip()

    db = SessionLocal()
    b = db.get(Binder, args.binder_id) if args.binder_id else \
        db.scalar(select(Binder).where(Binder.agreement_number == args.agreement))
    if b is None:
        print("Binder no encontrado.")
        return

    # Mapa de siniestros del binder para casar. El certificate puede estar DUPLICADO entre
    # siniestros (varios claims del mismo certificado), así que casamos primero por el par
    # (certificate, reference) y, como respaldo, por reference o certificate SOLO si son únicos.
    sins = db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all()
    cert_cnt = Counter((s.certificate or "").strip() for s in sins if s.certificate)
    ref_cnt = Counter((s.reference or "").strip() for s in sins if s.reference)
    por_par = {((s.certificate or "").strip(), (s.reference or "").strip()): s.id for s in sins}
    por_cert = {(s.certificate or "").strip(): s.id for s in sins
                if s.certificate and cert_cnt[(s.certificate or "").strip()] == 1}
    por_ref = {(s.reference or "").strip(): s.id for s in sins
               if s.reference and ref_cnt[(s.reference or "").strip()] == 1}

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    # periodo -> lista de (fila_dict, meta)
    por_periodo: dict[str, list] = {}
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        ix = {_normh(c.value): i for i, c in enumerate(ws[1]) if c.value}
        if "Certificate Reference" not in ix or "Reporting Period (End Date)" not in ix:
            continue
        filas = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[ix["Certificate Reference"]]]
        # Periodo. Las hojas NIL (cabecera pero sin filas) no tienen celda de periodo, así que se
        # toma del nombre de la hoja (igual que --periodo-desde hoja).
        if hoja in overrides:
            anio, mes = (int(x) for x in overrides[hoja].split("-"))
        elif args.periodo_desde == "hoja" or not filas:
            ay = _periodo_de_hoja(hoja, args.anio_defecto)
            if ay is None:
                if filas:
                    print(f"  [!] hoja {hoja!r}: no se reconoce el mes/año, omitida")
                continue
            anio, mes = ay
        else:
            rp = filas[0][ix["Reporting Period (End Date)"]]
            if not isinstance(rp, (dt.datetime, dt.date)):
                print(f"  [!] hoja {hoja!r}: sin Reporting Period válido, omitida")
                continue
            anio, mes = rp.year, rp.month
        periodo = f"{anio:04d}-{mes:02d}"
        po = anio * 100 + mes
        if not filas:
            por_periodo[periodo] = []   # mes presentado en blanco (NIL)
            continue
        registros = []
        for r in filas:
            def g(h):
                return r[ix[h]] if h in ix else None
            cert = str(g("Certificate Reference") or "").strip()
            ref = str(g("Claim Reference / Number") or "").strip()
            sid = por_par.get((cert, ref)) or por_ref.get(ref) or por_cert.get(cert)
            if sid is None and args.omitir_sin_casar:
                continue
            fila = {h: _safe(h, g(h)) for h in HEADERS}
            meta = {
                "siniestro_id": sid,
                "paid_indemnity_acum": _num(g("Previously Paid - Indemnity")) + _num(g("Paid this month - Indemnity")),
                "paid_fees_acum": _num(g("Previously Paid - Fees")) + _num(g("Paid this month - Fees")),
                "to_pay_indemnity": _num(g("Paid this month - Indemnity")),
                "to_pay_fees": _num(g("Paid this month - Fees")),
                "reserves_indemnity": _num(g("Reserve - Indemnity")),
                "reserves_fees": _num(g("Reserve - Fees")),
                "status": g("Claim Status"),
            }
            registros.append((fila, meta, po))
        por_periodo[periodo] = registros  # un periodo = una hoja (la última gana si se repite)

    periodos = sorted(por_periodo)
    print(f"== Histórico Claims BDX — binder {b.umr} (DRY-RUN={'NO' if args.apply else 'SÍ'}) ==")
    print(f"Periodos detectados: {len(periodos)}")
    sin_match = 0
    n_nil = 0
    for p in periodos:
        regs = por_periodo[p]
        if not regs:
            n_nil += 1
            print(f"  {p}: NIL (presentado en blanco)")
            continue
        nm = sum(1 for _, m, _ in regs if m["siniestro_id"] is None)
        sin_match += nm
        print(f"  {p}: {len(regs)} siniestro(s)" + (f"  [!] {nm} sin casar" if nm else ""))
    print(f"Filas sin casar con un siniestro del binder: {sin_match} · meses NIL: {n_nil}")

    if not args.apply:
        db.close()
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply para volcar el histórico.")
        return

    total = 0
    for p in periodos:
        regs = por_periodo[p]
        po = int(p[:4]) * 100 + int(p[5:7])
        db.execute(delete(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == p))
        if not regs:
            # Mes presentado en blanco (NIL): placeholder sin siniestro.
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=p, periodo_ord=po, siniestro_id=None,
                paid_indemnity_acum=Decimal("0"), paid_fees_acum=Decimal("0"),
                to_pay_indemnity=Decimal("0"), to_pay_fees=Decimal("0"),
                reserves_indemnity=Decimal("0"), reserves_fees=Decimal("0"), status="Nil",
                fila_json=json.dumps({"nil": True, "report": f"{p} — presentado en blanco"}, ensure_ascii=False),
                fecha_presentacion=None, usuario="histórico-nil"))
            total += 1
            if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == p)):
                db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=p))
            continue
        for fila, meta, _ in regs:
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=p, periodo_ord=po, siniestro_id=meta["siniestro_id"],
                paid_indemnity_acum=Decimal(str(meta["paid_indemnity_acum"])),
                paid_fees_acum=Decimal(str(meta["paid_fees_acum"])),
                to_pay_indemnity=Decimal(str(meta["to_pay_indemnity"])),
                to_pay_fees=Decimal(str(meta["to_pay_fees"])),
                reserves_indemnity=Decimal(str(meta["reserves_indemnity"])),
                reserves_fees=Decimal(str(meta["reserves_fees"])),
                status=(str(meta["status"]) if meta["status"] is not None else None),
                fila_json=json.dumps(fila, ensure_ascii=False, default=str),
                fecha_presentacion=None, usuario="histórico",
            ))
            total += 1
        # Bloquear el mes (presentado).
        if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == p)):
            db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=p))
    db.commit()
    print(f"\nAPLICADO: {total} filas de presentación en {len(periodos)} periodos. Meses bloqueados.")
    db.close()


if __name__ == "__main__":
    main()
