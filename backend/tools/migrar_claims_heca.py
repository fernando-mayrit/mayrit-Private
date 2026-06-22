"""
Importa el histórico de presentaciones de Claims del formato CRAWFORD/AXIS organizado en
**una carpeta por mes** (caso Heca AXIS PI13 YOA2022): cada subcarpeta '<N>. <Mes [Año]>' tiene
un Excel BDX combinado (o, en algunos meses, sólo ficheros por sección E3/E5/E7/E9 que se unen).

Particularidades resueltas (vistas con datos reales):
  - Dos sub-formatos de columnas que conviven: "Crawford e-Claims" (col 'Claim Number', estado
    'File Status') y "AxisCMS Monthly" (col 'Claim Reference / Number', estado 'Claim Status',
    filas en blanco intercaladas y VARIOS binders en hojas separadas). Se mapea por NOMBRE de
    cabecera con alias y se FILTRA por Agreement = agreement del binder.
  - Cabecera en la fila 3 (1-idx), no en la 1.
  - Ficheros .xlsx con workbook.xml corrupto que openpyxl no abre (0 hojas) -> se REPARAN con
    Excel COM (abrir + guardar) antes de leer.
  - Periodo: de la celda 'Reporting Period (End Date)'; respaldo, del número de carpeta.
  - Dedup por referencia de claim dentro de cada mes (evita doble conteo Total+secciones).
  - Empareja el siniestro por (certificate, reference) y respaldos únicos; con --crear-siniestros
    da de alta los que no existan (a partir de la última fila vista del claim).
  - Bloquea cada mes presentado (BdxBloqueo tipo='claims').

DRY-RUN por defecto. Uso:
  py -m tools.migrar_claims_heca --carpeta "RUTA\\Claims" --binder-id 40 [--crear-siniestros] [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import time
from collections import Counter
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.maestras import BdxBloqueo, Binder, ClaimsPresentacion, Siniestro
from app.routers.claims_bdx import HEADERS

# Campos de HEADERS que son fechas (para serializarlos como 'aaaa-mm-dd' en el fila_json).
H_FECHA = {
    "Binding authority or coverholder appointment agreement inception date",
    "Binding authority or coverholder appointment agreement expiry date",
    "Reporting Period (End Date)", "Risk Inception Date", "Risk Expiry Date",
    "Date Claim First Advised/Date Claim Made", "Date Claim Opened", "Date Closed",
}

# Alias: nombre canónico (HEADERS / lógicos) -> otros nombres con que aparece en el origen.
ALIAS = {
    "Coverholder Name": ["Cover Holder"],
    "Claim Reference / Number": ["Claim Number", "Claimnumber"],
    "Insured Full Name or Company Name": ["Insured name"],
    "Loss Description": ["Loss Description (max. 1000 characters)"],
    "Date Claim First Advised/Date Claim Made": ["Date Claim Made"],
    "Claim Status": ["File Status"],
    "Denial (Y/N)": ["Denial"],
    "Claimant Name": ["Claimant"],
    "Date Closed": ["Date File Closed / Finalized"],
    "UCR": ["Claimnumber", "Claim Reference / Number", "Claim Number"],
    "Agreement No.": ["Agreement No", "Agreement Number"],
    "Lloyd's Risk Code": ["Lloyd's Risk Code"],
    "Unique Market Reference (UMR)": ["Unit Market Reference (UMR)"],
    # GES40 nombra el pagado del mes como "To pay this month indemnity/fees".
    "Paid this month - Indemnity": ["To pay this month indemnity"],
    "Paid this month - Fees": ["To pay this month fees"],
}


_MES = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july", "august",
     "september", "october", "november", "december"], start=1)}
_MES.update({m: i for i, m in enumerate(
    ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
     "septiembre", "octubre", "noviembre", "diciembre"], start=1)})


def _periodo_carpeta(nombre: str, anio_defecto: int | None):
    """Saca (año, mes) del nombre de una carpeta de mes, p. ej. '42. Agosto 2025 NO HAY' ->
    (2025, 8) o '1. Enero' -> (anio_defecto, 1). Reconoce meses en inglés y español y busca el
    año en cualquier posición. Devuelve None si no reconoce el mes."""
    s = re.sub(r"^\s*\d+\.\s*", "", str(nombre)).strip()
    toks = s.split()
    mes = _MES.get(toks[0].lower()) if toks else None
    if mes is None:
        return None
    m = re.search(r"(?:19|20)\d{2}", s)
    anio = int(m.group()) if m else anio_defecto
    return (anio, mes) if anio else None


def _normh(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


def _key(s) -> str:
    # Cabecera normalizada: minúsculas y guiones/barras/guiones-bajos -> espacio (colapsado). Así
    # casan las variantes según el origen: "Reserve - Indemnity" = "Reserve Indemnity",
    # "Previously Paid - Indemnity" = "Previously paid indemnity", "Claim Reference / Number" =
    # "Claim Reference Number", "TOTAL_INCURRED_INDEMNITY" = "total incurred indemnity".
    return re.sub(r"\s+", " ", re.sub(r"[-/_]+", " ", _normh(s).lower())).strip()


def _tok(s) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(s or "")).upper()


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _dec(v) -> Decimal:
    return Decimal(str(_num(v))).quantize(Decimal("0.01"), ROUND_HALF_UP)


def _fecha(v) -> dt.date | None:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if not v:
        return None
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def _safe(h: str, v):
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    if h in H_FECHA:
        return str(v)[:10]
    return v


# ── Excel COM para reparar .xlsx con workbook.xml corrupto ───────────────────────────────────
_XL = {"app": None}


def _excel():
    if _XL["app"] is None:
        import pythoncom
        import win32com.client as w
        pythoncom.CoInitialize()
        xl = w.gencache.EnsureDispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        _XL["app"] = xl
    return _XL["app"]


def _retry(fn, n=12):
    for _ in range(n):
        try:
            return fn()
        except Exception:
            time.sleep(0.5)
    return fn()


def _reparar(path: str) -> str:
    """Abre el xlsx roto con Excel y lo reguarda en un temporal sano. Devuelve la ruta temporal."""
    xl = _excel()
    tmp = os.path.join(os.path.expanduser("~"), ".mayrit", "tmp")
    os.makedirs(tmp, exist_ok=True)
    out = os.path.join(tmp, "heca_" + _tok(os.path.basename(path))[-24:] + ".xlsx")
    wb = _retry(lambda: xl.Workbooks.Open(os.path.abspath(path)))
    _retry(lambda: wb.SaveAs(out, FileFormat=51))
    wb.Close(False)
    return out


def _abrir(path: str):
    """openpyxl read_only; si el workbook viene corrupto (0 hojas) lo repara con Excel COM."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if not wb.sheetnames:
        wb.close()
        wb = openpyxl.load_workbook(_reparar(path), read_only=True, data_only=True)
    return wb


# ── Lectura de una hoja: localiza cabecera y devuelve filas (dict canónico) del binder ────────
def _idx_map(hdr):
    return {_key(h): i for i, h in enumerate(hdr) if h is not None}


def _resolver(nidx, canonico):
    for nm in [canonico] + ALIAS.get(canonico, []):
        i = nidx.get(_key(nm))
        if i is not None:
            return i
    return None


def _leer_hoja(ws, agr_tok):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hr = None
    for i in range(min(8, len(rows))):
        celdas = {_key(c) for c in rows[i] if c is not None}
        tiene_ref = bool(celdas & {"claim number", "claimnumber", "claim reference number"}) or \
            any(k.startswith("claim reference") for k in celdas)
        tiene_ctx = any(k.startswith("agreement no") for k in celdas) or "reporting period (end date)" in celdas
        if tiene_ref and tiene_ctx:
            hr = i
            break
    if hr is None:
        return []
    nidx = _idx_map(rows[hr])
    i_ref = _resolver(nidx, "Claim Reference / Number")
    if i_ref is None:
        # Respaldo: cualquier columna que empiece por "claim reference" (variantes por TPA, p. ej.
        # "Claim Reference AULÉS"). Se evita "...status" y similares quedándonos en la primera.
        for k, i in sorted(nidx.items(), key=lambda kv: kv[1]):
            if k.startswith("claim reference"):
                i_ref = i
                break
    i_agr = _resolver(nidx, "Agreement No.")
    i_umr = _resolver(nidx, "Unique Market Reference (UMR)")
    i_ucr = nidx.get(_key("UCR"))
    out = []
    for r in rows[hr + 1:]:
        if i_ref is None or i_ref >= len(r):
            continue
        ref = str(r[i_ref]).strip() if r[i_ref] is not None else ""
        if not ref or _key(ref).startswith("claim"):
            continue  # fila sin referencia (en blanco o basura): no se puede casar -> se omite
        agr = _tok(r[i_agr]) if (i_agr is not None and i_agr < len(r)) else ""
        umr = _tok(r[i_umr]) if (i_umr is not None and i_umr < len(r)) else ""
        ucr = _tok(r[i_ucr]) if (i_ucr is not None and i_ucr < len(r)) else ""
        # El binder se identifica por Agreement, UMR o UCR (según el origen trae uno u otro).
        if agr_tok not in agr and agr_tok not in umr and agr_tok not in ucr:
            continue  # otra agencia/binder en el mismo fichero
        def g(canonico, _r=r, _ni=nidx):
            i = _resolver(_ni, canonico)
            return _r[i] if (i is not None and i < len(_r)) else None
        out.append((ref, g))
    return out


# ── Selección de ficheros de un mes ──────────────────────────────────────────────────────────
def _es_seccion(fn):
    return re.search(r"(^|[ _])E\d", fn) is not None


def _ficheros_mes(carpeta, fuente="ges40"):
    """Ficheros Excel a leer de una carpeta de mes, según la fuente:
      - 'ges40' (formato Heca/AXIS): el combinado del mes + las secciones E3/E5/E7/E9 (se unen).
      - 'aules': se IGNORA el común 'YOA*' (resumen) y se usan los ficheros por risk code
        (E7/E9/D3/CY…), que se agrupan en el snapshot del mes."""
    xs = [f for f in os.listdir(carpeta) if f.lower().endswith((".xlsx", ".xls"))
          and not re.search(r"timesheet|invoice|triangul|template|^~\$", f, re.I)]
    if fuente == "aules":
        return sorted([f for f in xs if not re.match(r"\s*YOA", f, re.I)]), 0
    combinados = [f for f in xs if not _es_seccion(f)]
    secciones = [f for f in xs if _es_seccion(f)]
    # Se leen el combinado Y las secciones (la lectura une y deduplica por referencia). El combinado
    # a veces viene incompleto (solo una sección); las secciones E3/E5/E7/E9 son el desglose fiable,
    # así que unirlos recupera el roster completo sin doble conteo.
    elegidos = []
    if combinados:
        # Preferir la versión 'amended' (corregida) si hay más de un combinado.
        amended = [f for f in combinados if "amend" in f.lower()]
        elegidos.append(amended[0] if amended else sorted(combinados)[0])
    elegidos += sorted(secciones)
    return elegidos, len(combinados)


def leer_carpeta(carpeta, agr_tok, anio_defecto=None, overrides=None, fuente="ges40"):
    """Lee las subcarpetas-mes de `carpeta` y devuelve [{per, po, claims:{ref:getter}, carpeta}]
    (una entrada por carpeta; la fusión por periodo se hace aparte con fusionar())."""
    overrides = overrides or {}
    subs = sorted(
        [d for d in os.listdir(carpeta) if os.path.isdir(os.path.join(carpeta, d))
         and re.match(r"\s*\d+\.", d)],
        key=lambda d: int(d.split(".")[0]),
    )
    meses = []
    for d in subs:
        cp = os.path.join(carpeta, d)
        ficheros, _ = _ficheros_mes(cp, fuente)
        claims = {}
        per_celda = None
        for f in ficheros:
            try:
                wb = _abrir(os.path.join(cp, f))
            except Exception as e:
                print(f"  [!] {d}/{f}: no se pudo abrir ({type(e).__name__})")
                continue
            for sn in wb.sheetnames:
                for ref, g in _leer_hoja(wb[sn], agr_tok):
                    claims[ref] = g  # dedup por ref; la última gana
                    if per_celda is None:
                        rp = _fecha(g("Reporting Period (End Date)"))
                        if rp:
                            per_celda = rp
            wb.close()
        ov = overrides.get(d) or overrides.get(d.split(".")[0].strip())
        if ov:
            anio, mes = int(ov[:4]), int(ov[5:7])
        elif per_celda:
            anio, mes = per_celda.year, per_celda.month
        else:
            ay = _periodo_carpeta(d, anio_defecto)
            if ay is None:
                print(f"  [!] {d}: sin filas y no se reconoce el mes en el nombre, omitida")
                continue
            anio, mes = ay
        meses.append({"per": f"{anio:04d}-{mes:02d}", "po": anio * 100 + mes, "claims": claims, "carpeta": d})
    return meses


def fusionar(meses):
    """Une entradas que caen en el mismo periodo (carpetas repetidas y, en el modelo de dos fuentes,
    las dos fuentes): los claims se combinan por referencia (la entrada procesada después gana)."""
    fusion: dict[str, dict] = {}
    for m in meses:
        if m["per"] in fusion:
            fusion[m["per"]]["claims"].update(m["claims"])
            if m["carpeta"] not in fusion[m["per"]]["carpeta"]:
                fusion[m["per"]]["carpeta"] += " + " + m["carpeta"]
        else:
            fusion[m["per"]] = dict(m)
    return sorted(fusion.values(), key=lambda x: x["po"])


def volcar(db, b, meses, alias_ref="", crear_siniestros=False, apply=False, etiqueta="Claims (carpetas)"):
    """Empareja los claims de `meses` con los siniestros del binder, opcionalmente crea los
    huérfanos, y vuelca las presentaciones (ClaimsPresentacion) + bloquea cada mes. DRY-RUN si
    apply=False (solo imprime el resumen)."""
    alias = {}
    for par in alias_ref.split(","):
        if "=" in par:
            k, v = par.split("=", 1)
            alias[_tok(k)] = _tok(v)

    sins = db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all()
    ref_cnt = Counter(_tok(s.reference) for s in sins if s.reference)
    por_par = {((s.certificate or "").strip(), _tok(s.reference)): s.id for s in sins}
    por_ref = {_tok(s.reference): s.id for s in sins if s.reference and ref_cnt[_tok(s.reference)] == 1}

    def casar(cert, ref):
        n = alias.get(_tok(ref), _tok(ref))
        return por_par.get((cert, n)) or por_ref.get(n)

    huerfanos = {}
    for m in meses:
        for ref, g in m["claims"].items():
            cert = str(g("Certificate Reference") or "").strip()
            if casar(cert, ref) is None and (ref not in huerfanos or m["po"] > huerfanos[ref][0]):
                huerfanos[ref] = (m["po"], g, cert)

    print(f"== {etiqueta} — binder {b.umr} (DRY-RUN={'NO' if apply else 'SÍ'}) ==")
    for m in meses:
        nm = sum(1 for ref, g in m["claims"].items()
                 if casar(str(g("Certificate Reference") or "").strip(), ref) is None and ref in huerfanos)
        et = f"{len(m['claims'])} claim(s)" if m["claims"] else "NIL"
        print(f"  {m['carpeta']:30} {m['per']}: {et}"
              + (f"  [+ {nm} nuevo(s)]" if nm and crear_siniestros else (f"  [! {nm} sin casar]" if nm else "")))
    print(f"Claims huérfanos (sin siniestro): {sorted(huerfanos)}"
          + ("  -> se CREARÁN" if crear_siniestros else "  -> se OMITIRÁN (usa --crear-siniestros para darlos de alta)"))

    if not apply:
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply.")
        return

    nuevos = 0
    if crear_siniestros:
        for ref, (po, g, cert) in huerfanos.items():
            db.add(Siniestro(
                binder_id=b.id, certificate=cert or None, reference=ref,
                insured=(_normh(g("Insured Full Name or Company Name")) or None),
                risk_code=(_normh(g("Lloyd's Risk Code")) or None),
                currency=(_normh(g("Original Currency")) or None),
                claimant=(_normh(g("Claimant Name")) or None),
                reporting_period=f"{po//100:04d}-{po%100:02d}",
                risk_inception=_fecha(g("Risk Inception Date")),
                risk_expiry=_fecha(g("Risk Expiry Date")),
                claim_first_advised=_fecha(g("Date Claim First Advised/Date Claim Made")),
                description=(_normh(g("Loss Description")) or None),
                status=(_normh(g("Claim Status")) or None),
                yoa=b.yoa,
                date_opened=_fecha(g("Date Claim Opened")),
                date_closed=_fecha(g("Date Closed")),
                amount_claimed=_dec(g("Amount Claimed")),
                paid_indemnity=_dec(_num(g("Previously Paid - Indemnity")) + _num(g("Paid this month - Indemnity"))),
                paid_fees=_dec(_num(g("Previously Paid - Fees")) + _num(g("Paid this month - Fees"))),
                reserves_indemnity=_dec(g("Reserve - Indemnity")),
                reserves_fees=_dec(g("Reserve - Fees")),
                total_indemnity=_dec(_num(g("Previously Paid - Indemnity")) + _num(g("Paid this month - Indemnity")) + _num(g("Reserve - Indemnity"))),
                total_fees=_dec(_num(g("Previously Paid - Fees")) + _num(g("Paid this month - Fees")) + _num(g("Reserve - Fees"))),
                ucr=(_normh(g("UCR")) or None),
            ))
            nuevos += 1
        db.flush()
        for s in db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all():
            por_par[((s.certificate or "").strip(), _tok(s.reference))] = s.id

    total = 0
    for m in meses:
        p, po = m["per"], m["po"]
        db.execute(delete(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == p))
        if not m["claims"]:
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=p, periodo_ord=po, siniestro_id=None,
                paid_indemnity_acum=Decimal("0"), paid_fees_acum=Decimal("0"),
                to_pay_indemnity=Decimal("0"), to_pay_fees=Decimal("0"),
                reserves_indemnity=Decimal("0"), reserves_fees=Decimal("0"), status="Nil",
                fila_json=json.dumps({"nil": True, "report": f"{p} — presentado en blanco"}, ensure_ascii=False),
                fecha_presentacion=None, usuario="histórico-nil"))
            total += 1
        else:
            for ref, g in m["claims"].items():
                cert = str(g("Certificate Reference") or "").strip()
                sid = casar(cert, ref)
                if sid is None:
                    continue
                fila = {h: _safe(h, g(h)) for h in HEADERS}
                db.add(ClaimsPresentacion(
                    binder_id=b.id, periodo=p, periodo_ord=po, siniestro_id=sid,
                    paid_indemnity_acum=_dec(_num(g("Previously Paid - Indemnity")) + _num(g("Paid this month - Indemnity"))),
                    paid_fees_acum=_dec(_num(g("Previously Paid - Fees")) + _num(g("Paid this month - Fees"))),
                    to_pay_indemnity=_dec(g("Paid this month - Indemnity")),
                    to_pay_fees=_dec(g("Paid this month - Fees")),
                    reserves_indemnity=_dec(g("Reserve - Indemnity")),
                    reserves_fees=_dec(g("Reserve - Fees")),
                    status=(_normh(g("Claim Status")) or None),
                    fila_json=json.dumps(fila, ensure_ascii=False, default=str),
                    fecha_presentacion=None, usuario="histórico"))
                total += 1
        if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == p)):
            db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=p))
    db.commit()
    if _XL["app"] is not None:
        _XL["app"].Quit()
        _XL["app"] = None
    print(f"\nAPLICADO: {total} presentaciones en {len(meses)} meses. Siniestros creados: {nuevos}. Meses bloqueados.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--carpeta", required=True, help="Carpeta 'Claims' con subcarpetas por mes.")
    ap.add_argument("--binder-id", type=int)
    ap.add_argument("--agreement")
    ap.add_argument("--crear-siniestros", action="store_true",
                    help="Da de alta los siniestros que aparezcan en los snapshots y no existan.")
    ap.add_argument("--anio-defecto", type=int, default=None,
                    help="Año para carpetas cuyo nombre es solo el mes (p. ej. 'Enero' sin año). "
                         "Sólo se usa como respaldo cuando el mes no tiene filas con periodo.")
    ap.add_argument("--alias-ref", default="",
                    help="Mapea referencias de snapshot a la referencia del siniestro cuando el "
                         "claim fue renumerado. Formato: 'refOrigen=refSiniestro,otra=otra'. "
                         "Se compara ignorando espacios/guiones (p. ej. '116498=116498001').")
    ap.add_argument("--periodo-override", default="",
                    help="Corrige el periodo de una carpeta cuya celda Reporting Period viene mal en "
                         "origen. Formato: 'NombreCarpeta=AAAA-MM' o 'NumCarpeta=AAAA-MM' "
                         "(p. ej. '1. Enero=2024-01' o '1=2024-01').")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    db = SessionLocal()
    b = db.get(Binder, args.binder_id) if args.binder_id else \
        db.scalar(select(Binder).where(Binder.agreement_number == args.agreement))
    if b is None:
        print("Binder no encontrado.")
        return
    agr_tok = _tok(b.agreement_number)

    overrides = {}
    for par in args.periodo_override.split(","):
        if "=" in par:
            k, v = par.split("=", 1)
            overrides[k.strip()] = v.strip()

    meses = fusionar(leer_carpeta(args.carpeta, agr_tok, args.anio_defecto, overrides, "ges40"))
    volcar(db, b, meses, args.alias_ref, args.crear_siniestros, args.apply, "Claims Heca (carpetas)")
    db.close()


if __name__ == "__main__":
    main()
