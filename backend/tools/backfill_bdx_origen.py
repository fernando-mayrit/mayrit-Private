"""
Recuperación (backfill) de columnas que NO migraron a BdxLinea, leyéndolas de los Risk BDX de origen
en SharePoint (carpeta sincronizada en local). NO usa el importador antiguo.

Causa de la pérdida: el importador casaba columnas por nombre exacto y se saltaba en silencio las que
no coincidían (p. ej. 'Insured Country (ISO code list)' vs el esperado 'see code list'). El dato sigue
en los Excel de origen; aquí lo recuperamos casando por Certificate Ref y rellenando SOLO celdas vacías.

Por defecto es DRY-RUN (no escribe). Con --apply escribe en la BD (solo campos vacíos, nunca pisa dato).

Uso:
    python -m tools.backfill_bdx_origen                 # dry-run, todos los binders
    python -m tools.backfill_bdx_origen --umr PI2525    # dry-run, solo binders cuyo agreement contenga eso
    python -m tools.backfill_bdx_origen --apply         # ESCRIBE (solo celdas vacías)
"""
from __future__ import annotations

import argparse
import os
import re
from collections import defaultdict

import openpyxl
from sqlalchemy import select

from app.db import SessionLocal
from app.models.maestras import Bdx, BdxLinea, Binder

BASE = r"C:\Users\ferna\Mayrit Insurance Broker\Mayrit - Negocio - Documentos\Agencias de Suscripcion"

# Campos a recuperar y cómo detectar su columna en el encabezado del origen (por palabras clave,
# robusto a variantes de texto). Cada detector: (claves_que_deben_estar, claves_que_NO_deben_estar).
DETECTORES = {
    "umr":                    (["unique market reference"], []),
    "certificate_ref":        (["certificate", "ref"], []),
    "section_no":             (["section"], ["sub"]),
    "risk_code":              (["risk code"], []),
    "insured_province":       (["insured", "sub-division"], ["location"]),
    "insured_postcode":       (["insured", "postcode"], []),
    "insured_country":        (["insured", "country"], ["sub-division", "sub division"]),
    "location_risk_province": (["location", "risk", "sub-division"], []),
    "location_risk_country":  (["location", "risk", "country"], ["sub-division", "sub division"]),
    "sum_insured_total":      (["sum insured amount"], []),
    "written_line_pct":       (["written line"], []),
}
# Campos constantes por certificado (mismo asegurado → mismo valor en todas sus líneas).
POR_CERTIFICADO = ["insured_province", "insured_postcode", "insured_country",
                   "location_risk_province", "location_risk_country"]


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _cert(s) -> str:
    return re.sub(r"\s+", "", str(s or "").strip()).upper()


def _detectar_columnas(header) -> dict[str, int]:
    """Mapea campo -> índice 0-based de columna, detectando por palabras clave (robusto a variantes)."""
    cols: dict[str, int] = {}
    norm = [_norm(h) for h in header]
    for campo, (si, no) in DETECTORES.items():
        for j, h in enumerate(norm):
            if h and all(k in h for k in si) and not any(k in h for k in no):
                cols[campo] = j
                break
    return cols


def _strip(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def _todos_risk_files():
    out = []
    for root, _dirs, files in os.walk(BASE):
        if os.path.basename(root).lower() != "risk":
            continue
        for fn in files:
            if fn.lower().endswith(".xlsx") and not fn.startswith("~$"):
                out.append(os.path.join(root, fn))
    return out


def _leer_risk(full):
    """Abre un Risk BDX. Devuelve (umr, cols, filas_datos) o None si no parece un bordereau Lloyd's."""
    wb = openpyxl.load_workbook(full, data_only=True, read_only=True)
    ws = wb["BDX"] if "BDX" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_idx = next((i for i, r in enumerate(rows[:15])
                    if any("certificate" in _norm(c) for c in r)), None)
    if hdr_idx is None:
        return None
    cols = _detectar_columnas(rows[hdr_idx])
    if "certificate_ref" not in cols or "umr" not in cols:
        return None
    datos = rows[hdr_idx + 1:]
    ju = cols["umr"]
    umr = next((_strip(r[ju]) for r in datos if len(r) > ju and r[ju]), "")
    return umr, cols, datos


def _recopilar_origen(binders):
    """Casa cada Risk BDX con su binder por el UMR de dentro del fichero (no por el nombre de carpeta).
    Acumula, por binder, los valores de los campos por certificado."""
    umr2bid = {_strip(b.umr): b.id for b in binders if b.umr}
    src: dict[int, dict] = defaultdict(lambda: defaultdict(dict))   # bid -> campo -> {cert: valor}
    stats: dict[int, dict] = defaultdict(lambda: {"files": 0, "rows": 0})
    sin_match = ilegibles = 0
    for full in _todos_risk_files():
        try:
            res = _leer_risk(full)
        except Exception:
            ilegibles += 1
            continue
        if not res:
            continue
        umr, cols, datos = res
        bid = umr2bid.get(umr)
        if not bid:
            sin_match += 1
            continue
        stats[bid]["files"] += 1
        ic = cols["certificate_ref"]
        for r in datos:
            cert = _cert(r[ic]) if len(r) > ic else ""
            if not cert:
                continue
            stats[bid]["rows"] += 1
            for campo in POR_CERTIFICADO:
                j = cols.get(campo)
                if j is not None and len(r) > j and r[j] not in (None, ""):
                    src[bid][campo].setdefault(cert, r[j])
    return src, stats, sin_match, ilegibles


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--umr", default=None, help="filtra binders cuyo agreement/UMR contenga esto")
    ap.add_argument("--apply", action="store_true", help="ESCRIBE en la BD (por defecto solo dry-run)")
    args = ap.parse_args()

    db = SessionLocal()
    binders = list(db.scalars(select(Binder)).all())
    by_id = {b.id: b for b in binders}
    if args.umr:
        f = args.umr.upper()
        binders = [b for b in binders if f in (b.agreement_number or "").upper() or f in (b.umr or "").upper()]

    print(f"== Backfill BDX origen ({'APPLY' if args.apply else 'DRY-RUN'}) ==  (match por UMR de contenido)")
    src, stats, sin_match, ilegibles = _recopilar_origen(binders)
    print(f"Binders con Risk BDX localizados: {len(stats)} | ficheros sin binder (legacy/otros): {sin_match} | ilegibles: {ilegibles}")

    tot_fill = defaultdict(int)
    tot_apply = 0
    for bid in sorted(stats, key=lambda x: by_id[x].umr or ""):
        b = by_id[bid]
        lineas = db.scalars(
            select(BdxLinea).join(Bdx, BdxLinea.bdx_id == Bdx.id).where(Bdx.binder_id == bid)
        ).all()
        fill_binder = defaultdict(int)
        for l in lineas:
            cert = _cert(l.certificate_ref)
            if not cert:
                continue
            for campo in POR_CERTIFICADO:
                if getattr(l, campo) in (None, "") and cert in src[bid].get(campo, {}):
                    fill_binder[campo] += 1
                    tot_fill[campo] += 1
                    if args.apply:
                        setattr(l, campo, src[bid][campo][cert])
                        tot_apply += 1
        if fill_binder:
            resumen = ", ".join(f"{k}={v}" for k, v in fill_binder.items())
            st = stats[bid]
            print(f"  {b.umr or b.agreement_number}: {st['files']} fichero(s), {st['rows']} filas origen, {len(lineas)} líneas BD -> rellenaría: {resumen}")

    if args.apply:
        db.commit()
        print(f"\nAPLICADO. Celdas escritas: {tot_apply}")
    else:
        print("\n== TOTAL que se rellenaría (dry-run) ==")
        for campo, n in tot_fill.items():
            print(f"  {campo}: {n}")
        print("\n(no se ha escrito nada — añade --apply para ejecutar)")
    db.close()


if __name__ == "__main__":
    main()
