"""
Importa el BDX (Risk) de UN binder desde un Excel exportado de su lista 'Mayrit - <UMR>'.

Reutiliza el mapeo de columnas de app.sharepoint (estándar Lloyd's + columnas de control) y la
inserción de app.bdx_import (coerción de tipos, % en fracción ×100, idempotente por sp_old_id).
Por defecto DRY-RUN; para escribir: --apply.

Uso:
  py -m tools.migrar_bdx_excel "RUTA.xlsx" [--umr B1634...] [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
import os

import openpyxl
from sqlalchemy import select

from app.bdx_import import _num, importar_filas
from app.db import SessionLocal
from app.models.maestras import Binder
from app.sharepoint import DATE_FIELDS, MAPEO, _norm, _solo_fecha

UMR_TITULO = "Unique Market Reference (UMR)"


def leer_excel(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [(_norm(str(c)).lower() if c is not None else "", i) for i, c in enumerate(rows[0])]

    def resolver(aliases):
        opts = [aliases] if isinstance(aliases, str) else aliases
        normados = [_norm(a).lower() for a in opts]
        for a in normados:
            for t, i in hdr:
                if t == a:
                    return i
        for a in normados:
            for t, i in hdr:
                if t and t.startswith(a):
                    return i
        return None

    idx = {campo: resolver(al) for campo, al in MAPEO.items()}
    if idx.get("sp_old_id") is None:  # en el Excel la columna del Id es "Id" (no "_OldID")
        idx["sp_old_id"] = next((i for t, i in hdr if t == "id"), None)
    umr_idx = resolver(UMR_TITULO)

    filas, umrs = [], set()
    for r in rows[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        if umr_idx is not None and umr_idx < len(r) and r[umr_idx]:
            umrs.add(str(r[umr_idx]).strip())
        fila = {}
        for campo, i in idx.items():
            v = r[i] if (i is not None and i < len(r)) else None
            if isinstance(v, (dt.datetime, dt.date)):
                v = v.isoformat()
            fila[campo] = _solo_fecha(v) if campo in DATE_FIELDS else v
        filas.append(fila)
    return filas, sorted(umrs)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--umr", default=None)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    filas, umrs = leer_excel(args.excel)
    umr = args.umr or (umrs[0] if len(umrs) == 1 else None)
    if not umr:
        print(f"No se pudo determinar el UMR (UMRs en el fichero: {umrs}). Usa --umr.")
        return

    db = SessionLocal()
    binder = db.scalar(select(Binder).where(Binder.umr == umr))
    print(f"== Importar Risk BDX de {umr} (DRY-RUN={'NO' if args.apply else 'SÍ'}) ==")
    print(f"Fichero: {os.path.basename(args.excel)}")
    print(f"UMR(s) en el fichero: {umrs}")
    if binder is None:
        print(f"⚠ Binder {umr} NO existe en la BD. Créalo antes de importar.")
        db.close()
        return
    print(f"Binder encontrado: id={binder.id}")
    print(f"Líneas en el Excel: {len(filas)}")
    con_id = sum(1 for f in filas if f.get("sp_old_id") not in (None, ""))
    print(f"Con Id (idempotencia): {con_id}/{len(filas)}")
    gwp = sum((_num(f.get("gross_written_premium")) or 0) for f in filas)
    print(f"GWP total (origen): {round(float(gwp), 2)}")
    print("Muestra (primeras 3):")
    for f in filas[:3]:
        print(f"   · {f.get('certificate_ref')} | risk {f.get('reporting_period_start')} | GWP {f.get('gross_written_premium')} | brokerage {f.get('brokerage_amount')} | recibo {f.get('recibo')}")

    if not args.apply:
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply para importar.")
        db.close()
        return

    res = importar_filas(db, binder, filas, origen=os.path.basename(args.excel))
    print("\nAPLICADO. Conciliación:")
    print(f"   insertadas={res['insertadas']} actualizadas={res['actualizadas']} sin_old_id={res['sin_old_id']}")
    c = res["conciliacion"]
    print(f"   líneas: origen={c['lineas_sharepoint']} / bd={c['lineas_postgres']} ok={c['lineas_ok']}")
    print(f"   GWP: origen={c['gwp_sharepoint']} / bd={c['gwp_postgres']} ok={c['gwp_ok']}")
    print(f"   periodos: {res['periodos']}")
    db.close()


if __name__ == "__main__":
    main()
