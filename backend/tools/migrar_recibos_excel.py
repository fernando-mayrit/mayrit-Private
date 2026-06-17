"""
Migración de recibos históricos desde un Excel exportado de 'Mayrit - TRecibos'.

- Importa SOLO los recibos de tipo **Binder** (enlazados al binder por NumeroPoliza = UMR).
- Los de tipo Póliza / Slip de Reaseguro (Open Market) se REPORTAN como pendientes (no se tocan;
  se migrarán cuando exista el módulo Pólizas con `poliza_id`).
- Idempotente: casa por `sp_old_id` (Id de SharePoint) y por `numero`.
- Por defecto es DRY-RUN (no escribe). Para aplicar: --apply.

Uso:
  py -m tools.migrar_recibos_excel "RUTA.xlsx" --anios 2017,2018 [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from sqlalchemy import select

from app.db import SessionLocal
from app.models.maestras import Binder, Recibo

# Columna Excel (TRecibos) -> campo del modelo Recibo. Importes/fechas/textos directos.
MAP_DIRECTO = {
    "NumeroRecibo": "numero", "Referencia": "referencia", "NombreMercado": "nombre_mercado",
    "Mercado": "mercado", "NumeroPoliza": "numero_poliza", "Asegurado": "asegurado",
    "Corredor": "corredor", "Ramo": "ramo", "TipoPoliza": "tipo_poliza", "Produccion": "produccion",
    "FechaEfecto": "fecha_efecto", "FechaVencimiento": "fecha_vencimiento", "Pago": "pago",
    "Moneda": "moneda", "PrimaNetaPoliza": "prima_neta_poliza", "Recibo": "recibo_num",
    "RecibosTotales": "recibos_totales", "FechaEfectoRecibo": "fecha_efecto_recibo",
    "FechaVctoRecibo": "fecha_vcto_recibo", "PrimaNetaRecibo": "prima_neta_recibo",
    "ImpuestossobreRecibo": "impuestos_sobre_recibo", "OtrosImpuestos": "otros_impuestos",
    "ImpuestosRecibo": "impuestos_recibo", "PrimaBrutaRecibo": "prima_bruta_recibo",
    "DeduccionTotal": "deduccion_total", "Honorarios": "honorarios", "ComisionCedida": "comision_cedida",
    "ComisionRetenida": "comision_retenida", "Pagador": "pagador", "PrimaAdeudada": "prima_adeudada",
    "PrimaCobrada": "prima_cobrada", "PrimaFechaCobro": "prima_fecha_cobro",
    "ComisionRetenidaCobrada": "comision_retenida_cobrada",
    "ComisionRetenidaTraspasada": "comision_retenida_traspasada",
    "ComisionFechaTraspaso": "comision_fecha_traspaso", "ComisionPendienteCobro": "comision_pendiente_cobro",
    "Liquidar": "liquidar", "LiquidarCobrado": "liquidar_cobrado",
    "LiquidarPendienteCobro": "liquidar_pendiente_cobro", "LiquidarLiquidado": "liquidar_liquidado",
    "LiquidarFechaLiquidacion": "liquidar_fecha_liquidacion", "ComisionCedidaaPagar": "comision_cedida_a_pagar",
    "ComisionCedidaPagada": "comision_cedida_pagada", "ComisionCedidaFechaPago": "comision_cedida_fecha_pago",
    "Notas": "notas", "Cuenta": "cuenta", "FechaContable": "fecha_contable", "YOA": "yoa",
}
# Columnas % (en TRecibos vienen como fracción 0.05 -> 5): se multiplican por 100.
MAP_PORC = {
    "ImpuestosPorc": "impuestos_porc", "ImpuestossobreTotalPorc": "impuestos_sobre_total_porc",
    "ImpuestossobreReciboPorc": "impuestos_sobre_recibo_porc", "DeduccionTotalPorc": "deduccion_total_porc",
    "ComisionCedidaPorc": "comision_cedida_porc", "ComisionRetenidaPorc": "comision_retenida_porc",
    "Participacion": "participacion",
}
FECHAS = {"fecha_efecto", "fecha_vencimiento", "fecha_efecto_recibo", "fecha_vcto_recibo",
          "prima_fecha_cobro", "comision_fecha_traspaso", "liquidar_fecha_liquidacion",
          "comision_cedida_fecha_pago", "fecha_contable"}
IMPORTES = {"prima_neta_poliza", "prima_neta_recibo", "otros_impuestos", "impuestos_recibo",
            "prima_bruta_recibo", "deduccion_total", "honorarios", "comision_cedida", "comision_retenida",
            "prima_adeudada", "prima_cobrada", "comision_retenida_cobrada", "comision_retenida_traspasada",
            "comision_pendiente_cobro", "liquidar", "liquidar_cobrado", "liquidar_pendiente_cobro",
            "liquidar_liquidado", "comision_cedida_a_pagar", "comision_cedida_pagada"}


def _fecha(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _dec(v, places="0.01"):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v)).quantize(Decimal(places), ROUND_HALF_UP)
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--anios", default="2017,2018")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    anios = {a.strip() for a in args.anios.split(",")}

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}

    def get(r, col):
        i = ix.get(col)
        return r[i] if i is not None and i < len(r) else None

    db = SessionLocal()
    binders = db.scalars(select(Binder)).all()
    by_umr = {}
    for b in binders:
        for k in (b.umr, b.agreement_number):
            if k:
                by_umr[k.strip()] = b.id

    importables, om_pendientes, sin_binder, ya_existen, otros_anios = [], [], [], [], 0
    for r in rows[1:]:
        numero = get(r, "NumeroRecibo")
        if not numero:
            continue
        numero = str(numero).strip()
        anio = numero[:4]
        if anio not in anios:
            otros_anios += 1
            continue
        tipo = str(get(r, "TipoPoliza") or "").strip()
        if tipo != "Binder":
            om_pendientes.append((numero, tipo, str(get(r, "NumeroPoliza") or "")))
            continue
        umr = str(get(r, "NumeroPoliza") or "").strip()
        binder_id = by_umr.get(umr)
        if not binder_id:
            sin_binder.append((numero, umr))
            continue
        fer = _fecha(get(r, "FechaEfectoRecibo")) or _fecha(get(r, "FechaContable"))
        periodo = fer.strftime("%Y-%m") if fer else None
        old_id = get(r, "Id")
        existe = db.scalar(select(Recibo).where(Recibo.numero == numero))
        if not existe and binder_id and periodo:
            existe = db.scalar(select(Recibo).where(Recibo.binder_id == binder_id, Recibo.periodo == periodo))
        if existe:
            ya_existen.append(numero)
            continue
        importables.append({"numero": numero, "binder_id": binder_id, "umr": umr, "periodo": periodo, "old_id": old_id, "row": r, "get": get})

    print(f"== Migración recibos {sorted(anios)} (DRY-RUN={'NO' if args.apply else 'SÍ'}) ==")
    print(f"Importables (Binder enlazado): {len(importables)}")
    print(f"OM pendientes (Póliza/Slip): {len(om_pendientes)}")
    for n, t, p in om_pendientes:
        print(f"   · {n} [{t}] póliza {p}")
    print(f"Binder NO encontrado: {len(sin_binder)}")
    for n, u in sin_binder:
        print(f"   · {n} UMR {u}")
    print(f"Ya existen (omitidos): {len(ya_existen)} {ya_existen}")
    print(f"(otros años, ignorados: {otros_anios})")
    print("Importables detalle:")
    for it in importables:
        print(f"   · {it['numero']} -> binder {it['binder_id']} ({it['umr']}) periodo {it['periodo']}")

    if not args.apply:
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply para migrar.")
        db.close()
        return

    creados = 0
    for it in importables:
        r, get = it["row"], it["get"]
        rec = Recibo(numero=it["numero"], binder_id=it["binder_id"], periodo=it["periodo"],
                     anio=int(it["numero"][:4]), sp_old_id=it["old_id"], estado="Emitido")
        for col, campo in MAP_DIRECTO.items():
            if campo in ("numero",):
                continue
            v = get(r, col)
            if campo in FECHAS:
                v = _fecha(v)
            elif campo in IMPORTES:
                v = _dec(v)
            elif campo == "impuestos_sobre_recibo":
                v = bool(v)
            elif campo == "yoa":
                v = int(v) if v not in (None, "") else None
            elif campo == "recibo_num":
                v = int(v) if v not in (None, "") else None
            else:
                v = str(v).strip() if v not in (None, "") else None
            if v is not None:
                setattr(rec, campo, v)
        for col, campo in MAP_PORC.items():
            v = get(r, col)
            if v not in (None, ""):
                d = _dec(Decimal(str(v)) * 100, "0.0001")
                if d is not None:
                    setattr(rec, campo, d)
        db.add(rec)
        creados += 1
    db.commit()
    print(f"\nAPLICADO: {creados} recibos creados.")
    db.close()


if __name__ == "__main__":
    main()
