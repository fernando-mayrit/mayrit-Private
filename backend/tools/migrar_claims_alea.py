"""
Importa Claims del FORMATO ANTIGUO de Alea ("Claims Brx … <agencia>.xlsx"): un workbook con una
pestaña por mes, filas de título ALEA arriba, cabecera en INGLÉS (fila ~7) y ESPAÑOL (fila ~8), y
datos debajo. Columnas por POSICIÓN fija (Claim Number, Policy/Slip Reference, …), distintas del
estándar Lloyd's.

A diferencia del migrador canónico, aquí el binder NO tiene siniestros: se CREAN/actualizan desde
los snapshots (upsert por (certificate, reference), cronológico → el último mes define el estado
actual) y se vuelca una presentación por mes (con bloqueo). Las hojas sin datos reales = mes NIL.

DRY-RUN por defecto. Uso:
  py -m tools.migrar_claims_alea --excel "RUTA.xlsx" --binder-id 20 [--anio-defecto 2018] [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
from collections import Counter
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.maestras import BdxBloqueo, Binder, ClaimsPresentacion, Siniestro
from app.routers.claims_bdx import HEADERS
from tools.migrar_claims_presentaciones import _periodo_de_hoja

# Posición (0-indexada) de cada dato en la fila del bordereau antiguo de Alea.
COL = {
    "reference": 0,          # Claim Number
    "made": 1,               # Date Of Claim made against Insured
    "certificate": 3,        # Policy/Slip Reference
    "yoa": 4,
    "insured": 5,
    "currency": 9,           # Limit Currency
    "risk_code": 12,
    "claimant": 13,
    "risk_inception": 14,    # Inception
    "risk_expiry": 15,       # Expiry Date
    "amount_claimed": 16,
    "description": 18,       # Brief Description of Claim
    "to_pay_ind": 20, "paid_ind": 21, "reserve_ind": 22,
    "to_pay_fee": 23, "paid_fee": 24, "reserve_fee": 25,
    "status": 30,            # Open / Closed
}


def _norm(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _dec(v):
    return Decimal(str(_num(v))).quantize(Decimal("0.01"), ROUND_HALF_UP)


def _fecha(v):
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def _g(row, key):
    c = COL[key]
    return row[c] if c < len(row) else None


def extraer(ws):
    """(claims, cabecera_ok). Localiza la fila de cabecera (col0 == 'Claim Number') y lee los datos."""
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    hr = next((r for r in range(min(12, len(grid))) if _norm(grid[r][0] if grid[r] else "") == "Claim Number"), None)
    if hr is None:
        return [], False
    claims = []
    for r in range(hr + 1, len(grid)):
        row = grid[r]
        ref = _norm(_g(row, "reference"))
        cert = _norm(_g(row, "certificate"))
        if ref.lower() in ("", "numero de siniestro", "none") or cert.lower() in ("", "none"):
            continue
        claims.append({"ref": ref, "cert": cert, "row": row})
    return claims, True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--binder-id", type=int)
    ap.add_argument("--agreement")
    ap.add_argument("--anio-defecto", type=int, default=None)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    db = SessionLocal()
    b = db.get(Binder, args.binder_id) if args.binder_id else \
        db.scalar(select(Binder).where(Binder.agreement_number == args.agreement))
    if b is None:
        print("Binder no encontrado.")
        return

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    meses = []
    for hoja in wb.sheetnames:
        claims, cab = extraer(wb[hoja])
        ay = _periodo_de_hoja(hoja, args.anio_defecto)
        if ay is None:
            if claims:
                print(f"  [!] hoja {hoja!r}: nombre no es '<Mes> <Año>', omitida")
            continue
        anio, mes = ay
        meses.append({"hoja": hoja, "per": f"{anio:04d}-{mes:02d}", "po": anio * 100 + mes,
                      "claims": claims, "nil": cab and not claims})
    meses.sort(key=lambda m: m["po"])

    print(f"== Claims Alea (antiguo) — binder {b.umr} (DRY-RUN={'NO' if args.apply else 'SÍ'}) ==")
    refs = set()
    n_nil = 0
    for m in meses:
        if m["nil"]:
            n_nil += 1
            print(f"  {m['per']}: NIL (presentado en blanco)")
        else:
            rs = sorted({c["ref"] for c in m["claims"]})
            refs |= {(c["cert"], c["ref"]) for c in m["claims"]}
            print(f"  {m['per']}: {len(m['claims'])} claim(s) {rs}")
    print(f"Claims distintos (certificate, reference): {len(refs)} · meses NIL: {n_nil}")

    if not args.apply:
        db.close()
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply.")
        return

    # ── PASO 1: upsert de siniestros por (certificate, reference) cronológico (último mes gana) ──
    sins = {((s.certificate or "").strip(), (s.reference or "").strip()): s
            for s in db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all()}
    for m in meses:
        for cl in m["claims"]:
            row = cl["row"]
            key = (cl["cert"], cl["ref"])
            s = sins.get(key)
            if s is None:
                s = Siniestro(binder_id=b.id)
                db.add(s)
                sins[key] = s
            s.certificate = cl["cert"]
            s.reference = cl["ref"]
            s.insured = _norm(_g(row, "insured")) or None
            s.risk_code = _norm(_g(row, "risk_code")) or None
            s.currency = _norm(_g(row, "currency")) or None
            s.claimant = _norm(_g(row, "claimant")) or None
            s.reporting_period = m["per"]
            s.risk_inception = _fecha(_g(row, "risk_inception"))
            s.risk_expiry = _fecha(_g(row, "risk_expiry"))
            s.claim_first_advised = _fecha(_g(row, "made"))
            s.description = _norm(_g(row, "description")) or None
            s.status = _norm(_g(row, "status")) or None
            s.yoa = int(_num(_g(row, "yoa"))) or None
            s.amount_claimed = _dec(_g(row, "amount_claimed"))
            s.paid_indemnity = _dec(_g(row, "paid_ind"))
            s.paid_fees = _dec(_g(row, "paid_fee"))
            s.reserves_indemnity = _dec(_g(row, "reserve_ind"))
            s.reserves_fees = _dec(_g(row, "reserve_fee"))
            s.total_indemnity = _dec(_num(_g(row, "paid_ind")) + _num(_g(row, "reserve_ind")))
            s.total_fees = _dec(_num(_g(row, "paid_fee")) + _num(_g(row, "reserve_fee")))
    db.flush()

    # ── PASO 2: presentaciones por mes + bloqueo ──
    total = 0
    for m in meses:
        db.execute(delete(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == m["per"]))
        if m["nil"]:
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=m["per"], periodo_ord=m["po"], siniestro_id=None,
                paid_indemnity_acum=Decimal("0"), paid_fees_acum=Decimal("0"),
                to_pay_indemnity=Decimal("0"), to_pay_fees=Decimal("0"),
                reserves_indemnity=Decimal("0"), reserves_fees=Decimal("0"), status="Nil",
                fila_json=json.dumps({"nil": True, "report": f"{m['per']} — presentado en blanco"}, ensure_ascii=False),
                fecha_presentacion=None, usuario="histórico-nil"))
            total += 1
        else:
            for cl in m["claims"]:
                row = cl["row"]
                s = sins.get((cl["cert"], cl["ref"]))
                paid_i = _num(_g(row, "paid_ind")); paid_f = _num(_g(row, "paid_fee"))
                tp_i = _num(_g(row, "to_pay_ind")); tp_f = _num(_g(row, "to_pay_fee"))
                # fila_json canónica (mejor esfuerzo) para el snapshot congelado.
                fila = {h: None for h in HEADERS}
                fila.update({
                    "Reporting Period (End Date)": m["per"],
                    "Certificate Reference": cl["cert"], "Claim Reference / Number": cl["ref"],
                    "Insured Full Name or Company Name": _norm(_g(row, "insured")) or None,
                    "Lloyd's Risk Code": _norm(_g(row, "risk_code")) or None,
                    "Original Currency": _norm(_g(row, "currency")) or None,
                    "Claimant Name": _norm(_g(row, "claimant")) or None,
                    "Loss Description": _norm(_g(row, "description")) or None,
                    "Claim Status": _norm(_g(row, "status")) or None,
                    "Amount Claimed": _num(_g(row, "amount_claimed")),
                    "Paid this month - Indemnity": tp_i, "Paid this month - Fees": tp_f,
                    "Previously Paid - Indemnity": paid_i - tp_i, "Previously Paid - Fees": paid_f - tp_f,
                    "Reserve - Indemnity": _num(_g(row, "reserve_ind")), "Reserve - Fees": _num(_g(row, "reserve_fee")),
                    "Total Incurred - Indemnity": paid_i + _num(_g(row, "reserve_ind")),
                    "Total Incurred - Fees": paid_f + _num(_g(row, "reserve_fee")),
                })
                db.add(ClaimsPresentacion(
                    binder_id=b.id, periodo=m["per"], periodo_ord=m["po"], siniestro_id=(s.id if s else None),
                    paid_indemnity_acum=_dec(paid_i), paid_fees_acum=_dec(paid_f),
                    to_pay_indemnity=_dec(tp_i), to_pay_fees=_dec(tp_f),
                    reserves_indemnity=_dec(_g(row, "reserve_ind")), reserves_fees=_dec(_g(row, "reserve_fee")),
                    status=_norm(_g(row, "status")) or None,
                    fila_json=json.dumps(fila, ensure_ascii=False, default=str),
                    fecha_presentacion=None, usuario="histórico"))
                total += 1
        if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == m["per"])):
            db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=m["per"]))
    db.commit()
    print(f"\nAPLICADO: {len(sins)} siniestros (upsert) y {total} filas de presentación en {len(meses)} meses.")
    db.close()


if __name__ == "__main__":
    main()
