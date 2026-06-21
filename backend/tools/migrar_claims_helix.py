"""
Importa los snapshots de Claims del formato HELIX/TME ("Claims Bdx Helix …xlsx"): un workbook con
una pestaña por mes, cabecera en la fila 1 con columnas propias de Helix (Claim Number, Policy
Number, Net Reserve …, Payment …, OPEN/CLOSED). Los siniestros YA existen (importados de SharePoint):
se CASAN por el Claim Number normalizado (Helix 'CLA1989A122' ↔ siniestro 'CLA/1989/A1/22') y se
vuelca una presentación por mes. Payment es acumulado → 'a pagar este mes' = diferencia cronológica.

DRY-RUN por defecto. Uso:
  py -m tools.migrar_claims_helix --excel "RUTA.xlsx" --binder-id 34 [--anio-defecto 2022] [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.maestras import BdxBloqueo, Binder, ClaimsPresentacion, Siniestro
from app.routers.claims_bdx import HEADERS
from tools.migrar_claims_presentaciones import _periodo_de_hoja

COL = {
    "insured": 0, "certificate": 3, "inception": 7, "expiry": 8,
    "reference": 10, "broker_ref": 11, "description": 13, "received": 14,
    "status": 16, "closed": 17, "claimant": 23,
    "reserve_ind": 30, "reserve_fee": 31, "paid_ind": 33, "paid_fee": 34,
}


def _norm(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


def _nref(s) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(s or "")).upper()


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _dec(v) -> Decimal:
    return Decimal(str(_num(v))).quantize(Decimal("0.01"), ROUND_HALF_UP)


def _g(row, key):
    c = COL[key]
    return row[c] if c < len(row) else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--binder-id", type=int)
    ap.add_argument("--agreement")
    ap.add_argument("--anio-defecto", type=int, default=None)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    db = SessionLocal()
    b = db.get(Binder, args.binder_id) if args.binder_id else \
        db.scalar(select(Binder).where(Binder.agreement_number == args.agreement))
    if b is None:
        print("Binder no encontrado.")
        return
    por_ref = {_nref(s.reference): s for s in db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all() if s.reference}

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    meses = []
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        hr = next((r for r in range(min(6, len(grid)))
                   if "Claim Number" in [_norm(v) for v in grid[r]]), None)
        if hr is None:
            continue
        rows = [grid[r] for r in range(hr + 1, len(grid)) if _norm(_g(grid[r], "reference"))]
        ay = _periodo_de_hoja(hoja, args.anio_defecto)
        if ay is None:
            if rows:
                print(f"  [!] hoja {hoja!r}: nombre no es '<Mes> <Año>', omitida")
            continue
        anio, mes = ay
        meses.append({"per": f"{anio:04d}-{mes:02d}", "po": anio * 100 + mes, "rows": rows, "nil": not rows})
    meses.sort(key=lambda m: m["po"])

    print(f"== Claims Helix — binder {b.umr} (DRY-RUN={'NO' if args.apply else 'SÍ'}) ==")
    sin_casar = 0
    for m in meses:
        nm = sum(1 for r in m["rows"] if _nref(_g(r, "reference")) not in por_ref)
        sin_casar += nm
        et = "NIL" if m["nil"] else f"{len(m['rows'])} claim(s)" + (f"  [!] {nm} sin casar" if nm else "")
        print(f"  {m['per']}: {et}")
    print(f"Filas sin casar con un siniestro: {sin_casar}")

    if not args.apply:
        db.close()
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply.")
        return

    prev_i: dict = {}
    prev_f: dict = {}
    total = 0
    for m in meses:
        db.execute(delete(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == m["per"]))
        if m["nil"]:
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=m["per"], periodo_ord=m["po"], siniestro_id=None,
                paid_indemnity_acum=Decimal("0"), paid_fees_acum=Decimal("0"),
                to_pay_indemnity=Decimal("0"), to_pay_fees=Decimal("0"),
                reserves_indemnity=Decimal("0"), reserves_fees=Decimal("0"), status="Nil",
                fila_json=json.dumps({"nil": True, "report": f"{m['per']} — presentado en blanco"}, ensure_ascii=False),
                fecha_presentacion=None, usuario="histórico-nil"))
            total += 1
        else:
            for r in m["rows"]:
                nr = _nref(_g(r, "reference"))
                s = por_ref.get(nr)
                pi = _num(_g(r, "paid_ind")); pf = _num(_g(r, "paid_fee"))
                ri_ = _num(_g(r, "reserve_ind")); rf = _num(_g(r, "reserve_fee"))
                tp_i = pi - prev_i.get(nr, 0.0); tp_f = pf - prev_f.get(nr, 0.0)
                prev_i[nr] = pi; prev_f[nr] = pf
                fila = {h: None for h in HEADERS}
                fila.update({
                    "Reporting Period (End Date)": m["per"],
                    "Certificate Reference": _norm(_g(r, "certificate")) or None,
                    "Claim Reference / Number": _norm(_g(r, "reference")) or None,
                    "Insured Full Name or Company Name": _norm(_g(r, "insured")) or None,
                    "Claimant Name": _norm(_g(r, "claimant")) or None,
                    "Loss Description": _norm(_g(r, "description")) or None,
                    "Claim Status": _norm(_g(r, "status")) or None,
                    "Paid this month - Indemnity": tp_i, "Paid this month - Fees": tp_f,
                    "Previously Paid - Indemnity": pi - tp_i, "Previously Paid - Fees": pf - tp_f,
                    "Reserve - Indemnity": ri_, "Reserve - Fees": rf,
                    "Total Incurred - Indemnity": pi + ri_, "Total Incurred - Fees": pf + rf,
                })
                db.add(ClaimsPresentacion(
                    binder_id=b.id, periodo=m["per"], periodo_ord=m["po"], siniestro_id=(s.id if s else None),
                    paid_indemnity_acum=_dec(pi), paid_fees_acum=_dec(pf),
                    to_pay_indemnity=_dec(tp_i), to_pay_fees=_dec(tp_f),
                    reserves_indemnity=_dec(ri_), reserves_fees=_dec(rf),
                    status=_norm(_g(r, "status")) or None,
                    fila_json=json.dumps(fila, ensure_ascii=False, default=str),
                    fecha_presentacion=None, usuario="histórico"))
                total += 1
        if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == m["per"])):
            db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=m["per"]))
    db.commit()
    print(f"\nAPLICADO: {total} filas de presentación en {len(meses)} meses.")
    db.close()


if __name__ == "__main__":
    main()
