"""
Importa el histórico de Claims de un binder cuyos bordereaux son .xls MENSUALES (un fichero por
mes, plantilla "Lloyd's Claims Template" V5.x de VAMMOS). A diferencia de migrar_claims_presentaciones
(un workbook con pestañas-mes), aquí:

- Cada mes es un .xls distinto en su carpeta; se barre un directorio base.
- Muchos .xls (plantillas V5.1/V5.2) no los lee xlrd → se CONVIERTEN a .xlsx con Excel (COM) y se
  leen con openpyxl.
- La cabecera NO está en la 1ª fila (hay una fila 'Ref' encima): se autodetecta la fila que
  contiene 'Claim Reference / Number'.
- Los datos están en una pestaña por territorio (p. ej. "Lloyd's Brussels and Europe"); se escanean
  todas las pestañas y se cogen las filas con un Claim Reference real (se ignoran filas de guía).
- El periodo se toma de la celda 'Reporting Period (End Date)'.
- El binder normalmente NO tiene siniestros: se CREAN/actualizan desde estos Excel (upsert por
  Claim Reference, en orden cronológico → el último mes define el estado actual) y luego se crea
  una presentación por mes (con bloqueo). El certificate puede repetirse → se casa por reference.

DRY-RUN por defecto. Para aplicar: --apply.
Uso:  py -m tools.migrar_claims_vammos --dir "RUTA\\Claims" --binder-id 24 [--apply]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import tempfile
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.maestras import BdxBloqueo, Binder, ClaimsPresentacion, Siniestro
from app.routers.claims_bdx import HEADERS

H_FECHA = {
    "Binding authority or coverholder appointment agreement inception date",
    "Binding authority or coverholder appointment agreement expiry date",
    "Reporting Period (End Date)", "Risk Inception Date", "Risk Expiry Date",
    "Date Claim First Advised/Date Claim Made", "Date Claim Opened", "Date Closed",
}
A_SINIESTRO = {
    "Lloyd's Risk Code": "risk_code", "Original Currency": "currency",
    "Certificate Reference": "certificate", "Claim Reference / Number": "reference",
    "Insured Full Name or Company Name": "insured", "Insured Country": "insured_country",
    "Reporting Period (End Date)": "reporting_period", "Risk Inception Date": "risk_inception",
    "Risk Expiry Date": "risk_expiry", "Loss Description": "description",
    "Date Claim First Advised/Date Claim Made": "claim_first_advised", "Claim Status": "status",
    "Refer to Underwriters": "refer", "Denial (Y/N)": "denial", "Claimant Name": "claimant",
    "Date Claim Opened": "date_opened", "Date Closed": "date_closed", "Amount Claimed": "amount_claimed",
    "Reserve - Indemnity": "reserves_indemnity", "Reserve - Fees": "reserves_fees",
    "Total Incurred - Indemnity": "total_indemnity", "Total Incurred - Fees": "total_fees",
}
S_TEXTO = {"risk_code", "currency", "certificate", "reference", "insured", "reporting_period",
           "description", "status", "refer", "denial", "claimant"}
S_FECHA = {"risk_inception", "risk_expiry", "claim_first_advised", "date_opened", "date_closed"}
S_NUM = {"amount_claimed", "reserves_indemnity", "reserves_fees", "total_indemnity", "total_fees"}


def _norm(s) -> str:
    return " ".join(str(s).split()) if s is not None else ""


# Variantes de cabecera en las plantillas V5.x -> nombre canónico de HEADERS.
ALIAS = {
    "Claim Status - See Drop Down List": "Claim Status",
    "Denial": "Denial (Y/N)",
}


def _canon(s) -> str:
    h = _norm(s)
    return ALIAS.get(h, h)


def _num(v) -> float:
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v).replace(",", ".")).quantize(Decimal("0.01"), ROUND_HALF_UP)
    except Exception:
        return None


def _fecha(v):
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)) and v > 0:
        try:
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).date()
        except Exception:
            return None
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def _ref_valida(ref: str) -> bool:
    # Un Claim Reference real es un código corto sin espacios; descarta filas de guía/placeholder.
    return bool(ref) and " " not in ref and len(ref) <= 40 and ref.lower() not in ("none", "field", "ref")


def convertir_xls(paths: list[str]) -> dict[str, str]:
    """Convierte cada .xls a .xlsx (Excel COM) en un temp. Devuelve {original: xlsx}. Los .xlsx se dejan tal cual."""
    import win32com.client as win32
    out_dir = os.path.join(tempfile.gettempdir(), "vammos_xlsx")
    os.makedirs(out_dir, exist_ok=True)
    res: dict[str, str] = {}
    xls = [p for p in paths if p.lower().endswith(".xls")]
    res.update({p: p for p in paths if p.lower().endswith(".xlsx")})
    if not xls:
        return res
    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        for i, p in enumerate(xls):
            try:
                wb = xl.Workbooks.Open(p, ReadOnly=True)
                out = os.path.join(out_dir, f"{i:03d}.xlsx")
                wb.SaveAs(out, FileFormat=51)
                wb.Close(False)
                res[p] = out
            except Exception as e:  # noqa: BLE001
                print(f"  [X] conv {os.path.basename(p)}: {str(e)[:70]}")
    finally:
        xl.Quit()
    return res


def extraer(xlsx_path: str):
    """Lee un .xlsx y devuelve (claims, cabecera_ok). cabecera_ok=True si alguna pestaña tenía
    la cabecera del bordereau (aunque no haya filas → mes NIL)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    claims = []
    cabecera_ok = False
    for sh in wb.sheetnames:
        ws = wb[sh]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        hr = ix = None
        for r in range(min(8, len(grid))):
            fila = [_canon(v) for v in grid[r]]
            if "Claim Reference / Number" in fila and "Certificate Reference" in fila:
                hr = r
                ix = {h: c for c, h in enumerate(fila) if h}
                break
        if hr is None:
            continue
        cabecera_ok = True
        col_ref = ix["Claim Reference / Number"]
        for r in range(hr + 1, len(grid)):
            row = grid[r]
            ref = _norm(row[col_ref]) if col_ref < len(row) else ""
            if not _ref_valida(ref):
                continue

            def g(h, row=row):
                c = ix.get(h)
                return row[c] if (c is not None and c < len(row)) else None
            fila = {}
            for h in HEADERS:
                v = g(h)
                if h in H_FECHA:
                    f = _fecha(v)
                    fila[h] = f.isoformat() if f else None
                else:
                    fila[h] = None if v in (None, "") else v
            claims.append({"_sheet": sh, "ref": ref, "g": {h: g(h) for h in set(A_SINIESTRO) | set(HEADERS)}, "fila": fila})
    wb.close()
    return claims, cabecera_ok


def _periodo(claims, fallback):
    for cl in claims:
        f = _fecha(cl["g"].get("Reporting Period (End Date)"))
        if f:
            return f"{f.year:04d}-{f.month:02d}", f.year * 100 + f.month
    return fallback


def _periodo_de_nombre(nombre):
    m = re.search(r"(\d{2})[ _]+(\d{4})", nombre)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}", int(m.group(2)) * 100 + int(m.group(1))
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    ap.add_argument("--binder-id", type=int)
    ap.add_argument("--agreement")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    db = SessionLocal()
    b = db.get(Binder, args.binder_id) if args.binder_id else \
        db.scalar(select(Binder).where(Binder.agreement_number == args.agreement))
    if b is None:
        print("Binder no encontrado.")
        return

    ficheros = []
    for raiz, _, archs in os.walk(args.dir):
        for a in archs:
            if a.lower().endswith((".xls", ".xlsx")) and "resumen" not in a.lower():
                ficheros.append(os.path.join(raiz, a))

    print(f"Convirtiendo {len(ficheros)} ficheros a .xlsx con Excel…")
    conv = convertir_xls(ficheros)

    meses = []
    for path in ficheros:
        xlsx = conv.get(path)
        if not xlsx:
            continue
        try:
            claims, cab_ok = extraer(xlsx)
        except Exception as e:  # noqa: BLE001
            print(f"  [!] {os.path.basename(path)}: error al leer ({str(e)[:70]})")
            continue
        nombre_per = _periodo_de_nombre(os.path.basename(path)) or (None, None)
        if claims:
            per, po = _periodo(claims, nombre_per)
            nil = False
        elif cab_ok:
            # Cabecera presente pero sin filas reales -> mes presentado en BLANCO (NIL).
            per, po = nombre_per
            nil = True
        else:
            print(f"  [!] {os.path.basename(path)}: sin cabecera de bordereau (no se reconoce)")
            continue
        if per is None:
            print(f"  [!] {os.path.basename(path)}: sin periodo")
            continue
        meses.append({"path": path, "per": per, "po": po, "claims": claims, "nil": nil})

    # Dedup: si dos ficheros dieran el mismo periodo, gana el de más claims (más completo).
    by_per: dict[str, dict] = {}
    for m in meses:
        if m["per"] not in by_per or len(m["claims"]) > len(by_per[m["per"]]["claims"]):
            by_per[m["per"]] = m
    meses = sorted(by_per.values(), key=lambda m: m["po"])

    print(f"\n== Claims VAMMOS — binder {b.umr} (DRY-RUN={'NO' if args.apply else 'SÍ'}) ==")
    print(f"Meses con datos: {len(meses)}")
    refs_global = set()
    n_nil = 0
    for m in meses:
        if m["nil"]:
            n_nil += 1
            print(f"  {m['per']}: NIL (presentado en blanco)")
            continue
        refs = sorted({c["ref"] for c in m["claims"]})
        refs_global |= set(refs)
        sheets = sorted({c["_sheet"] for c in m["claims"]})
        print(f"  {m['per']}: {len(refs)} claim(s)  [{', '.join(sheets)}]")
    print(f"Claims distintos (por reference): {len(refs_global)} -> {sorted(refs_global)}")
    print(f"Meses NIL (en blanco): {n_nil}")

    if not args.apply:
        db.close()
        print("\nDRY-RUN: no se ha escrito nada. Repite con --apply.")
        return

    # ── PASO 1: upsert de siniestros por reference (cronológico → último mes gana) ──
    sins = {s.reference: s for s in db.scalars(select(Siniestro).where(Siniestro.binder_id == b.id)).all() if s.reference}
    for m in meses:
        for cl in m["claims"]:
            s = sins.get(cl["ref"])
            if s is None:
                s = Siniestro(binder_id=b.id, reference=cl["ref"])
                db.add(s)
                sins[cl["ref"]] = s
            for h, campo in A_SINIESTRO.items():
                v = cl["g"].get(h)
                if campo in S_FECHA:
                    v = _fecha(v)
                elif campo in S_NUM:
                    v = _dec(v)
                elif campo == "reporting_period":
                    f = _fecha(v)
                    v = f.isoformat() if f else (str(v).strip() or None)
                elif campo in S_TEXTO:
                    v = str(v).strip() if v not in (None, "") else None
                setattr(s, campo, v)
            s.paid_indemnity = _dec(_num(cl["g"].get("Previously Paid - Indemnity")) + _num(cl["g"].get("Paid this month - Indemnity")))
            s.paid_fees = _dec(_num(cl["g"].get("Previously Paid - Fees")) + _num(cl["g"].get("Paid this month - Fees")))
    db.flush()

    # ── PASO 2: presentaciones por mes + bloqueo ──
    total = 0
    for m in meses:
        db.execute(delete(ClaimsPresentacion).where(ClaimsPresentacion.binder_id == b.id, ClaimsPresentacion.periodo == m["per"]))
        if m["nil"]:
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=m["per"], periodo_ord=m["po"], siniestro_id=None,
                paid_indemnity_acum=Decimal("0"), paid_fees_acum=Decimal("0"),
                to_pay_indemnity=Decimal("0"), to_pay_fees=Decimal("0"),
                reserves_indemnity=Decimal("0"), reserves_fees=Decimal("0"), status="Nil",
                fila_json=json.dumps({"nil": True, "report": f"{m['per']} — presentado en blanco"}, ensure_ascii=False),
                fecha_presentacion=None, usuario="histórico-nil"))
            total += 1
            if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == m["per"])):
                db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=m["per"]))
            continue
        for cl in m["claims"]:
            s = sins.get(cl["ref"]); g = cl["g"]
            db.add(ClaimsPresentacion(
                binder_id=b.id, periodo=m["per"], periodo_ord=m["po"], siniestro_id=(s.id if s else None),
                paid_indemnity_acum=Decimal(str(_num(g.get("Previously Paid - Indemnity")) + _num(g.get("Paid this month - Indemnity")))),
                paid_fees_acum=Decimal(str(_num(g.get("Previously Paid - Fees")) + _num(g.get("Paid this month - Fees")))),
                to_pay_indemnity=Decimal(str(_num(g.get("Paid this month - Indemnity")))),
                to_pay_fees=Decimal(str(_num(g.get("Paid this month - Fees")))),
                reserves_indemnity=Decimal(str(_num(g.get("Reserve - Indemnity")))),
                reserves_fees=Decimal(str(_num(g.get("Reserve - Fees")))),
                status=(str(g.get("Claim Status")).strip() if g.get("Claim Status") not in (None, "") else None),
                fila_json=json.dumps(cl["fila"], ensure_ascii=False, default=str),
                fecha_presentacion=None, usuario="histórico",
            ))
            total += 1
        if not db.scalar(select(BdxBloqueo).where(BdxBloqueo.binder_id == b.id, BdxBloqueo.tipo == "claims", BdxBloqueo.periodo == m["per"])):
            db.add(BdxBloqueo(binder_id=b.id, tipo="claims", periodo=m["per"]))
    db.commit()
    print(f"\nAPLICADO: {len(sins)} siniestros (upsert) y {total} filas de presentación en {len(meses)} meses. Meses bloqueados.")
    db.close()


if __name__ == "__main__":
    main()
